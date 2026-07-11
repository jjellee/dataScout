#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
quarterly_earnings.py - 기업 분기별 실적 수집·엑셀 정리

트랙 A (잠정실적): 영업(잠정)실적(공정공시) 공시 HTML을 파싱해 속보 실적 축적
트랙 B (확정실적): OpenDART 재무제표 다중회사 API(fnlttMultiAcnt)로 전 상장사 확정 실적
수주잔고: periodic_report_saver.py가 만든 order_backlog_cache.json을 시트로 정리

사용법:
  python quarterly_earnings.py --backfill-prov   # 잠정실적 백필/증분 수집·파싱
  python quarterly_earnings.py --confirmed       # 확정실적 API 수집 (증분)
  python quarterly_earnings.py --build --upload  # 엑셀 생성 + 텔레그램 업로드
  python quarterly_earnings.py --daily           # 크론용: 증분 수집 + 신규 있으면 빌드·업로드
"""

import os
import re
import io
import sys
import json
import time
import glob
import zipfile
import argparse
import datetime
import logging

import requests
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[logging.StreamHandler(sys.stdout)])
logger = logging.getLogger("quarterly_earnings")

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(PROJECT_DIR, "data_dart")
PROV_CACHE = os.path.join(DATA_DIR, "earnings_prov_cache.json")
CONF_CACHE = os.path.join(DATA_DIR, "earnings_confirmed_cache.json")
BACKLOG_CACHE = os.path.join(DATA_DIR, "order_backlog_cache.json")

sys.path.insert(0, PROJECT_DIR)
# DART 키 로테이션·쿼터 예약 유틸 재사용
from download_important_historical import (  # noqa: E402
    get_api_key, rotate_api_key, QuotaExhausted, wait_until_quota_reset,
    note_request, DART_KEYS, load_env)

load_env()
TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_SUPPLY_DATA_CHAT_ID = os.getenv("TELEGRAM_SUPPLY_DATA_CHAT_ID")
TELEGRAM_TEST_CHAT_ID = os.getenv("TELEGRAM_TEST_CHAT_ID", "-1003843549676")

PROV_PAT = re.compile(r"영업\(잠정\)실적|잠정\)?실적\(공정공시")
EXCLUDE_PAT = re.compile(r"전망|예측")


def load_json(path, default):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def save_json(path, obj):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)
    os.replace(tmp, path)


def api_get(url, params, binary=False):
    """키 로테이션 + 쿼터 카운트 + 일시적 네트워크 오류 재시도가 적용된 API GET."""
    attempts = 0
    net_retries = 0
    while True:
        key = get_api_key()
        params['crtfc_key'] = key
        note_request(key)
        try:
            r = requests.get(url, params=params, timeout=40)
        except requests.exceptions.RequestException as e:
            net_retries += 1
            if net_retries > 5:
                logger.warning(f"network fail after retries: {e}")
                return None
            time.sleep(3 * net_retries)
            continue
        if binary:
            if r.content[:2] == b'PK':
                return r.content
            # 에러 응답 (XML/JSON)
            txt = r.content.decode('utf-8', errors='ignore')
            if '020' in txt or '021' in txt or '사용한도' in txt:
                attempts += 1
                if attempts >= len(DART_KEYS):
                    raise QuotaExhausted()
                rotate_api_key()
                continue
            return None
        d = r.json()
        if d.get("status") in ("020", "021"):
            attempts += 1
            if attempts >= len(DART_KEYS):
                raise QuotaExhausted()
            rotate_api_key()
            continue
        return d


# ---------------------------------------------------------------------------
# 트랙 A: 잠정실적
# ---------------------------------------------------------------------------

def find_prov_disclosures():
    """수집된 disclosures.json 전체에서 잠정실적 공시 목록 추출."""
    out = []
    for f in glob.glob(os.path.join(DATA_DIR, "20*", "disclosures.json")):
        try:
            items = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for x in items:
            nm = x.get("report_nm", "").replace(" ", "")
            if PROV_PAT.search(nm) and not EXCLUDE_PAT.search(nm):
                out.append(x)
    return out


def parse_num(s):
    s = str(s or '').strip().replace(',', '')
    if not s or s in ('-', ''):
        return None
    neg = s.startswith('-') or (s.startswith('(') and s.endswith(')'))
    s = s.strip('()-')
    try:
        return -float(s) if neg else float(s)
    except ValueError:
        return None


UNIT_MULT = {"원": 1, "천원": 1e3, "백만원": 1e6, "억원": 1e8}


def parse_prov_html(xml_text):
    """잠정실적 공시 원문에서 실적 테이블 파싱.

    Returns dict: {basis, quarter, unit, rows: {계정: {당기, 전기, qoq, 전년동기, yoy, 누계, 누계전년, 누계yoy}}}
    """
    soup = BeautifulSoup(xml_text, "html.parser")

    def clean(s):
        return re.sub(r"\s+", " ", s).strip()

    ACCOUNTS = ("매출액", "영업이익", "법인세비용차감전계속사업이익", "당기순이익")
    for table in soup.find_all(["table", "TABLE"]):
        txt = table.get_text()
        if "매출액" not in txt or ("당해실적" not in txt and "당기실적" not in txt):
            continue
        unit = 1e6
        m = re.search(r"단위\s*[:：]\s*([가-힣]+원)", txt)
        if m and m.group(1) in UNIT_MULT:
            unit = UNIT_MULT[m.group(1)]
        quarter = None
        m = re.search(r"(\d{4})\s*년\s*(\d)\s*분기", txt)
        if m:
            quarter = f"{m.group(1)}Q{m.group(2)}"
        rows = {}
        current_acct = None
        for tr in table.find_all(["tr", "TR"]):
            cells = [clean(td.get_text()) for td in tr.find_all(["td", "th", "TD", "TH"])]
            if not cells:
                continue
            first = cells[0].replace(" ", "")
            if first in [a.replace(" ", "") for a in ACCOUNTS]:
                current_acct = first
                body = cells[1:]
            elif first.startswith("누계") and current_acct:
                body = cells
            else:
                continue
            if not body:
                continue
            kind = body[0].replace(" ", "")  # 당해실적 / 누계실적
            vals = body[1:]
            nums = [parse_num(v) for v in vals]
            # [당기, 전기, qoq%, (흑전), 전년동기, yoy%, (흑전)] 순 — 숫자 위치로 매핑
            if kind.startswith("당해") or kind.startswith("당기"):
                r = rows.setdefault(current_acct, {})
                if len(nums) >= 1:
                    r["당기"] = nums[0]
                if len(nums) >= 2:
                    r["전기"] = nums[1]
                if len(nums) >= 3:
                    r["qoq"] = nums[2]
                # 전년동기: 뒤쪽에서 값·증감률 짝 찾기
                tail = [n for n in nums[3:] if n is not None]
                if len(tail) >= 2:
                    r["전년동기"], r["yoy"] = tail[0], tail[1]
                elif len(tail) == 1:
                    r["전년동기"] = tail[0]
            elif kind.startswith("누계"):
                r = rows.setdefault(current_acct, {})
                if len(nums) >= 1:
                    r["누계"] = nums[0]
                tail = [n for n in nums[1:] if n is not None]
                if len(tail) >= 2:
                    r["누계전년"], r["누계yoy"] = tail[0], tail[1]
        if rows:
            return {"quarter": quarter, "unit": unit, "rows": rows}
    return None


def backfill_prov():
    cache = load_json(PROV_CACHE, {})
    targets = find_prov_disclosures()
    todo = [x for x in targets if x["rcept_no"] not in cache]
    logger.info(f"잠정실적: 전체 {len(targets)}건, 신규 {len(todo)}건")
    new_count = 0
    for i, x in enumerate(todo, 1):
        rcept = x["rcept_no"]
        while True:
            try:
                content = api_get("https://opendart.fss.or.kr/api/document.xml",
                                  {"rcept_no": rcept}, binary=True)
                break
            except QuotaExhausted:
                save_json(PROV_CACHE, cache)
                wait_until_quota_reset()
        if not content:
            cache[rcept] = {"error": "no_document", "meta": x}
            continue
        try:
            z = zipfile.ZipFile(io.BytesIO(content))
            xml = z.read(z.namelist()[0]).decode("utf-8", errors="replace")
            parsed = parse_prov_html(xml)
        except Exception as e:
            logger.warning(f"parse fail {rcept}: {e}")
            parsed = None
        basis = "연결" if "연결" in x["report_nm"] else "별도"
        cache[rcept] = {"meta": {k: x.get(k) for k in ("corp_name", "stock_code", "rcept_dt", "report_nm", "corp_code")},
                        "basis": basis, "parsed": parsed}
        new_count += 1
        if i % 100 == 0:
            save_json(PROV_CACHE, cache)
            logger.info(f"  {i}/{len(todo)}")
        time.sleep(0.05)
    save_json(PROV_CACHE, cache)
    ok = sum(1 for v in cache.values() if v.get("parsed"))
    logger.info(f"잠정실적 캐시: {len(cache)}건 (파싱 성공 {ok}) — 신규 {new_count}")
    return new_count


# ---------------------------------------------------------------------------
# 트랙 B: 확정실적 (재무제표 다중회사 API)
# ---------------------------------------------------------------------------

REPRT_CODES = [("11013", "1Q"), ("11012", "반기"), ("11014", "3Q"), ("11011", "연간")]
CONF_ACCOUNTS = {"매출액", "영업이익", "당기순이익"}


def listed_corp_codes():
    """disclosures.json 메타에서 상장사 corp_code↔종목코드·회사명 매핑."""
    seen = {}
    for f in glob.glob(os.path.join(DATA_DIR, "20*", "disclosures.json")):
        try:
            items = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for x in items:
            cc = x.get("corp_code")
            sc = (x.get("stock_code") or "").strip()
            if cc and sc:
                seen[cc] = (x.get("corp_name"), sc)
    return seen


def collect_confirmed():
    cache = load_json(CONF_CACHE, {})
    corps = listed_corp_codes()
    codes = sorted(corps.keys())
    logger.info(f"확정실적: 상장사 {len(codes)}개")
    today = datetime.date.today()
    year_reprts = []
    for y in range(2024, today.year + 1):
        for rc, label in REPRT_CODES:
            year_reprts.append((str(y), rc, label))

    BATCH = 90
    new_rows = 0
    for year, rc, label in year_reprts:
        period_key = f"{year}_{rc}"
        done_key = f"_done_{period_key}"
        # 지난 기간은 완료 표시로 스킵, 최근 2개 기간은 항상 재조회(추가 제출 반영)
        recent = (int(year) >= today.year - (1 if today.month <= 6 else 0))
        if cache.get(done_key) and not recent:
            continue
        for bi in range(0, len(codes), BATCH):
            batch = codes[bi:bi + BATCH]
            while True:
                try:
                    d = api_get("https://opendart.fss.or.kr/api/fnlttMultiAcnt.json",
                                {"corp_code": ",".join(batch), "bsns_year": year, "reprt_code": rc})
                    break
                except QuotaExhausted:
                    save_json(CONF_CACHE, cache)
                    wait_until_quota_reset()
            if not d or d.get("status") != "000":
                continue
            for row in d.get("list", []):
                if row.get("account_nm") not in CONF_ACCOUNTS:
                    continue
                fs = row.get("fs_div")  # CFS 연결 / OFS 별도
                key = f"{row['corp_code']}|{year}|{rc}|{fs}|{row['account_nm']}"
                if key not in cache:
                    new_rows += 1
                cache[key] = {
                    "thstrm": row.get("thstrm_amount"),
                    "add": row.get("thstrm_add_amount"),
                }
            time.sleep(0.1)
        cache[done_key] = True
        save_json(CONF_CACHE, cache)
        logger.info(f"  {year} {label} 수집 완료")
    save_json(CONF_CACHE, cache)
    logger.info(f"확정실적 캐시 항목: {len(cache)} — 신규 {new_rows}")
    return new_rows


# ---------------------------------------------------------------------------
# 엑셀 빌드
# ---------------------------------------------------------------------------

def _fmt_mm(v):
    """원 단위 → 백만원 정수."""
    n = parse_num(v)
    return int(round(n / 1e6)) if n is not None else None


def build_excel(out_path):
    import pandas as pd
    from openpyxl.utils import get_column_letter

    # --- 잠정실적 ---
    prov = load_json(PROV_CACHE, {})
    prov_rows = []
    for rcept, v in prov.items():
        p = v.get("parsed")
        if not p:
            continue
        meta = v["meta"]
        unit_mm = p["unit"] / 1e6  # 백만원 환산 배수
        r = p["rows"]

        def g(acct, field):
            x = r.get(acct, {}).get(field)
            return round(x * unit_mm) if field in ("당기", "전기", "전년동기", "누계", "누계전년") and x is not None \
                else x
        prov_rows.append({
            "접수일자": f"{meta['rcept_dt'][:4]}-{meta['rcept_dt'][4:6]}-{meta['rcept_dt'][6:]}",
            "회사명": meta["corp_name"], "종목코드": meta["stock_code"],
            "기준": v["basis"], "대상분기": p.get("quarter"),
            "매출액": g("매출액", "당기"), "매출QoQ%": g("매출액", "qoq"), "매출YoY%": g("매출액", "yoy"),
            "영업이익": g("영업이익", "당기"), "영업이익QoQ%": g("영업이익", "qoq"), "영업이익YoY%": g("영업이익", "yoy"),
            "순이익": g("당기순이익", "당기"), "순이익YoY%": g("당기순이익", "yoy"),
            "접수번호": rcept,
            "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept}", "공시열람")',
        })
    df_prov = pd.DataFrame(prov_rows)
    if not df_prov.empty:
        df_prov = df_prov.sort_values("접수일자", ascending=False).reset_index(drop=True)

    # --- 확정실적 (연결 우선, 없으면 별도) ---
    conf = load_json(CONF_CACHE, {})
    corps = listed_corp_codes()
    acc = {}
    for key, v in conf.items():
        if key.startswith("_done_"):
            continue
        cc, year, rc, fs, acct = key.split("|")
        label = dict(REPRT_CODES)[rc]
        k = (cc, year, label)
        acc.setdefault(k, {})
        # CFS 우선
        cur = acc[k].get(acct)
        if cur is None or (cur[0] == "OFS" and fs == "CFS"):
            acc[k][acct] = (fs, _fmt_mm(v.get("thstrm")))
    conf_rows = []
    for (cc, year, label), accts in acc.items():
        name, sc = corps.get(cc, (cc, ""))
        row = {"회사명": name, "종목코드": sc, "연도": year, "보고서": label,
               "기준": "연결" if any(x[0] == "CFS" for x in accts.values()) else "별도"}
        for a in ("매출액", "영업이익", "당기순이익"):
            row[f"{a}(누적)"] = accts.get(a, (None, None))[1]
        conf_rows.append(row)
    df_conf = pd.DataFrame(conf_rows)
    if not df_conf.empty:
        order = {"1Q": 1, "반기": 2, "3Q": 3, "연간": 4}
        df_conf["_o"] = df_conf["보고서"].map(order)
        df_conf = df_conf.sort_values(["회사명", "연도", "_o"]).drop(columns="_o")
        # 분기 단독값 파생: 2Q=반기-1Q, 3Q표시는 누적이므로 3Q단독=3Q-반기, 4Q=연간-3Q
        for a in ("매출액", "영업이익", "당기순이익"):
            col = f"{a}(누적)"
            qcol = f"{a}(분기)"
            vals = []
            prev = {}
            for _, row in df_conf.iterrows():
                k = (row["회사명"], row["연도"])
                cum = row[col]
                p = prev.get(k)
                if row["보고서"] == "1Q" or cum is None:
                    q = cum
                else:
                    q = cum - p if p is not None else None
                vals.append(q)
                if cum is not None:
                    prev[k] = cum
            df_conf[qcol] = vals

    # --- 수주잔고 ---
    backlog = load_json(BACKLOG_CACHE, {})
    bl_rows = []
    for rcept, v in backlog.items():
        for row in v.get("rows", []):
            bl_rows.append({
                "회사명": v["corp_name"], "종목코드": v["stock_code"],
                "보고서": v["label"], "접수일자": v["rcept_dt"],
                "품목": row.get("품목"), "수주총액": row.get("수주총액"),
                "기납품액": row.get("기납품액"), "수주잔고": row.get("수주잔고"),
                "단위": v.get("unit", ""),
                "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept}", "공시열람")',
            })
    df_bl = pd.DataFrame(bl_rows)
    if not df_bl.empty:
        df_bl = df_bl.sort_values(["회사명", "보고서"], ascending=[True, False]).reset_index(drop=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        (df_prov if not df_prov.empty else pd.DataFrame([{"안내": "데이터 없음"}])).to_excel(
            writer, sheet_name="잠정실적(속보)", index=False)
        (df_conf if not df_conf.empty else pd.DataFrame([{"안내": "데이터 없음"}])).to_excel(
            writer, sheet_name="확정실적(분기)", index=False)
        (df_bl if not df_bl.empty else pd.DataFrame([{"안내": "periodic_report_saver 수집 대기"}])).to_excel(
            writer, sheet_name="수주잔고", index=False)
        for ws in writer.book:
            ws.freeze_panes = "A2"
            for ci, col in enumerate(ws[1], 1):
                width = max(10, min(30, len(str(col.value or "")) + 6))
                ws.column_dimensions[get_column_letter(ci)].width = width
            ws.auto_filter.ref = ws.dimensions
    logger.info(f"Excel saved: {out_path} (잠정 {len(df_prov)}, 확정 {len(df_conf)}, 수주 {len(df_bl)})")
    return len(df_prov) + len(df_conf) + len(df_bl)


def send_document(file_path, caption, test=False):
    chat = TELEGRAM_TEST_CHAT_ID if test else TELEGRAM_SUPPLY_DATA_CHAT_ID
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT4_TOKEN}/sendDocument"
    with open(file_path, "rb") as f:
        r = requests.post(url, data={"chat_id": chat, "caption": caption},
                          files={"document": f}, timeout=120)
    ok = r.json().get("ok")
    logger.info(f"Telegram upload {'OK' if ok else 'FAILED: ' + r.text[:200]}")
    return ok


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--backfill-prov", action="store_true")
    ap.add_argument("--confirmed", action="store_true")
    ap.add_argument("--build", action="store_true")
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--daily", action="store_true", help="증분 수집 + 신규 있으면 빌드·업로드")
    ap.add_argument("--test", action="store_true")
    args = ap.parse_args()

    out_path = os.path.join(DATA_DIR, "quarterly_earnings.xlsx")
    new = 0
    if args.backfill_prov or args.daily:
        new += backfill_prov()
    if args.confirmed or args.daily:
        new += collect_confirmed()
    if args.build or (args.daily and new > 0):
        total = build_excel(out_path)
        if args.upload or args.daily:
            prov = load_json(PROV_CACHE, {})
            caption = (f"📈 분기 실적 정리 ({datetime.date.today()})\n"
                       f"잠정실적(속보) + 확정실적(재무제표 API) + 수주잔고\n"
                       f"총 {total:,}행")
            send_document(out_path, caption, test=args.test)
    elif args.daily:
        logger.info("신규 데이터 없음 — 빌드 스킵")


if __name__ == "__main__":
    main()

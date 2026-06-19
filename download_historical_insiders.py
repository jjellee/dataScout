import os
import pandas as pd
import zipfile
import requests
import time
import urllib.parse
from bs4 import BeautifulSoup

def download_q1_dera():
    """Parses pre-extracted Q1 2026 DERA dataset files and returns a DataFrame."""
    dir_path = "/home/inhyuk/projects/dataScout/2026q1_extracted"
    
    if not os.path.exists(dir_path):
        print("Q1 extracted directory not found. Please ensure inspect_q1.py was run.")
        return pd.DataFrame()
        
    print("Loading Q1 DERA data from TSVs...")
    sub_df = pd.read_csv(os.path.join(dir_path, "SUBMISSION.tsv"), sep="\t", 
                         usecols=['ACCESSION_NUMBER', 'FILING_DATE', 'ISSUERNAME', 'ISSUERTRADINGSYMBOL', 'ISSUERCIK'])
    owner_df = pd.read_csv(os.path.join(dir_path, "REPORTINGOWNER.tsv"), sep="\t", 
                           usecols=['ACCESSION_NUMBER', 'RPTOWNERNAME', 'RPTOWNER_RELATIONSHIP', 'RPTOWNER_TITLE'])
    trans_df = pd.read_csv(os.path.join(dir_path, "NONDERIV_TRANS.tsv"), sep="\t", 
                           usecols=['ACCESSION_NUMBER', 'TRANS_SHARES', 'TRANS_PRICEPERSHARE', 'TRANS_ACQUIRED_DISP_CD', 'TRANS_FORM_TYPE'])
    
    print("Processing and joining Q1 DERA datasets...")
    trans_df = trans_df[trans_df['TRANS_FORM_TYPE'] == 4]
    df = trans_df.merge(sub_df, on='ACCESSION_NUMBER', how='inner')
    df = df.merge(owner_df, on='ACCESSION_NUMBER', how='inner')
    
    df = df[df['TRANS_SHARES'].notna() & df['TRANS_PRICEPERSHARE'].notna()]
    df = df[(df['TRANS_SHARES'] > 0) & (df['TRANS_PRICEPERSHARE'] > 0)]
    
    def map_role(row):
        rel = str(row['RPTOWNER_RELATIONSHIP'])
        title = str(row['RPTOWNER_TITLE']) if pd.notna(row['RPTOWNER_TITLE']) else ''
        if 'Director' in rel:
            if 'Officer' in rel and title:
                return title
            return 'Director'
        elif 'Officer' in rel:
            return title if title else 'Officer'
        elif 'TenPercentOwner' in rel:
            return '10% Owner'
        return title if title else 'Insider'
        
    df['role'] = df.apply(map_role, axis=1)
    df['filing_date'] = pd.to_datetime(df['FILING_DATE'], format='%d-%b-%Y').dt.strftime('%Y-%m-%d')
    df['ticker'] = df['ISSUERTRADINGSYMBOL'].fillna('Unknown')
    df['company'] = df['ISSUERNAME'].fillna('Unknown')
    df['insider'] = df['RPTOWNERNAME'].fillna('Unknown')
    df['type'] = df['TRANS_ACQUIRED_DISP_CD'].map({'A': 'BUY', 'D': 'SELL'})
    df['shares'] = df['TRANS_SHARES'].astype(int)
    df['price'] = df['TRANS_PRICEPERSHARE'].astype(float)
    df['value'] = df['shares'] * df['price']
    
    def make_filing_url(row):
        cik = str(row['ISSUERCIK']).zfill(10)
        acc = str(row['ACCESSION_NUMBER'])
        acc_no_hyphen = acc.replace('-', '')
        return f"https://www.sec.gov/Archives/edgar/data/{cik}/{acc_no_hyphen}/"
        
    df['filing_url'] = df.apply(make_filing_url, axis=1)
    
    final_cols = ['filing_date', 'ticker', 'company', 'insider', 'role', 'type', 'shares', 'price', 'value', 'filing_url']
    df_final = df[final_cols]
    print(f"Processed {len(df_final)} transactions from Q1 2026.")
    return df_final

def scrape_openinsider_range(start_date, end_date):
    """Scrapes a specific date range from OpenInsider."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    fdr_val = f"{start_date} - {end_date}"
    fdr_encoded = urllib.parse.quote(fdr_val)
    url = f"http://openinsider.com/screener?s=&o=&pl=&ph=&ll=&lh=&fd=-1&fdr={fdr_encoded}&td=0&tdr=&fdoy=&rp=&rel=&type%5B%5D=p&type%5B%5D=s&cl=&co=&isc=&ic=&hh=&cnt=5000"
    
    print(f"Scraping OpenInsider for filing date range: {fdr_val}...")
    try:
        resp = requests.get(url, headers=headers, timeout=45)
        if resp.status_code != 200:
            print(f"Failed to fetch {fdr_val}: Status {resp.status_code}")
            return []
            
        soup = BeautifulSoup(resp.content, "html.parser")
        table = soup.find("table", {"class": "tinytable"})
        if not table:
            print(f"No transactions found for range {fdr_val}")
            return []
            
        rows = table.find_all("tr")
        transactions = []
        for r in rows[1:]:
            tds = r.find_all("td")
            if len(tds) < 13:
                continue
                
            filing_date_tag = tds[1].find("a")
            if not filing_date_tag:
                continue
            filing_date_str = filing_date_tag.get_text().strip().split()[0]  # YYYY-MM-DD
            filing_url = filing_date_tag.get("href", "").strip()
            if filing_url.startswith("http://"):
                filing_url = filing_url.replace("http://", "https://")
                
            trade_date = tds[2].get_text().strip()
            ticker = tds[3].get_text().strip()
            company = tds[4].get_text().strip()
            insider = tds[5].get_text().strip()
            role = tds[6].get_text().strip()
            
            trade_type_str = tds[7].get_text().strip()
            if "Purchase" in trade_type_str or trade_type_str.startswith("P -"):
                trade_type = "BUY"
            elif "Sale" in trade_type_str or trade_type_str.startswith("S -"):
                trade_type = "SELL"
            else:
                continue
                
            price_str = tds[8].get_text().strip().replace("$", "").replace(",", "")
            price = float(price_str) if price_str else 0.0
            
            qty_str = tds[9].get_text().strip().replace("+", "").replace(",", "")
            shares = int(qty_str) if qty_str else 0
            shares = abs(shares)
            
            value = shares * price
            if value == 0:
                continue
                
            transactions.append({
                'filing_date': filing_date_str,
                'ticker': ticker,
                'company': company,
                'insider': insider,
                'role': role,
                'type': trade_type,
                'shares': shares,
                'price': price,
                'value': value,
                'filing_url': filing_url
            })
        print(f"Successfully parsed {len(transactions)} transactions.")
        return transactions
    except Exception as e:
        print(f"Error scraping range {fdr_val}: {e}")
        return []

def get_q2_openinsider():
    """Scrapes Q2 2026 from OpenInsider in 10-day chunks."""
    date_ranges = [
        ("04/01/2026", "04/10/2026"),
        ("04/11/2026", "04/20/2026"),
        ("04/21/2026", "04/30/2026"),
        ("05/01/2026", "05/10/2026"),
        ("05/11/2026", "05/20/2026"),
        ("05/21/2026", "05/31/2026"),
        ("06/01/2026", "06/10/2026"),
        ("06/11/2026", "06/19/2026")
    ]
    
    all_transactions = []
    for start, end in date_ranges:
        txs = scrape_openinsider_range(start, end)
        all_transactions.extend(txs)
        time.sleep(2.0)  # Polite sleep to avoid rate limiting
        
    return pd.DataFrame(all_transactions)

def main():
    print("Starting Historical Insider Transactions Collector (2026)...")
    
    # 1. Parse Q1 DERA
    df_q1 = download_q1_dera()
    
    # 2. Scrape Q2 OpenInsider
    df_q2 = get_q2_openinsider()
    
    # 3. Combine and Deduplicate
    excel_path = "/home/inhyuk/projects/dataScout/data_us/us_insider_transactions.xlsx"
    
    df_old = pd.DataFrame()
    if os.path.exists(excel_path):
        try:
            df_old = pd.read_excel(excel_path)
            print(f"Loaded {len(df_old)} existing transactions from Excel.")
        except Exception as e:
            print(f"Error reading existing Excel: {e}")
            
    dfs_to_concat = []
    if not df_old.empty:
        dfs_to_concat.append(df_old)
    if not df_q1.empty:
        dfs_to_concat.append(df_q1)
    if not df_q2.empty:
        dfs_to_concat.append(df_q2)
        
    if not dfs_to_concat:
        print("No transactions to save.")
        return
        
    df_combined = pd.concat(dfs_to_concat, ignore_index=True)
    
    # Clean data types
    df_combined['shares'] = df_combined['shares'].astype(int)
    df_combined['price'] = df_combined['price'].astype(float)
    df_combined['value'] = df_combined['value'].astype(float)
    
    df_combined['filing_url'] = df_combined['filing_url'].fillna('')
    df_combined['ticker'] = df_combined['ticker'].fillna('')
    df_combined['insider'] = df_combined['insider'].fillna('')
    df_combined['type'] = df_combined['type'].fillna('')
    
    # Deduplicate based on unique transaction key
    print(f"Total rows before deduplication: {len(df_combined)}")
    df_combined = df_combined.sort_values(by='filing_date', ascending=False)
    df_combined = df_combined.drop_duplicates(
        subset=['filing_url', 'ticker', 'insider', 'shares', 'price', 'type'],
        keep='first'
    )
    df_combined = df_combined.sort_values(by='filing_date', ascending=False)
    print(f"Total rows after deduplication: {len(df_combined)}")
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    # Save to Excel with formatted column widths and colored type cells
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_combined.to_excel(writer, index=False, sheet_name='Transactions')
            worksheet = writer.sheets['Transactions']
            
            # Define fills
            buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
            sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # soft red
            meaningful_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # soft yellow/gold
            
            # Color Column 6 (type) based on BUY/SELL and highlight meaningful values in Column 9 (value)
            for row_idx in range(2, worksheet.max_row + 1):
                type_cell = worksheet.cell(row=row_idx, column=6)
                shares_cell = worksheet.cell(row=row_idx, column=7)
                price_cell = worksheet.cell(row=row_idx, column=8)
                value_cell = worksheet.cell(row=row_idx, column=9)
                
                # Apply 3-digit comma formatting and currency units ($)
                shares_cell.number_format = '#,##0'
                price_cell.number_format = '$#,##0.00'
                value_cell.number_format = '$#,##0'
                
                if type_cell.value == "BUY":
                    type_cell.fill = buy_fill
                    try:
                        val_num = float(value_cell.value)
                        if val_num >= 30000:
                            value_cell.fill = meaningful_fill
                    except (ValueError, TypeError):
                        pass
                elif type_cell.value == "SELL":
                    type_cell.fill = sell_fill
                    try:
                        val_num = float(value_cell.value)
                        if val_num >= 200000:
                            value_cell.fill = meaningful_fill
                    except (ValueError, TypeError):
                        pass
            
            # Adjust column widths
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                width = min(max(max_len + 3, 10), 50)
                if col_idx in [3, 4, 5]: # company, insider, role
                    width = width / 2
                worksheet.column_dimensions[col_letter].width = width
                
            # Apply AutoFilter
            worksheet.auto_filter.ref = worksheet.dimensions
            
        print(f"Successfully saved cumulative insider transactions to {excel_path}")
    except Exception as e:
        print(f"Failed to save cumulative Excel file: {e}")

if __name__ == "__main__":
    main()

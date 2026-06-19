#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import datetime
import requests
import time
import pandas as pd
import argparse
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Setup logging
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Custom env loader
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    val_str = val.strip().strip("'").strip('"')
                    os.environ[key.strip()] = val_str

load_env()
DART_API_KEY = os.getenv("DART_API_KEY")
TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_JJANG_GU_CHAT_ID = os.getenv("TELEGRAM_JJANG_GU_CHAT_ID") or "-1003877753638"

def fetch_daily_disclosures_list(target_date):
    """Fetches list of all disclosures for a specific date from DART."""
    url = "https://opendart.fss.or.kr/api/list.json"
    page_no = 1
    page_count = 100
    all_reports = []

    if not DART_API_KEY:
        logger.error("DART_API_KEY is missing in environment variables.")
        return []

    while True:
        params = {
            'crtfc_key': DART_API_KEY,
            'bgn_de': target_date,
            'end_de': target_date,
            'page_no': page_no,
            'page_count': page_count
        }
        
        try:
            response = requests.get(url, params=params, timeout=15)
            if response.status_code != 200:
                break
                
            data = response.json()
            status = data.get("status")
            
            if status == "013": # No results
                break
            elif status != "000":
                logger.error(f"DART API Error ({status}) for date {target_date}: {data.get('message')}")
                break
                
            reports = data.get("list", [])
            all_reports.extend(reports)
            
            total_page = int(data.get("total_page", 1))
            if page_no >= total_page:
                break
            page_no += 1
            
        except Exception as e:
            logger.error(f"Request failed for date {target_date}: {e}")
            break
            
    return all_reports

def fetch_elestock_details(corp_code):
    """Fetches executive and major shareholder stock status for a company."""
    url = "https://opendart.fss.or.kr/api/elestock.json"
    params = {
        'crtfc_key': DART_API_KEY,
        'corp_code': corp_code
    }
    
    try:
        response = requests.get(url, params=params, timeout=15)
        if response.status_code != 200:
            return []
            
        data = response.json()
        status = data.get("status")
        if status == "000":
            return data.get("list", [])
        return []
    except Exception as e:
        logger.error(f"Failed to fetch elestock details for {corp_code}: {e}")
        return []

def fetch_majorstock_details(corp_code):
    """Fetches 5% block holdings details for a company."""
    url = "https://opendart.fss.or.kr/api/majorstock.json"
    params = {
        'crtfc_key': DART_API_KEY,
        'corp_code': corp_code
    }
    
    try:
        response = requests.get(url, params=params, timeout=15)
        if response.status_code != 200:
            return []
            
        data = response.json()
        status = data.get("status")
        if status == "000":
            return data.get("list", [])
        return []
    except Exception as e:
        logger.error(f"Failed to fetch majorstock details for {corp_code}: {e}")
        return []

def get_kr_close_price(ticker, date_str):
    """Tries to load close price from daily investor trend CSV, falls back to pykrx."""
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    formatted_date = date_str.replace("-", "")
    
    # Try local CSV first
    csv_path = os.path.join(workspace_dir, "data_kr", formatted_date, "all_stocks_investor_trend.csv")
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path, dtype={'티커': str})
            df = df.set_index('티커')
            ticker_str = str(ticker).zfill(6)
            if ticker_str in df.index:
                return float(df.loc[ticker_str]['종가'])
        except Exception:
            pass
            
    # Fallback to pykrx
    try:
        from pykrx import stock
        df = stock.get_market_close_by_ticker(formatted_date)
        ticker_str = str(ticker).zfill(6)
        if ticker_str in df.index:
            return float(df.loc[ticker_str]['종가'])
    except Exception:
        pass
        
    return 0.0

def send_telegram_document(token, chat_id, file_path, caption=None):
    """Sends a document file to Telegram."""
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    data = {"chat_id": chat_id}
    if caption:
        data["caption"] = caption
        data["parse_mode"] = "Markdown"
        
    try:
        with open(file_path, "rb") as f:
            files = {"document": f}
            resp = requests.post(url, data=data, files=files, timeout=300)
            return resp.json()
    except Exception as e:
        logger.error(f"Failed to send telegram document: {e}")
        return None

def process_date_range(start_date, end_date):
    """Collects insider and block disclosures for a date range and queries details."""
    start_dt = datetime.datetime.strptime(start_date.replace("-", ""), "%Y%m%d")
    end_dt = datetime.datetime.strptime(end_date.replace("-", ""), "%Y%m%d")
    
    current_dt = start_dt
    all_disclosures = []
    
    logger.info(f"Step 1: Fetching disclosures from {start_date} to {end_date}...")
    while current_dt <= end_dt:
        date_str = current_dt.strftime("%Y%m%d")
        reports = fetch_daily_disclosures_list(date_str)
        
        # Filter target reports:
        # A. Executive & Major Shareholder (임원ㆍ주요주주)
        # B. Largest Shareholder change reports (최대주주등소유주식변동신고서)
        # C. 5% Block Holdings (주식등의대량보유상황보고서)
        filtered = [
            r for r in reports 
            if ("임원ㆍ주요주주" in str(r.get("report_nm", ""))) or
               ("최대주주등소유주식변동" in str(r.get("report_nm", ""))) or
               ("대량보유상황보고서" in str(r.get("report_nm", "")))
        ]
        
        if filtered:
            logger.info(f"  Date {date_str}: Found {len(filtered)} target disclosures.")
            all_disclosures.extend(filtered)
            
        current_dt += datetime.timedelta(days=1)
        time.sleep(0.1)
        
    if not all_disclosures:
        logger.info("No target disclosures found in this date range.")
        return [], []
        
    # Group disclosures
    elestock_disclosures = [r for r in all_disclosures if "임원ㆍ주요주주" in r['report_nm']]
    largest_disclosures = [r for r in all_disclosures if "최대주주등소유주식변동" in r['report_nm']]
    block_disclosures = [r for r in all_disclosures if "대량보유상황보고서" in r['report_nm']]
    
    # Process unique corp codes for elestock (Sheet 1)
    unique_elestock_corps = list(set([r['corp_code'] for r in elestock_disclosures if r.get('corp_code')]))
    logger.info(f"Step 2A: Fetching elestock details for {len(unique_elestock_corps)} unique companies...")
    elestock_cache = {}
    for idx, corp_code in enumerate(unique_elestock_corps):
        if idx % 20 == 0 and idx > 0:
            logger.info(f"  Processed {idx}/{len(unique_elestock_corps)} companies...")
        details = fetch_elestock_details(corp_code)
        if details:
            elestock_cache[corp_code] = details
        time.sleep(0.15)
        
    # Process unique corp codes for majorstock (Sheet 2)
    unique_block_corps = list(set([r['corp_code'] for r in block_disclosures if r.get('corp_code')]))
    logger.info(f"Step 2B: Fetching majorstock (5% block) details for {len(unique_block_corps)} unique companies...")
    majorstock_cache = {}
    for idx, corp_code in enumerate(unique_block_corps):
        if idx % 20 == 0 and idx > 0:
            logger.info(f"  Processed {idx}/{len(unique_block_corps)} companies...")
        details = fetch_majorstock_details(corp_code)
        if details:
            majorstock_cache[corp_code] = details
        time.sleep(0.15)
        
    # Match and parse
    logger.info("Step 3: Matching disclosures with details and prices...")
    
    # Sheet 1: InsiderTrades
    insider_trades = []
    
    # Add elestock matches
    for r in elestock_disclosures:
        rcept_no = str(r['rcept_no'])
        corp_code = str(r['corp_code'])
        stock_code = str(r.get('stock_code', '')).zfill(6)
        rcept_dt = str(r['rcept_dt'])
        
        details = elestock_cache.get(corp_code, [])
        match = next((d for d in details if str(d.get('rcept_no')) == rcept_no), None)
        
        if match:
            try:
                change_shares_str = str(match.get('sp_stock_lmp_irds_cnt', '0')).replace(',', '')
                change_shares = int(change_shares_str) if change_shares_str and change_shares_str != '-' else 0
                if change_shares == 0:
                    continue
                    
                total_shares_str = str(match.get('sp_stock_lmp_cnt', '0')).replace(',', '')
                total_shares = int(total_shares_str) if total_shares_str and total_shares_str != '-' else 0
                
                change_rate_str = str(match.get('sp_stock_lmp_irds_rate', '0')).replace(',', '')
                change_rate = float(change_rate_str) if change_rate_str and change_rate_str != '-' else 0.0
                
                total_rate_str = str(match.get('sp_stock_lmp_rate', '0')).replace(',', '')
                total_rate = float(total_rate_str) if total_rate_str and total_rate_str != '-' else 0.0
                
                close_price = get_kr_close_price(stock_code, rcept_dt)
                approx_value = abs(change_shares) * close_price
                
                role_parts = []
                if match.get('isu_exctv_rgist_at') and match['isu_exctv_rgist_at'] != '-':
                    role_parts.append(match['isu_exctv_rgist_at'])
                if match.get('isu_exctv_ofcps') and match['isu_exctv_ofcps'] != '-':
                    role_parts.append(match['isu_exctv_ofcps'])
                if match.get('isu_main_shrholdr') and match['isu_main_shrholdr'] != '-':
                    role_parts.append(match['isu_main_shrholdr'])
                role = " / ".join(role_parts) if role_parts else "임원/주주"
                
                insider_trades.append({
                    'filing_date': f"{rcept_dt[:4]}-{rcept_dt[4:6]}-{rcept_dt[6:]}",
                    'ticker': stock_code,
                    'company': r['corp_name'],
                    'insider': match['repror'],
                    'role': role,
                    'type': 'BUY' if change_shares > 0 else 'SELL',
                    'shares': abs(change_shares),
                    'price': close_price,
                    'value': approx_value,
                    'total_shares': total_shares,
                    'total_rate': total_rate,
                    'filing_url': f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}"
                })
            except Exception as e:
                logger.error(f"Error parsing elestock match {rcept_no}: {e}")
                
    # Add largest shareholder reports directly (no detailed JSON, but metadata with direct link)
    for r in largest_disclosures:
        try:
            rcept_no = str(r['rcept_no'])
            stock_code = str(r.get('stock_code', '')).zfill(6)
            rcept_dt = str(r['rcept_dt'])
            
            insider_trades.append({
                'filing_date': f"{rcept_dt[:4]}-{rcept_dt[4:6]}-{rcept_dt[6:]}",
                'ticker': stock_code,
                'company': r['corp_name'],
                'insider': r['flr_nm'],
                'role': "최대주주등 (공시)",
                'type': "-",
                'shares': None,
                'price': None,
                'value': None,
                'total_shares': None,
                'total_rate': None,
                'filing_url': f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}"
            })
        except Exception as e:
            logger.error(f"Error parsing largest shareholder report {rcept_no}: {e}")

    # Sheet 2: BlockHoldings (5% Rule)
    block_holdings = []
    
    for r in block_disclosures:
        rcept_no = str(r['rcept_no'])
        corp_code = str(r['corp_code'])
        stock_code = str(r.get('stock_code', '')).zfill(6)
        rcept_dt = str(r['rcept_dt'])
        
        details = majorstock_cache.get(corp_code, [])
        match = next((d for d in details if str(d.get('rcept_no')) == rcept_no), None)
        
        if match:
            try:
                change_shares_str = str(match.get('stkqy_irds', '0')).replace(',', '')
                change_shares = int(change_shares_str) if change_shares_str and change_shares_str != '-' else 0
                if change_shares == 0:
                    continue
                    
                total_shares_str = str(match.get('stkqy', '0')).replace(',', '')
                total_shares = int(total_shares_str) if total_shares_str and total_shares_str != '-' else 0
                
                change_rate_str = str(match.get('stkrt_irds', '0')).replace(',', '')
                change_rate = float(change_rate_str) if change_rate_str and change_rate_str != '-' else 0.0
                
                total_rate_str = str(match.get('stkrt', '0')).replace(',', '')
                total_rate = float(total_rate_str) if total_rate_str and total_rate_str != '-' else 0.0
                
                close_price = get_kr_close_price(stock_code, rcept_dt)
                approx_value = abs(change_shares) * close_price
                
                block_holdings.append({
                    'filing_date': f"{rcept_dt[:4]}-{rcept_dt[4:6]}-{rcept_dt[6:]}",
                    'ticker': stock_code,
                    'company': r['corp_name'],
                    'insider': match['repror'],
                    'role': f"5% 보유자 ({match.get('report_tp', '일반')})",
                    'type': 'BUY' if change_shares > 0 else 'SELL',
                    'shares': abs(change_shares),
                    'price': close_price,
                    'value': approx_value,
                    'total_shares': total_shares,
                    'total_rate': total_rate,
                    'report_reason': match.get('report_resn', '-'),
                    'filing_url': f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}"
                })
            except Exception as e:
                logger.error(f"Error parsing majorstock match {rcept_no}: {e}")
                
    return insider_trades, block_holdings

def save_to_excel(insider_trades, block_holdings):
    """Saves compiled transactions to Excel sheet cumulatively with premium formatting."""
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(workspace_dir, "data_kr")
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, "kr_insider_transactions.xlsx")
    
    # Load old sheets if exist
    old_insiders = pd.DataFrame()
    old_blocks = pd.DataFrame()
    
    if os.path.exists(excel_path):
        try:
            with pd.ExcelFile(excel_path) as xls:
                if 'InsiderTrades' in xls.sheet_names:
                    old_insiders = pd.read_excel(xls, 'InsiderTrades')
                if 'BlockHoldings' in xls.sheet_names:
                    old_blocks = pd.read_excel(xls, 'BlockHoldings')
        except Exception as e:
            logger.error(f"Failed to read existing Excel sheets: {e}")
            
    # Merge Insider Trades
    df_new_insiders = pd.DataFrame(insider_trades)
    if not df_new_insiders.empty or not old_insiders.empty:
        df_combined_insiders = pd.concat([old_insiders, df_new_insiders], ignore_index=True)
        dup_cols = ['ticker', 'insider', 'shares', 'type']
        if 'filing_url' in df_combined_insiders.columns:
            dup_cols.append('filing_url')
        df_combined_insiders = df_combined_insiders.drop_duplicates(subset=dup_cols, keep='first')
        if 'filing_date' in df_combined_insiders.columns:
            df_combined_insiders = df_combined_insiders.sort_values(by='filing_date', ascending=False)
    else:
        df_combined_insiders = pd.DataFrame()
        
    # Merge Block Holdings
    df_new_blocks = pd.DataFrame(block_holdings)
    if not df_new_blocks.empty or not old_blocks.empty:
        df_combined_blocks = pd.concat([old_blocks, df_new_blocks], ignore_index=True)
        dup_cols_block = ['ticker', 'insider', 'shares', 'type']
        if 'filing_url' in df_combined_blocks.columns:
            dup_cols_block.append('filing_url')
        df_combined_blocks = df_combined_blocks.drop_duplicates(subset=dup_cols_block, keep='first')
        if 'filing_date' in df_combined_blocks.columns:
            df_combined_blocks = df_combined_blocks.sort_values(by='filing_date', ascending=False)
    else:
        df_combined_blocks = pd.DataFrame()
        
    # Determine if file is updated (new transactions added)
    is_updated = False
    if len(df_combined_insiders) > len(old_insiders) or len(df_combined_blocks) > len(old_blocks):
        is_updated = True
        
    if df_combined_insiders.empty and df_combined_blocks.empty:
        logger.info("No data to save.")
        return False
        
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write InsiderTrades
            if not df_combined_insiders.empty:
                df_combined_insiders.to_excel(writer, index=False, sheet_name='InsiderTrades')
                ws_insiders = writer.sheets['InsiderTrades']
                format_sheet(ws_insiders, is_block=False)
            
            # Write BlockHoldings
            if not df_combined_blocks.empty:
                df_combined_blocks.to_excel(writer, index=False, sheet_name='BlockHoldings')
                ws_blocks = writer.sheets['BlockHoldings']
                format_sheet(ws_blocks, is_block=True)
                
        logger.info(f"Successfully saved KR insider transactions Excel to {excel_path}")
        return is_updated
    except Exception as e:
        logger.error(f"Failed to save cumulative Excel file: {e}")
        return False

def format_sheet(worksheet, is_block=False):
    """Applies styles, number formats, auto filters, and width settings to a sheet."""
    # Define fills
    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
    sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # soft red
    meaningful_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # soft yellow/gold
    
    # 50M KRW for insiders, 500M KRW for block holdings
    meaningful_buy_limit = 500000000 if is_block else 50000000
    # 200M KRW for insiders, 2B KRW for block holdings
    meaningful_sell_limit = 2000000000 if is_block else 200000000
    
    for row_idx in range(2, worksheet.max_row + 1):
        type_cell = worksheet.cell(row=row_idx, column=6)
        shares_cell = worksheet.cell(row=row_idx, column=7)
        price_cell = worksheet.cell(row=row_idx, column=8)
        value_cell = worksheet.cell(row=row_idx, column=9)
        total_shares_cell = worksheet.cell(row=row_idx, column=10)
        total_rate_cell = worksheet.cell(row=row_idx, column=11)
        
        # Apply 3-digit comma and currency/unit formatting
        shares_cell.number_format = '#,##0'
        price_cell.number_format = '₩#,##0'
        value_cell.number_format = '₩#,##0'
        total_shares_cell.number_format = '#,##0'
        total_rate_cell.number_format = '0.00"%"'
        
        # Skip if empty or non-numeric (e.g., largest shareholder report row)
        if type_cell.value == "-" or type_cell.value is None:
            continue
            
        # Color code type & value
        if type_cell.value == "BUY":
            type_cell.fill = buy_fill
            try:
                val_num = float(value_cell.value)
                if val_num >= meaningful_buy_limit:
                    value_cell.fill = meaningful_fill
            except (ValueError, TypeError):
                pass
        elif type_cell.value == "SELL":
            type_cell.fill = sell_fill
            try:
                val_num = float(value_cell.value)
                if val_num >= meaningful_sell_limit:
                    value_cell.fill = meaningful_fill
            except (ValueError, TypeError):
                pass
                
    # Adjust column widths
    for col_idx, col in enumerate(worksheet.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        width = min(max(max_len + 3, 10), 50)
        if col_idx in [3, 4, 5]: # company, insider, role
            width = width / 2
        worksheet.column_dimensions[col_letter].width = width
        
    # Apply AutoFilter
    worksheet.auto_filter.ref = worksheet.dimensions

def main():
    parser = argparse.ArgumentParser(description="DART Korean Insider Transactions Collector")
    parser.add_argument("--days", type=int, default=0, help="Number of days to collect (default: 0 = today only)")
    parser.add_argument("--start", type=str, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", type=str, help="End date (YYYY-MM-DD)")
    args = parser.parse_args()
    
    if args.start and args.end:
        start_date = args.start
        end_date = args.end
    else:
        # Determine range by days
        days_to_collect = args.days if args.days > 0 else 1
        end_dt = datetime.datetime.now()
        start_dt = end_dt - datetime.timedelta(days=days_to_collect - 1)
        
        start_date = start_dt.strftime("%Y-%m-%d")
        end_date = end_dt.strftime("%Y-%m-%d")
        
    logger.info(f"Starting KR Insider/Block Collector for range: {start_date} to {end_date}...")
    
    insider_trades, block_holdings = process_date_range(start_date, end_date)
    is_updated = save_to_excel(insider_trades, block_holdings)
    
    if is_updated:
        workspace_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(workspace_dir, "data_kr", "kr_insider_transactions.xlsx")
        
        if TELEGRAM_BOT4_TOKEN and TELEGRAM_JJANG_GU_CHAT_ID:
            logger.info(f"Uploading updated Korean Insider/Block Excel sheet to Telegram chat: {TELEGRAM_JJANG_GU_CHAT_ID}...")
            caption_text = (
                "📁 *[국내 내부자 거래 & 5% 지분 변동 Excel 업데이트]*\n"
                "새로운 지분 변동 내역이 누적 엑셀 파일에 추가되어 전송합니다.\n"
                "• *InsiderTrades*: 임원 및 대주주, 최대주주 주식 변동 내역\n"
                "• *BlockHoldings*: 5% 이상 대량 보유자(기관, 펀드 등) 변동 내역"
            )
            res_file = send_telegram_document(TELEGRAM_BOT4_TOKEN, TELEGRAM_JJANG_GU_CHAT_ID, excel_path, caption=caption_text)
            if res_file and res_file.get("ok"):
                logger.info("Successfully uploaded Korean Excel sheet to Telegram.")
            else:
                logger.error(f"Failed to upload Korean Excel sheet to Telegram: {res_file}")
    else:
        logger.info("Excel file was not updated (no new unique transactions found).")

if __name__ == "__main__":
    main()

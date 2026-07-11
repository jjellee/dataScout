#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
quarterly_financials_excel.py - 한국·미국 기업 분기별 매출액/영업이익 엑셀 (각각 별도 파일)

- 한국: earnings_confirmed_cache.json (OpenDART 재무제표 API 수집분) → 분기 단독값 파생
        출력: data_dart/kr_quarterly_financials.xlsx (시트: 매출액(억원)/영업이익(억원), wide 형식)
- 미국: SEC XBRL frames API (전 기업 일괄) → 시총 $10B+ 및 watchlist 필터
        출력: data_us/us_quarterly_financials.xlsx (시트: Revenue($M)/OperatingIncome($M))
        Q4는 연간 − (Q1+Q2+Q3)로 파생 (미국은 Q4 단독 XBRL 미보고가 일반적)

사용법: python quarterly_financials_excel.py [--upload] [--test]
"""

import os
import sys
import json
import time
import argparse
import datetime
import logging

import requests

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[logging.StreamHandler(sys.stdout)])
logger = logging.getLogger("quarterly_financials")

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_DIR)
from quarterly_earnings import load_json, listed_corp_codes, parse_num, load_env  # noqa: E402

load_env()
TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_SUPPLY_DATA_CHAT_ID = os.getenv("TELEGRAM_SUPPLY_DATA_CHAT_ID")
TELEGRAM_TEST_CHAT_ID = os.getenv("TELEGRAM_TEST_CHAT_ID", "-1003843549676")

SEC_UA = {"User-Agent": "dataScout research heyork12@gmail.com"}
CONF_CACHE = os.path.join(PROJECT_DIR, "data_dart", "earnings_confirmed_cache.json")

REPRT_ORDER = {"11013": 1, "11012": 2, "11014": 3, "11011": 4}


def quarters_range():
    """2024Q1 ~ 직전 완결 분기 라벨 목록."""
    today = datetime.date.today()
    last_q_end = datetime.date(today.year, ((today.month - 1) // 3) * 3, 1) if today.month > 3 \
        else datetime.date(today.year - 1, 10, 1)
    out = []
    y, q = 2024, 1
    while (y, q) <= (last_q_end.year, (last_q_end.month - 1) // 3 + 1):
        out.append(f"{y}Q{q}")
        q += 1
        if q == 5:
            y, q = y + 1, 1
    return out


# ---------------------------------------------------------------------------
# 한국
# ---------------------------------------------------------------------------

def build_kr(out_path, quarters):
    conf = load_json(CONF_CACHE, {})
    corps = listed_corp_codes()
    # (corp, year, acct, 보고서차수) -> (3개월값, 누적값) — CFS 우선
    cum = {}
    addv = {}
    fs_pref = {}
    for key, v in conf.items():
        if key.startswith("_done_"):
            continue
        cc, year, rc, fs, acct = key.split("|")
        if acct not in ("매출액", "영업이익"):
            continue
        val = parse_num(v.get("thstrm"))
        av = parse_num(v.get("add"))
        if val is None and av is None:
            continue
        k = (cc, year, acct, REPRT_ORDER[rc])
        prev_fs = fs_pref.get(k)
        if prev_fs == "CFS" and fs != "CFS":
            continue
        if val is not None:
            cum[k] = val
        if av is not None:
            addv[k] = av
        fs_pref[k] = fs

    # 계정별 분기값 맵: {corp_code: {quarter_label: 원단위 값}}
    metric_maps = {}
    for acct in ("매출액", "영업이익"):
        rows = {}
        adds = {}
        for (cc, year, a, ro), val in cum.items():
            if a == acct:
                rows.setdefault(cc, {})[(year, ro)] = val
        for (cc, year, a, ro), val in addv.items():
            if a == acct:
                adds.setdefault(cc, {})[(year, ro)] = val
        qmap = {}
        for cc, series in rows.items():
            aser = adds.get(cc, {})
            out = {}
            for ql in quarters:
                y, q = ql.split("Q")
                q = int(q)
                # thstrm = 해당 3개월치(연간 보고서만 12개월), add = 누적치
                if q < 4:
                    v = series.get((y, q))
                    if v is None:
                        # 폴백: 누적 차분 (신규상장 등으로 직전 분기 보고서가 없는 경우)
                        aq, ap = aser.get((y, q)), (aser.get((y, q - 1)) if q > 1 else 0)
                        if aq is not None and ap is not None:
                            v = aq - ap
                else:
                    ann = series.get((y, 4))
                    v = None
                    if ann is not None:
                        add3 = aser.get((y, 3))
                        q123 = [series.get((y, i)) for i in (1, 2, 3)]
                        if add3 is not None:
                            v = ann - add3  # 3분기 누적 기반 (Q1·Q2 보고서 불필요)
                        elif all(x is not None for x in q123):
                            v = ann - sum(q123)
                if v is not None:
                    out[ql] = v
            if out:
                qmap[cc] = out
        metric_maps[acct] = qmap

    markets = kr_market_map()
    ids = {}
    for cc in set(metric_maps["매출액"]) | set(metric_maps["영업이익"]):
        name, sc = corps.get(cc, (cc, ""))
        ids[cc] = {"종목코드": sc, "종목명": name, "시장": markets.get(cc, "")}
    df, meta = assemble_combined(ids, metric_maps["매출액"], metric_maps["영업이익"], quarters,
                                 unit_div=1e8)
    _write_screening_excel(df, meta, out_path, "실적", "억원")
    logger.info(f"KR Excel: {out_path} ({len(df)}개사, T={meta['T']})")
    return len(df)


def kr_market_map():
    """corp_code → 시장구분 (KOSPI/KOSDAQ/KONEX)."""
    import glob as _glob
    m = {}
    label = {"Y": "KOSPI", "K": "KOSDAQ", "N": "KONEX"}
    for f in _glob.glob(os.path.join(PROJECT_DIR, "data_dart", "20*", "disclosures.json")):
        try:
            items = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        for x in items:
            cc = x.get("corp_code")
            cls = x.get("corp_cls")
            if cc and cls in label:
                m[cc] = label[cls]
    return m


# ---------------------------------------------------------------------------
# 미국
# ---------------------------------------------------------------------------

def _frames(tag, period):
    url = f"https://data.sec.gov/api/xbrl/frames/us-gaap/{tag}/USD/{period}.json"
    for attempt in range(4):
        try:
            r = requests.get(url, headers=SEC_UA, timeout=90)
            if r.status_code == 200:
                return {int(x["cik"]): x["val"] for x in r.json().get("data", [])}
            if r.status_code == 404:
                return {}
        except requests.exceptions.RequestException:
            time.sleep(3 * (attempt + 1))
    return {}


def us_universe():
    """시총 $10B+ 또는 watchlist 티커의 CIK 매핑."""
    from us_disclosure_summary import fetch_market_caps, fetch_cik_map, load_watchlist_tickers, MARKET_CAP_MIN
    caps = fetch_market_caps()
    cik2t = fetch_cik_map()
    watch = load_watchlist_tickers()
    out = {}
    for cik, tickers in cik2t.items():
        for t in tickers:
            if t in watch or caps.get(t, 0) >= MARKET_CAP_MIN:
                out[cik] = (t, caps.get(t, 0))
                break
    return out


def build_us(out_path, quarters):
    uni = us_universe()
    logger.info(f"US universe: {len(uni)} CIKs")
    rev = {}   # quarter -> {cik: val}
    oi = {}
    years = sorted({q.split("Q")[0] for q in quarters})
    for ql in quarters:
        period = f"CY{ql.replace('Q', 'Q')}"
        a = _frames("RevenueFromContractWithCustomerExcludingAssessedTax", period)
        b = _frames("Revenues", period)
        b.update(a)  # RFCWC 우선
        rev[ql] = b
        oi[ql] = _frames("OperatingIncomeLoss", period)
        logger.info(f"  {ql}: revenue {len(b)}, OI {len(oi[ql])}")
        time.sleep(0.3)
    # 연간 (Q4 파생용)
    ann_rev, ann_oi = {}, {}
    for y in years:
        a = _frames("RevenueFromContractWithCustomerExcludingAssessedTax", f"CY{y}")
        b = _frames("Revenues", f"CY{y}")
        b.update(a)
        ann_rev[y] = b
        ann_oi[y] = _frames("OperatingIncomeLoss", f"CY{y}")
        time.sleep(0.3)

    def q4_derive(metric_q, metric_ann, cik, y):
        fy = metric_ann.get(y, {}).get(cik)
        if fy is None:
            return None
        s = 0
        for q in (1, 2, 3):
            v = metric_q.get(f"{y}Q{q}", {}).get(cik)
            if v is None:
                return None
            s += v
        return fy - s

    # {cik: {quarter: 값}} 형태로 변환 (Q4 파생 포함)
    def to_map(metric_q, metric_ann):
        out = {}
        for cik in uni:
            series = {}
            for ql in quarters:
                y, qn = ql.split("Q")
                v = metric_q.get(ql, {}).get(cik)
                if v is None and qn == "4":
                    v = q4_derive(metric_q, metric_ann, cik, y)
                if v is not None:
                    series[ql] = v
            if series:
                out[cik] = series
        return out

    rev_map = to_map(rev, ann_rev)
    oi_map = to_map(oi, ann_oi)
    ids = {cik: {"Ticker": t, "MCap($B)": round(mc / 1e9, 1)}
           for cik, (t, mc) in uni.items() if cik in rev_map or cik in oi_map}
    df, meta = assemble_combined(ids, rev_map, oi_map, quarters, unit_div=1e6)
    _write_screening_excel(df, meta, out_path, "실적", "$M")
    logger.info(f"US Excel: {out_path} ({len(df)} companies, T={meta['T']})")
    return len(df)


# ---------------------------------------------------------------------------

def _q_to_yyyymm(ql):
    """'2024Q1' → '202403'"""
    y, q = ql.split("Q")
    return f"{y}{int(q)*3:02d}"


def assemble_combined(ids, rev_map, oi_map, quarters, unit_div, min_rev_for_sort=0):
    """'1Q26_실적 스크리닝' 포맷의 실적 블록: 메트릭별 그룹 컬럼.

    컬럼: [ID...] + [매출액 분기 시계열(YYYYMM)] + [매출 YoY T-4..T]
                  + [영업이익 분기 시계열] + [영업이익 YoY T-4..T]
    T = 채움율 30% 이상인 최신 분기. 시계열은 2024Q1~T.
    Returns (df, meta) — meta는 그룹헤더 작성용 {"rev_qs":[..], "yoy_qs":[..], "T": ql}.
    """
    import pandas as pd

    # T 결정: 채움율 30% 이상 최신 분기
    def fill_ratio(ql):
        n = sum(1 for s in rev_map.values() if ql in s)
        return n / max(len(rev_map), 1)
    T = None
    for ql in reversed(quarters):
        if fill_ratio(ql) >= 0.3:
            T = ql
            break
    if T is None:
        T = quarters[-1]
    display_qs = [q for q in quarters if q <= T]
    ti = display_qs.index(T)
    yoy_qs = display_qs[max(0, ti - 4):ti + 1]  # T-4 .. T

    def yoy(ser, ql):
        y, qn = ql.split("Q")
        cur, prev = ser.get(ql), ser.get(f"{int(y)-1}Q{qn}")
        if cur is None or prev is None:
            return None
        if prev > 0:
            return round((cur / prev - 1) * 100, 1)
        if prev <= 0 < cur:
            return "흑전"
        return None

    recs = []
    for key, idv in ids.items():
        rser = rev_map.get(key, {})
        oser = oi_map.get(key, {})
        if not rser and not oser:
            continue
        rec = dict(idv)
        for ql in display_qs:
            v = rser.get(ql)
            rec[f"매출_{_q_to_yyyymm(ql)}"] = round(v / unit_div, 1) if v is not None else None
        for i, ql in enumerate(yoy_qs):
            rec[f"매출YoY_T-{len(yoy_qs)-1-i}" if ql != T else "매출YoY_T"] = yoy(rser, ql)
        for ql in display_qs:
            v = oser.get(ql)
            rec[f"영업이익_{_q_to_yyyymm(ql)}"] = round(v / unit_div, 1) if v is not None else None
        for i, ql in enumerate(yoy_qs):
            rec[f"영업이익YoY_T-{len(yoy_qs)-1-i}" if ql != T else "영업이익YoY_T"] = yoy(oser, ql)
        recs.append(rec)
    df = pd.DataFrame(recs)
    # 정렬: T분기 매출 내림차순 (저기저 가드는 YoY 정렬이 아니므로 불필요)
    sort_col = f"매출_{_q_to_yyyymm(T)}"
    if sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=False, na_position="last")
    meta = {"display_qs": display_qs, "yoy_qs": yoy_qs, "T": T}
    return df.reset_index(drop=True), meta


def _write_screening_excel(df, meta, out_path, sheet_name, unit):
    """참조 포맷(그룹헤더 2줄) 실적 시트 작성."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    gf = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    cols = list(df.columns)
    id_cols = [c for c in cols if "_" not in str(c)]
    groups = [
        (f"매출액({unit})", [c for c in cols if str(c).startswith("매출_")]),
        ("매출액 전년동기대비 증감률(YoY, %)", [c for c in cols if str(c).startswith("매출YoY_")]),
        (f"영업이익({unit})", [c for c in cols if str(c).startswith("영업이익_") and "YoY" not in str(c)]),
        ("영업이익 전년동기대비 증감률(YoY, %)", [c for c in cols if str(c).startswith("영업이익YoY_")]),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 1행: 그룹 헤더 (병합)
    ci = len(id_cols) + 1
    for title, gcols in groups:
        if not gcols:
            continue
        ws.cell(row=1, column=ci, value=title).fill = gf
        ws.cell(row=1, column=ci).font = Font(bold=True, size=10)
        ws.cell(row=1, column=ci).alignment = Alignment(horizontal="center")
        if len(gcols) > 1:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + len(gcols) - 1)
        ci += len(gcols)
    # 2행: 컬럼 헤더
    for j, c in enumerate(cols, 1):
        label = str(c).split("_", 1)[1] if "_" in str(c) else str(c)
        cell = ws.cell(row=2, column=j, value=label)
        cell.fill, cell.font = hf, Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 13 if j <= len(id_cols) else 11
    # 데이터
    yoy_cols = {j for j, c in enumerate(cols, 1) if "YoY" in str(c)}
    num_cols = {j for j, c in enumerate(cols, 1) if "_" in str(c)}
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=j, value=row[c])
            if j in yoy_cols:
                cell.number_format = "0.0;[Red]-0.0"
            elif j in num_cols:
                cell.number_format = "#,##0;[Red]-#,##0"
    ws.freeze_panes = ws.cell(row=3, column=len(id_cols) + 1)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df)+2}"
    wb.save(out_path)


def _write_wide_excel(sheets, out_path):
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]
            ws.freeze_panes = "C2"
            n_id_cols = sum(1 for c in df.columns if not str(c)[0].isdigit())
            for ci, col in enumerate(df.columns, 1):
                hc = ws.cell(row=1, column=ci)
                hc.fill, hc.font = hf, Font(color="FFFFFF", bold=True, size=10)
                hc.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = 17 if ci <= n_id_cols else 11
            # 분기 숫자 열: 천단위 콤마 + 음수 빨강 / YoY 열: 소수 1자리
            col_names = list(df.columns)
            for row in ws.iter_rows(min_row=2, min_col=n_id_cols + 1):
                for c in row:
                    name_c = col_names[c.column - 1]
                    if str(name_c).endswith("YoY%"):
                        c.number_format = "0.0;[Red]-0.0"
                    else:
                        c.number_format = "#,##0;[Red]-#,##0"
            ws.auto_filter.ref = ws.dimensions


def send_document(file_path, caption, test=False):
    chat = TELEGRAM_TEST_CHAT_ID if test else TELEGRAM_SUPPLY_DATA_CHAT_ID
    with open(file_path, "rb") as f:
        r = requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT4_TOKEN}/sendDocument",
                          data={"chat_id": chat, "caption": caption},
                          files={"document": f}, timeout=120)
    logger.info(f"upload {os.path.basename(file_path)}: {r.json().get('ok')}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--test", action="store_true")
    args = ap.parse_args()

    quarters = quarters_range()
    logger.info(f"quarters: {quarters}")

    kr_path = os.path.join(PROJECT_DIR, "data_dart", "kr_quarterly_financials.xlsx")
    us_path = os.path.join(PROJECT_DIR, "data_us", "us_quarterly_financials.xlsx")
    n_kr = build_kr(kr_path, quarters)
    n_us = build_us(us_path, quarters)

    if args.upload:
        send_document(kr_path, f"🇰🇷 한국 상장사 분기별 매출액·영업이익 ({quarters[0]}~{quarters[-1]}, "
                               f"{n_kr:,}개사, 단위 억원, 연결 우선)", test=args.test)
        send_document(us_path, f"🇺🇸 미국 기업 분기별 Revenue·Operating Income ({quarters[0]}~{quarters[-1]}, "
                               f"{n_us:,}개사, 단위 $M, 시총 $10B+ 및 watchlist)", test=args.test)


if __name__ == "__main__":
    main()

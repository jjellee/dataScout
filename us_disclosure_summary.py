#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
us_disclosure_summary.py - 미국 SEC 공시를 유형별로 분류해 엑셀 리포트 생성·텔레그램 전송.

DART 분류 엑셀(dart_classifier.py)의 미국판.
- 대상: 시가총액 $10B 이상 + us_disclosure_watchlist.json 종목
  (단, 13D·공개매수·합병 등 저빈도 중요 공시는 전 종목 커버)
- 소스: EDGAR full-text search API (efts.sec.gov) — 8-K item 코드 포함
- 시가총액: Nasdaq screener API (전 미국 상장사)
- CIK↔티커 매핑: SEC company_tickers.json

사용법:
  python us_disclosure_summary.py                # 직전 미국 영업일 수집, 엑셀만 생성
  python us_disclosure_summary.py --upload       # 생성 후 텔레그램 업로드
  python us_disclosure_summary.py --date 2026-07-07 --upload
  python us_disclosure_summary.py --test         # 테스트 채널로 전송
"""

import os
import re
import sys
import json
import time
import argparse
import datetime
import logging

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("us_disclosure_summary")

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(PROJECT_DIR, "data_us")
UNIVERSE_CACHE = os.path.join(DATA_DIR, "us_universe_cache.json")

MARKET_CAP_MIN = 10_000_000_000  # $10B

SEC_UA = {"User-Agent": "dataScout research heyork12@gmail.com"}
EFTS_URL = "https://efts.sec.gov/LATEST/search-index"


def load_env():
    env_path = os.path.join(PROJECT_DIR, ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip().strip("'\"")


load_env()
from llm_client import deepseek_chat  # noqa: E402

TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_SUPPLY_DATA_CHAT_ID = os.getenv("TELEGRAM_SUPPLY_DATA_CHAT_ID")
TELEGRAM_TEST_CHAT_ID = os.getenv("TELEGRAM_TEST_CHAT_ID", "-1003843549676")

# ---------------------------------------------------------------------------
# 공시 유형 정의
# ---------------------------------------------------------------------------

# 전 종목 커버 (저빈도·고중요도)
ALL_CAP_FORMS = [
    "SCHEDULE 13D", "SCHEDULE 13D/A",
    "SC TO-T", "SC TO-I", "SC 14D9",
    "DEFM14A", "S-4", "425", "25", "25-NSE",
]

# 유니버스(시총 기준 + watchlist)만 커버
UNIVERSE_FORMS = [
    "8-K", "8-K/A",
    "S-1", "S-3", "F-1", "F-3",
    # 424B2 제외: 대형은행 구조화채권(structured notes) 일일 수백 건 노이즈
    "424B1", "424B3", "424B4", "424B5", "424B7", "424B8",
    "SCHEDULE 13G", "SCHEDULE 13G/A",
    "10-K", "10-Q", "20-F",
    "NT 10-K", "NT 10-Q",
]

# 폼 → 시트 매핑 (8-K는 item 기반이라 별도)
FORM_CATEGORY = {
    "S-1": "자금조달", "S-3": "자금조달", "F-1": "자금조달", "F-3": "자금조달",
    "424B1": "자금조달", "424B2": "자금조달", "424B3": "자금조달",
    "424B4": "자금조달", "424B5": "자금조달", "424B7": "자금조달", "424B8": "자금조달",
    "SCHEDULE 13D": "5%지분", "SCHEDULE 13D/A": "5%지분",
    "SCHEDULE 13G": "5%지분", "SCHEDULE 13G/A": "5%지분",
    "SC TO-T": "M&A_경영권", "SC TO-I": "M&A_경영권", "SC 14D9": "M&A_경영권",
    "DEFM14A": "M&A_경영권", "S-4": "M&A_경영권", "425": "M&A_경영권",
    "25": "M&A_경영권", "25-NSE": "M&A_경영권",
    "10-K": "실적_정기", "10-Q": "실적_정기", "20-F": "실적_정기",
    "NT 10-K": "실적_정기", "NT 10-Q": "실적_정기",
}

# 8-K item → (시트, 한글 설명). 미등재 item은 무시(루틴 공시).
ITEM_MAP = {
    "1.01": ("계약_사업", "주요 계약 체결"),
    "1.02": ("계약_사업", "주요 계약 해지"),
    "1.03": ("리스크신호", "파산·법정관리"),
    "2.01": ("M&A_경영권", "자산·사업 취득/처분 완료"),
    "2.02": ("실적_정기", "실적 발표"),
    "2.03": ("자금조달", "채무 발생(차입·사채)"),
    "2.04": ("리스크신호", "채무 조기상환 트리거"),
    "2.05": ("리스크신호", "구조조정 비용"),
    "2.06": ("리스크신호", "자산 손상"),
    "3.01": ("리스크신호", "상장유지 기준 미달"),
    "3.02": ("자금조달", "주식 사모 발행"),
    "4.01": ("리스크신호", "감사인 변경"),
    "4.02": ("리스크신호", "기존 재무제표 신뢰불가"),
    "5.01": ("M&A_경영권", "지배권 변동"),
    "5.02": ("경영진변동", "임원·이사 선임/사임"),
}

SHEET_ORDER = ["자금조달", "계약_사업", "M&A_경영권", "5%지분",
               "경영진변동", "실적_정기", "리스크신호"]

# 구조화채권(structured notes)을 상시 발행하는 금융사 — 424B* 노이즈 제외 대상
STRUCTURED_NOTE_ISSUERS = {"JPM", "C", "GS", "MS", "BAC", "WFC", "UBS", "RY",
                           "TD", "BCS", "HSBC", "DB", "BMO", "BNS", "CM", "MUFG", "NMR"}

FORM_KO = {
    "S-1": "신규 증권등록(IPO 등)", "S-3": "일괄 증권등록", "F-1": "외국계 증권등록", "F-3": "외국계 일괄등록",
    "424B1": "발행조건 확정", "424B2": "발행조건 확정", "424B3": "발행조건 확정",
    "424B4": "발행조건 확정", "424B5": "발행조건 확정", "424B7": "발행조건 확정", "424B8": "발행조건 확정",
    "SCHEDULE 13D": "5% 이상 취득(경영참여)", "SCHEDULE 13D/A": "5% 보유 변동(경영참여)",
    "SCHEDULE 13G": "5% 이상 취득(단순투자)", "SCHEDULE 13G/A": "5% 보유 변동(단순투자)",
    "SC TO-T": "공개매수(제3자)", "SC TO-I": "공개매수(자사)", "SC 14D9": "공개매수 의견표명",
    "DEFM14A": "합병 주총 위임장", "S-4": "합병·교환 증권등록", "425": "합병 관련 커뮤니케이션",
    "25": "상장폐지 신청", "25-NSE": "상장폐지(거래소)",
    "10-K": "연간 보고서", "10-Q": "분기 보고서", "20-F": "외국기업 연간보고서",
    "NT 10-K": "연간보고서 제출지연", "NT 10-Q": "분기보고서 제출지연",
    "8-K": "수시공시", "8-K/A": "수시공시(정정)",
}


# ---------------------------------------------------------------------------
# 유니버스 (시가총액 + watchlist)
# ---------------------------------------------------------------------------

def fetch_market_caps():
    """Nasdaq screener에서 전 종목 시가총액. 실패 시 캐시 사용."""
    try:
        r = requests.get(
            "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=25&download=true",
            headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"}, timeout=60)
        rows = r.json()["data"]["rows"]
        caps = {}
        for row in rows:
            try:
                caps[row["symbol"].strip().upper()] = float(row["marketCap"] or 0)
            except (ValueError, AttributeError):
                continue
        if len(caps) > 1000:
            with open(UNIVERSE_CACHE, "w", encoding="utf-8") as f:
                json.dump(caps, f)
            logger.info(f"Market caps fetched: {len(caps)} tickers.")
            return caps
    except Exception as e:
        logger.warning(f"Nasdaq screener failed: {e}")
    if os.path.exists(UNIVERSE_CACHE):
        with open(UNIVERSE_CACHE, "r", encoding="utf-8") as f:
            caps = json.load(f)
        logger.info(f"Market caps loaded from cache: {len(caps)} tickers.")
        return caps
    return {}


def fetch_cik_map():
    """SEC company_tickers.json → {cik(int): [ticker,...]}"""
    r = requests.get("https://www.sec.gov/files/company_tickers.json",
                     headers=SEC_UA, timeout=60)
    d = r.json()
    cik2tickers = {}
    for rec in d.values():
        cik2tickers.setdefault(int(rec["cik_str"]), []).append(rec["ticker"].upper())
    return cik2tickers


def load_watchlist_tickers():
    path = os.path.join(PROJECT_DIR, "us_disclosure_watchlist.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return set(t.upper() for t in json.load(f).keys())
    except Exception as e:
        logger.warning(f"watchlist load failed: {e}")
        return set()


def build_universe():
    """returns (universe_ciks:set, cik2tickers, caps)"""
    caps = fetch_market_caps()
    cik2tickers = fetch_cik_map()
    watch = load_watchlist_tickers()
    universe = set()
    for cik, tickers in cik2tickers.items():
        for t in tickers:
            if t in watch or caps.get(t, 0) >= MARKET_CAP_MIN:
                universe.add(cik)
                break
    logger.info(f"Universe: {len(universe)} CIKs "
                f"(mcap >= ${MARKET_CAP_MIN/1e9:.0f}B or watchlist {len(watch)}).")
    return universe, cik2tickers, caps


# ---------------------------------------------------------------------------
# EDGAR 수집
# ---------------------------------------------------------------------------

def efts_search(form, date_str):
    """해당 일자·폼의 전체 공시를 페이지네이션으로 수집."""
    hits, frm = [], 0
    while True:
        params = {"q": "", "forms": form, "startdt": date_str, "enddt": date_str}
        if frm:
            params["from"] = str(frm)
        data = None
        for attempt in range(4):
            try:
                r = requests.get(EFTS_URL, params=params, headers=SEC_UA, timeout=30)
                data = r.json()
                if "hits" in data:
                    break
            except Exception:
                pass
            time.sleep(1 + attempt)
        if not data or "hits" not in data:
            logger.warning(f"efts search failed: {form} from={frm}")
            break
        page = data["hits"]["hits"]
        total = data["hits"]["total"]["value"]
        hits.extend(page)
        frm += len(page)
        if frm >= total or not page or frm >= 9990:
            break
        time.sleep(0.15)
    return hits


def collect_filings(date_str, universe, cik2tickers, caps):
    """일자별 공시 수집 → 분류된 행 리스트 {sheet: [row,...]}"""
    sheets = {s: [] for s in SHEET_ORDER}
    seen = set()  # (adsh, sheet) 중복 방지

    def display_info(src):
        ciks = [int(c) for c in src.get("ciks", [])]
        names = src.get("display_names", [])
        # ciks[0] = 대상회사(subject). 13D/G·공개매수에서 신고자가 아닌 대상 기준.
        main_cik = ciks[0] if ciks else None
        ticker, mcap = "", None
        for c in ciks:
            ts = [t for t in cik2tickers.get(c, []) if t in caps]
            if ts and (ticker == "" or c == main_cik):
                ticker, mcap = ts[0], caps[ts[0]]
        name = "; ".join(re.sub(r"\s*\([^)]*\)\s*$", "", n.split("(CIK")[0]).strip()
                         for n in names)[:80]
        return main_cik, ticker, name, mcap

    def make_row(src, sheet, item_desc="", doc_id=""):
        adsh = src.get("adsh", "")
        if (adsh, sheet) in seen:
            return
        seen.add((adsh, sheet))
        cik, ticker, name, mcap = display_info(src)
        form = src.get("file_type") or (src.get("root_forms") or [""])[0]
        form_desc = FORM_KO.get(form, "")
        if not form_desc and form.endswith("/A"):
            base = FORM_KO.get(form[:-2], "")
            form_desc = f"{base}(정정)" if base else ""
        link = ""
        if cik and adsh:
            link = (f"https://www.sec.gov/Archives/edgar/data/{cik}/"
                    f"{adsh.replace('-', '')}/{adsh}-index.htm")
        desc_txt = (src.get("file_description") or "").strip()[:120]
        # 폼 이름만 반복하는 무의미한 설명은 생략
        if desc_txt.upper() in (form.upper(), f"FORM {form}".upper(), "CURRENT REPORT"):
            desc_txt = ""
        sheets[sheet].append({
            "접수일자": src.get("file_date", ""),
            "티커": ticker,
            "회사명": name,
            "시가총액($B)": round(mcap / 1e9, 1) if mcap else None,
            "공시유형": form,
            "유형설명": item_desc or form_desc,
            "핵심내용": desc_txt,
            "신고자": None, "보유비율(%)": None, "보유주식수": None,
            "링크": link,
            "_adsh": adsh, "_cik": cik,
            "_doc": doc_id.split(":", 1)[1] if ":" in doc_id else "",
            "_form": form,
        })

    # 1) 전 종목 커버 폼
    for form in ALL_CAP_FORMS:
        hits = efts_search(form, date_str)
        for h in hits:
            src = h["_source"]
            make_row(src, FORM_CATEGORY[form], doc_id=h.get("_id", ""))
        if hits:
            logger.info(f"[{form}] {len(hits)} filings (all-cap).")
        time.sleep(0.2)

    # 2) 유니버스 한정 폼
    for form in UNIVERSE_FORMS:
        hits = efts_search(form, date_str)
        kept = 0
        for h in hits:
            src = h["_source"]
            ciks = [int(c) for c in src.get("ciks", [])]
            # 대상회사(ciks[0]) 기준 유니버스 판정 — 13G처럼 신고자만 대형인 건 제외
            if not ciks or ciks[0] not in universe:
                continue
            if form.startswith("424B"):
                subj_tickers = set(cik2tickers.get(ciks[0], []))
                if subj_tickers & STRUCTURED_NOTE_ISSUERS:
                    continue
            if form.startswith("8-K"):
                items = src.get("items") or []
                mapped = [(ITEM_MAP[i], i) for i in items if i in ITEM_MAP]
                for (sheet, desc), item_no in mapped:
                    make_row(src, sheet, f"[{item_no}] {desc}", doc_id=h.get("_id", ""))
                kept += 1 if mapped else 0
            else:
                make_row(src, FORM_CATEGORY[form], doc_id=h.get("_id", ""))
                kept += 1
        if kept:
            logger.info(f"[{form}] {kept}/{len(hits)} filings in universe.")
        time.sleep(0.2)

    return sheets


# ---------------------------------------------------------------------------
# 상세 파싱 (13D/G 구조화 XML + 원문 DeepSeek 요약)
# ---------------------------------------------------------------------------

def _fetch_doc_text(cik, adsh, filename, max_chars):
    """공시 원문 문서를 받아 텍스트로 변환."""
    url = (f"https://www.sec.gov/Archives/edgar/data/{cik}/"
           f"{adsh.replace('-', '')}/{filename}")
    try:
        r = requests.get(url, headers=SEC_UA, timeout=30)
        if r.status_code != 200:
            return ""
        text = BeautifulSoup(r.content, "html.parser").get_text(" ", strip=True)
        return re.sub(r"\s+", " ", text)[:max_chars]
    except Exception as e:
        logger.warning(f"doc fetch failed {url}: {e}")
        return ""


def enrich_13dg(rows):
    """SCHEDULE 13D/G primary_doc.xml → 신고자·보유비율·보유주식수 + 13D 목적(item4)."""
    for row in rows:
        if not row["_form"].startswith("SCHEDULE 13") or not row["_doc"].endswith(".xml"):
            continue
        url = (f"https://www.sec.gov/Archives/edgar/data/{row['_cik']}/"
               f"{row['_adsh'].replace('-', '')}/{row['_doc']}")
        try:
            xml = requests.get(url, headers=SEC_UA, timeout=30).text
        except Exception as e:
            logger.warning(f"13D/G xml fetch failed: {e}")
            continue
        persons = re.findall(r"<reportingPersonName>(.*?)</reportingPersonName>", xml, re.DOTALL)
        percents = re.findall(r"<percentOfClass>([\d.]+)</percentOfClass>", xml)
        amounts = re.findall(r"<aggregateAmountOwned>([\d.]+)</aggregateAmountOwned>", xml)
        if persons:
            uniq = list(dict.fromkeys(p.strip() for p in persons))
            row["신고자"] = ", ".join(uniq[:2]) + (" 외" if len(uniq) > 2 else "")
        if percents:
            vals = [float(p) for p in percents]
            top = vals.index(max(vals))
            row["보유비율(%)"] = vals[top]
            # 최대 비율 신고자와 같은 인덱스의 주식수로 짝을 맞춘다
            if len(amounts) == len(percents):
                row["보유주식수"] = int(float(amounts[top]))
            elif amounts:
                row["보유주식수"] = int(max(float(a) for a in amounts))
        elif amounts:
            row["보유주식수"] = int(max(float(a) for a in amounts))
        # 13D 목적(item4) 원문 → LLM 요약 대상으로 저장
        m = re.search(r"<item4>(.*?)</item4>", xml, re.DOTALL)
        if m and "13D" in row["_form"]:
            purpose = re.sub(r"<[^>]+>", " ", m.group(1))
            row["_text"] = re.sub(r"\s+", " ", purpose).strip()[:3000]
        time.sleep(0.12)


LLM_SKIP_FORMS = {"10-K", "10-Q", "20-F", "SCHEDULE 13G", "SCHEDULE 13G/A"}


def enrich_llm(all_rows):
    """원문 텍스트를 DeepSeek으로 핵심내용(한국어 1~2문장) 추출."""
    targets = []
    for row in all_rows:
        if row["_form"] in LLM_SKIP_FORMS:
            continue
        if "_text" not in row:
            if not row["_doc"] or row["_doc"].endswith(".xml"):
                continue
            row["_text"] = _fetch_doc_text(row["_cik"], row["_adsh"], row["_doc"], 4000)
            time.sleep(0.12)
        if row.get("_text"):
            targets.append(row)
    logger.info(f"LLM extraction targets: {len(targets)} filings.")

    BATCH = 5
    for i in range(0, len(targets), BATCH):
        batch = targets[i:i + BATCH]
        parts = []
        for j, row in enumerate(batch, 1):
            parts.append(f"### 공시 {j} ({row['회사명'][:40]} / {row['_form']} / {row['유형설명']})\n"
                         f"{row['_text'][:3500]}")
        prompt = (
            "아래는 미국 SEC 공시 원문 발췌들이야. 각 공시의 핵심 사실을 한국어 1~2문장으로 추출해줘. "
            "금액·수량·상대방·조건 등 구체적 숫자가 있으면 반드시 포함하고, 원문에 없는 내용은 지어내지 마. "
            "예비신고서라 금액이 공란이면 무엇을 발행·계약하는지 요약하고 '(금액 미정)'을 붙여. "
            "본문이 목차·표지뿐이라 정말 아무것도 알 수 없을 때만 '본문 정보 부족'이라고 써.\n"
            "출력 형식: 각 공시마다 `[1] 내용` 형식 한 줄씩, 서론·꼬리말 없이.\n\n"
            + "\n\n".join(parts)
        )
        text = deepseek_chat(prompt, temperature=0.2, max_tokens=2048, timeout=120)
        if not text:
            continue
        for line in text.splitlines():
            m = re.match(r"\**\[(\d+)\]\**\s*(.+)", line.strip())
            if m:
                idx = int(m.group(1)) - 1
                if 0 <= idx < len(batch) and "본문 정보 부족" not in m.group(2):
                    batch[idx]["핵심내용"] = m.group(2).strip()[:300]
        logger.info(f"LLM extraction {min(i + BATCH, len(targets))}/{len(targets)}")


# ---------------------------------------------------------------------------
# 엑셀 생성
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
BASE_FONT = Font(size=10)

BASE_COLUMNS = ["접수일자", "티커", "회사명", "시가총액($B)", "공시유형", "유형설명", "핵심내용", "링크"]
HOLDER_COLUMNS = ["접수일자", "티커", "회사명", "시가총액($B)", "공시유형", "유형설명",
                  "신고자", "보유비율(%)", "보유주식수", "핵심내용", "링크"]
COL_WIDTHS = {"접수일자": 11, "티커": 8, "회사명": 32, "시가총액($B)": 12, "공시유형": 15,
              "유형설명": 24, "핵심내용": 70, "신고자": 26, "보유비율(%)": 11,
              "보유주식수": 13, "링크": 10}


def sheet_columns(sheet_name):
    return HOLDER_COLUMNS if sheet_name == "5%지분" else BASE_COLUMNS


def build_excel(sheets, date_str, out_path, presorted=False):
    wb = Workbook()
    wb.remove(wb.active)
    total_rows = 0
    for sheet_name in SHEET_ORDER:
        rows = sheets[sheet_name]
        columns = sheet_columns(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for ci, col in enumerate(columns, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill, c.font = HEADER_FILL, HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[col]
        if not presorted:
            rows.sort(key=lambda r: (-(r["시가총액($B)"] or 0), r["티커"]))
        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(columns, 1):
                val = row[col]
                c = ws.cell(row=ri, column=ci)
                c.border, c.font = THIN_BORDER, BASE_FONT
                if col == "링크" and val:
                    c.value, c.hyperlink, c.font = "EDGAR", val, LINK_FONT
                    c.alignment = Alignment(horizontal="center")
                elif col == "핵심내용":
                    c.value = val
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    c.value = val
                    if col in ("접수일자", "티커", "공시유형"):
                        c.alignment = Alignment(horizontal="center")
                    elif col == "시가총액($B)":
                        c.number_format = "#,##0.0"
                        c.alignment = Alignment(horizontal="right")
                    elif col == "보유비율(%)":
                        c.number_format = "0.00"
                        c.alignment = Alignment(horizontal="right")
                    elif col == "보유주식수":
                        c.number_format = "#,##0"
                        c.alignment = Alignment(horizontal="right")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(len(rows)+1, 2)}"
        total_rows += len(rows)
    wb.save(out_path)
    logger.info(f"Excel saved: {out_path} ({total_rows} rows).")
    return total_rows


# ---------------------------------------------------------------------------
# 텔레그램
# ---------------------------------------------------------------------------

def send_document(token, chat_id, file_path, caption):
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as f:
        r = requests.post(url, data={"chat_id": chat_id, "caption": caption},
                          files={"document": f}, timeout=120)
    ok = r.json().get("ok")
    logger.info(f"Telegram upload {'OK' if ok else 'FAILED: ' + r.text[:200]}")
    return ok


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def default_target_date():
    """직전 완료된 미국(ET) 영업일. EDGAR 접수 마감(22:00 ET) 이후면 당일."""
    now_et = datetime.datetime.utcnow() - datetime.timedelta(hours=4)
    d = now_et.date()
    if now_et.hour < 22:
        d -= datetime.timedelta(days=1)
    while d.weekday() >= 5:
        d -= datetime.timedelta(days=1)
    return d


ROW_CACHE = os.path.join(DATA_DIR, "us_disclosures_cache.json")


def load_row_cache():
    try:
        with open(ROW_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_row_cache(cache):
    tmp = ROW_CACHE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    os.replace(tmp, ROW_CACHE)


def collect_date_into_cache(target, universe, cik2tickers, caps, cache):
    """하루치 수집·보강 후 누적 캐시에 병합. 이미 캐시된 행은 재보강하지 않음."""
    date_str = target.isoformat()
    sheets = collect_filings(date_str, universe, cik2tickers, caps)

    new_rows = []
    for sheet in SHEET_ORDER:
        for row in sheets[sheet]:
            key = f"{row['_adsh']}|{sheet}"
            if key in cache:
                continue
            row["_sheet"] = sheet
            new_rows.append(row)
    # 신규 행만 보강 (13D/G XML + DeepSeek)
    enrich_13dg([r for r in new_rows if r["_sheet"] == "5%지분"])
    enrich_llm(new_rows)
    for row in new_rows:
        row.pop("_text", None)
        cache[f"{row['_adsh']}|{row['_sheet']}"] = row
    logger.info(f"[{date_str}] new rows merged: {len(new_rows)} (cache {len(cache)})")
    return len(new_rows)


def build_cumulative_excel(cache, out_path):
    sheets = {s: [] for s in SHEET_ORDER}
    for key, row in cache.items():
        sheet = row.get("_sheet") or key.split("|", 1)[1]
        if sheet in sheets:
            sheets[sheet].append(row)
    # 누적본은 최신 접수일 우선, 같은 날은 시총 순
    for s in sheets:
        sheets[s].sort(key=lambda r: (r.get("접수일자", ""), r.get("시가총액($B)") or 0), reverse=True)
    total = build_excel(sheets, "", out_path, presorted=True)
    return total, sheets


def main():
    parser = argparse.ArgumentParser(description="US SEC disclosure summary Excel (누적).")
    parser.add_argument("--date", help="대상 일자 (YYYY-MM-DD, ET 기준)")
    parser.add_argument("--backfill", help="백필 시작일 (YYYY-MM-DD) — 시작일~기본 대상일 순회")
    parser.add_argument("--upload", action="store_true", help="텔레그램 업로드")
    parser.add_argument("--test", action="store_true", help="테스트 채널로 전송")
    args = parser.parse_args()

    end = datetime.date.fromisoformat(args.date) if args.date else default_target_date()
    if args.backfill:
        start = datetime.date.fromisoformat(args.backfill)
        dates = []
        d = start
        while d <= end:
            if d.weekday() < 5:
                dates.append(d)
            d += datetime.timedelta(days=1)
    else:
        dates = [end]

    universe, cik2tickers, caps = build_universe()
    cache = load_row_cache()
    new_total = 0
    for target in dates:
        new_total += collect_date_into_cache(target, universe, cik2tickers, caps, cache)
        save_row_cache(cache)

    out_path = os.path.join(DATA_DIR, "us_disclosures_summary.xlsx")
    total, sheets = build_cumulative_excel(cache, out_path)
    counts = ", ".join(f"{s} {len(sheets[s])}" for s in SHEET_ORDER if sheets[s])
    dts = sorted({r.get("접수일자", "") for r in cache.values() if r.get("접수일자")})
    logger.info(f"Done: 누적 {total} rows, 신규 {new_total} ({counts})")

    if args.upload and total > 0:
        chat_id = TELEGRAM_TEST_CHAT_ID if args.test else TELEGRAM_SUPPLY_DATA_CHAT_ID
        caption = (f"🇺🇸 미국 공시 요약 누적본 ({dts[0]} ~ {dts[-1]})\n"
                   f"총 {total:,}건 (오늘 신규 {new_total}) — {counts}\n"
                   f"대상: 시총 $10B+ 및 관심종목 (13D·공개매수·합병은 전 종목)")
        send_document(TELEGRAM_BOT4_TOKEN, chat_id, out_path, caption)
    elif args.upload:
        logger.info("No rows — skip upload.")


if __name__ == "__main__":
    main()

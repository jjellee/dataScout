#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import glob
import json
import re
import time
import datetime
import logging
import argparse
import requests
import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dart_officer_parser import parse_officer_report_html
import warnings
from bs4 import XMLParsedAsHTMLWarning

warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Load env variables from .env file
def load_env():
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    env_path = os.path.join(workspace_dir, ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    val_str = val.strip().strip("'").strip('"')
                    os.environ[key.strip()] = val_str

load_env()
DART_API_KEY = os.getenv("DART_API_KEY")
TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_TEST_CHAT_ID = os.getenv("TELEGRAM_TEST_CHAT_ID")  # antbot channel

# Cache path for parsed disclosures to avoid repeated API hits or re-parsing HTML
CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_dart", "mezzanine_cache.json")

def load_cache():
    if os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load cache: {e}")
    return {}

def save_cache(cache):
    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Failed to save cache: {e}")

def classify_disclosure(report_nm):
    """Classifies a disclosure by checking keywords in report_nm."""
    nm = str(report_nm).replace(" ", "").strip()
    
    # 1. 정기공시
    if any(k in nm for k in ["사업보고서", "반기보고서", "분기보고서"]):
        return "정기공시"
        
    # 2. 지분공시
    if any(k in nm for k in ["주식등의대량보유상황보고서", "임원ㆍ주요주주소유주식변동보고서", "임원.주요주주소유주식변동보고서", "최대주주등소유주식변동신고서", "소유주식변동", "소유주식보고서"]):
        return "지분공시"
        
    # 3. 신규시설투자 (Dedicated Category)
    if "신규시설투자" in nm:
        return "신규시설투자"
        
    # 4. 자금조달_증자
    if any(k in nm for k in ["유상증자결정", "무상증자결정", "사채발행결정", "사채발행"]) or \
       (any(k in nm for k in ["전환사채", "신주인수권부사채", "교환사채"]) and "발행" in nm):
        return "자금조달_증자"
        
    # 5. 영업활동_계약
    if any(k in nm for k in ["단일판매ㆍ공급계약체결", "단일판매.공급계약체결", "공급계약체결", "영업정지", "특허권취득", "기술도입", "업무제휴", "공급계약"]):
        return "영업활동_계약"
        
    # 6. 재무_채무보증
    if any(k in nm for k in ["타인에대한채무보증결정", "금전대여결정", "담보제공결정", "채무보증", "금전대여"]):
        return "재무_채무보증"
        
    # 7. 경영권_지배구조
    if any(k in nm for k in ["최대주주변경", "합병결정", "회사분할결정", "분할결정", "주식교환", "영업양수결정", "영업양도결정", "경영권분쟁", "주주총회"]):
        return "경영권_지배구조"
        
    # 8. 재무_자기주식
    if any(k in nm for k in ["자기주식취득결정", "자기주식취득신탁계약", "자기주식신탁계약체결결정", "자기주식소각결정", "주식소각결정", "신탁계약체결결정"]):
        return "재무_자기주식"

    # 9. 자산취득_처분
    if any(k in nm for k in ["유형자산취득결정", "유형자산양수결정", "타법인주식및출자증권취득결정", "타법인주식및출자증권처분결정"]):
        return "자산취득_처분"

    return "기타공시"
 
def identify_base_report_type(report_nm):
    """Identifies the fine-grained document type."""
    nm = str(report_nm).replace(" ", "").strip()
    if "전환사채" in nm and "발행결정" in nm:
        return "CB"
    elif "신주인수권부사채" in nm and "발행결정" in nm:
        return "BW"
    elif "교환사채" in nm and "발행결정" in nm:
        return "EB"
    elif "유상증자결정" in nm:
        return "유상증자"
    elif "무상증자결정" in nm:
        return "무상증자"
    elif "단일판매" in nm or "공급계약" in nm:
        return "공급계약"
    elif "신규시설투자" in nm:
        return "시설투자"
    elif "채무보증결정" in nm:
        return "채무보증"
    elif "금전대여결정" in nm:
        return "금전대여"
    elif "자기주식취득결정" in nm:
        return "자기주식취득"
    elif "자기주식취득신탁계약" in nm or "신탁계약체결결정" in nm:
        return "자기주식신탁"
    elif "자기주식소각결정" in nm or "주식소각결정" in nm:
        return "자기주식소각"
    elif "유형자산취득결정" in nm:
        return "유형자산취득"
    elif "타법인주식및출자증권취득결정" in nm:
        return "타법인증권취득"
    elif "타법인주식및출자증권처분결정" in nm:
        return "타법인증권처분"
    return "기타"



def clean_numeric(val):
    """Cleans numeric values from OpenDART API for calculation/formatting in Excel."""
    if not val or val == "-" or val == "N/A" or val == "미해당" or val == "해당없음":
        return None
    try:
        cleaned = re.sub(r'[₩,%\s]', '', str(val))
        if not cleaned:
            return None
        if '.' in cleaned:
            return float(cleaned)
        else:
            return int(cleaned)
    except Exception:
        return val

def extract_dates(text):
    if not text:
        return []
    found_dates = []
    # Pattern 1: YYYY년 MM월 DD일
    for m in re.finditer(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일', text):
        date_str = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        found_dates.append((m.start(), date_str))
    # Pattern 2: YYYY-MM-DD
    for m in re.finditer(r'(\d{4})-(\d{2})-(\d{2})', text):
        date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        found_dates.append((m.start(), date_str))
    # Pattern 3: YYYY.MM.DD
    for m in re.finditer(r'(\d{4})\s*\.\s*(\d{1,2})\s*\.\s*(\d{1,2})', text):
        date_str = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        found_dates.append((m.start(), date_str))
    # Sort by start position
    found_dates.sort(key=lambda x: x[0])
    return [d[1] for d in found_dates]

def extract_option_start_date(text, opt_type="call"):
    """Extracts the first valid option request start date (청구일) from option text snippet."""
    if not text or text in ["N/A", "N/A (HTML 파일 없음)", "N/A (파일 에러)"] or text.replace(" ", "") in ["해당사항없음", "해당없음", "미해당", "없음"]:
        return "-"
    
    text_clean = re.sub(r'\s+', ' ', text)
    keywords = ["청구기간", "청구 시작", "청구일", "행사기간", "행사 시작"] if opt_type == "put" else ["행사기간", "청구기간", "청구일", "매도청구기간", "행사 시작"]
    
    for kw in keywords:
        pos = text_clean.find(kw)
        if pos != -1:
            sub = text_clean[pos:pos+500]
            dates = extract_dates(sub)
            if dates:
                return dates[0]
                
    dates = extract_dates(text)
    if dates:
        return dates[0]
    return "-"

def find_original_date_from_html(html_path):
    """Parses the HTML to find the original filing date of an amended ('정정') disclosure."""
    if not os.path.exists(html_path):
        return None
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        for table in soup.find_all('table'):
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    key = cols[0].replace(" ", "").replace("ㆍ", "").replace(".", "")
                    if any(pat in key for pat in ["정정관련공시서류제출일", "최초제출일"]):
                        val = cols[1]
                        date_match = re.search(r'(\d{4})[-년\s]*(\d{1,2})[-월\s]*(\d{1,2})', val)
                        if date_match:
                            y, m, d = date_match.groups()
                            return f"{y}{int(m):02d}{int(d):02d}"
    except Exception as e:
        logger.error(f"Error finding original date in amendment HTML {html_path}: {e}")
    return None

def parse_officer_report_html(html_path, report_type):
    """
    Parses 5%/officer report HTML files.
    report_type: '임원보고' or '대량보유'
    Returns list of dicts with keys:
        reporter_name, relationship, change_reason, shares_change, avg_price,
        shares_after, ownership_pct
    """
    results = []
    if not os.path.exists(html_path):
        return results

    try:
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
    except Exception as e:
        logger.error(f"Error reading officer report HTML {html_path}: {e}")
        return results

    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all('table')

    def clean_text(t):
        """Strip whitespace and normalize text."""
        return re.sub(r'\s+', ' ', t).strip()

    def parse_number(val):
        """Parse a number string with commas, parentheses for negatives."""
        if not val or val.strip() in ['-', '', '(   )']:
            return None
        val = val.strip()
        neg = False
        if val.startswith('-') and len(val) > 1 and not val[1:].strip().startswith('-'):
            neg = True
            val = val[1:]
        elif val.startswith('(') and val.endswith(')'):
            neg = True
            val = val[1:-1]
        cleaned = re.sub(r'[^\d.]', '', val)
        if not cleaned:
            return None
        try:
            num = float(cleaned) if '.' in cleaned else int(cleaned)
            return -num if neg else num
        except ValueError:
            return None

    if report_type == '임원보고':
        # --- Parse 임원보고 (Officer/Key Shareholder Report) ---
        reporter_name = "-"
        relationship = "-"
        ownership_pct = None
        shares_change_total = 0
        shares_after_total = 0
        avg_price = None
        change_reason = "-"

        for table in tables:
            rows = table.find_all('tr')
            if not rows:
                continue
            first_text = clean_text(rows[0].get_text())

            # Table with 보고구분 → reporter info
            if '보고구분' in first_text:
                for row in rows:
                    cells = row.find_all(recursive=False)
                    cell_texts = [clean_text(c.get_text()) for c in cells]
                    # 성명(명칭) row
                    if any('성명' in ct or '명칭' in ct for ct in cell_texts[:2]):
                        # Name is typically in the 3rd cell (한글 value)
                        for j, ct in enumerate(cell_texts):
                            if ct in ['한 글', '한글'] and j + 1 < len(cell_texts):
                                reporter_name = cell_texts[j + 1].replace(' ', '').strip()
                                break
                        if reporter_name == "-" and len(cell_texts) >= 3:
                            reporter_name = cell_texts[2].replace(' ', '').strip()
                    # 발행회사와의 관계 row
                    if any('발행회사와의' in ct or '관계' in ct for ct in cell_texts[:2]):
                        # relationship description (e.g. 임원(등기여부) / 비등기임원 / 직위명)
                        rel_parts = []
                        for ct in cell_texts[1:]:
                            ct_clean = ct.strip()
                            if ct_clean and ct_clean != '-':
                                rel_parts.append(ct_clean)
                        if rel_parts:
                            relationship = ' '.join(rel_parts)

            # Table with 소유비율(%) → ownership percentage
            if '소유비율' in first_text or '발행주식' in first_text:
                for row in rows:
                    cells = row.find_all(recursive=False)
                    cell_texts = [clean_text(c.get_text()) for c in cells]
                    # Look for the data row (not header) that has numeric values
                    if len(cell_texts) >= 3 and not any('비율' in ct for ct in cell_texts):
                        # ownership_pct is typically the 3rd or 4th value
                        for ct in cell_texts[2:]:
                            pct = parse_number(ct)
                            if pct is not None and 0 < pct < 100:
                                ownership_pct = pct
                                break

            # Table with 보고사유 → transaction details
            if '보고사유' in first_text:
                data_rows = []
                header_done = False
                for row in rows:
                    cells = row.find_all(recursive=False)
                    cell_texts = [clean_text(c.get_text()) for c in cells]
                    if not header_done:
                        # Skip header rows
                        if any('변동전' in ct for ct in cell_texts) or any('보고사유' in ct for ct in cell_texts):
                            header_done = '변동전' in ' '.join(cell_texts)
                            continue
                    else:
                        if any('합' in ct and '계' in ct for ct in cell_texts[:2]):
                            # Summary row: use for totals
                            for ct in cell_texts:
                                n = parse_number(ct)
                                if n is not None and abs(n) > 0:
                                    if shares_change_total == 0:
                                        # skip변동전 if first
                                        pass
                            continue
                        if len(cell_texts) >= 6:
                            data_rows.append(cell_texts)

                # Process data rows for this officer report
                total_change = 0
                total_after = 0
                weighted_price_sum = 0
                price_count = 0
                reasons = []

                for dr in data_rows:
                    # dr structure: [보고사유, 변동일, 종류, 변동전, 증감, 변동후, 단가, 비고, ...]
                    reason = dr[0] if len(dr) > 0 else "-"
                    if reason and reason != '-':
                        reasons.append(reason)

                    change = parse_number(dr[4]) if len(dr) > 4 else None
                    after = parse_number(dr[5]) if len(dr) > 5 else None
                    price_raw = dr[6] if len(dr) > 6 else None

                    # Clean price: remove parenthetical notes like "( 원)"
                    if price_raw:
                        price_val = parse_number(re.sub(r'\([^)]*\)', '', price_raw))
                    else:
                        price_val = None

                    if change is not None:
                        total_change += int(change)
                    if after is not None:
                        total_after = max(total_after, int(after))
                    if price_val is not None and change is not None and abs(change) > 0:
                        weighted_price_sum += price_val * abs(change)
                        price_count += abs(change)

                shares_change_total = total_change
                shares_after_total = total_after
                if price_count > 0:
                    avg_price = round(weighted_price_sum / price_count)
                change_reason = ', '.join(reasons) if reasons else "-"

        # Build ownership % from table 5 if not found in table 7
        if ownership_pct is None:
            for table in tables:
                rows = table.find_all('tr')
                all_text = table.get_text()
                if '이번보고서' in all_text and '비율' in all_text:
                    for row in rows:
                        cells = row.find_all(recursive=False)
                        cell_texts = [clean_text(c.get_text()) for c in cells]
                        if any('이번' in ct for ct in cell_texts):
                            for ct in reversed(cell_texts):
                                pct = parse_number(ct)
                                if pct is not None and 0 < pct < 100:
                                    ownership_pct = pct
                                    break
                            break

        results.append({
            "reporter_name": reporter_name,
            "relationship": relationship,
            "change_reason": change_reason,
            "shares_change": shares_change_total,
            "avg_price": avg_price,
            "shares_after": shares_after_total,
            "ownership_pct": ownership_pct
        })

    elif report_type == '대량보유':
        # --- Parse 대량보유 (5% Bulk Ownership Report) ---
        change_reason = "-"
        overall_pct = None

        # 1. Get 변동사유 / 변경사유
        for table in tables:
            all_text = table.get_text()
            if '변동사유' in all_text or '변동방법' in all_text:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(recursive=False)
                    cell_texts = [clean_text(c.get_text()) for c in cells]
                    if len(cell_texts) >= 2:
                        key = cell_texts[0].replace(' ', '')
                        val = cell_texts[1].strip()
                        if '변동사유' in key and val and val != '-':
                            change_reason = val
                        elif '변경사유' in key and val and val != '-' and change_reason == '-':
                            change_reason = val

        # 2. Get overall ownership % from summary table (이번보고서 row)
        for table in tables:
            all_text = table.get_text()
            if '이번보고서' in all_text and '비율' in all_text:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(recursive=False)
                    cell_texts = [clean_text(c.get_text()) for c in cells]
                    if any('이번' in ct for ct in cell_texts):
                        for ct in cell_texts:
                            pct = parse_number(ct)
                            if pct is not None and 0 < pct < 100:
                                overall_pct = pct
                                break
                        break

        # 3. Parse detail transaction table (변동일 + 취득/처분단가)
        detail_transactions = {}  # key: reporter_name -> list of (change, price)

        for table in tables:
            all_text = table.get_text()
            if '변동일' not in all_text or '취득/처분단가' not in all_text:
                continue
            rows = table.find_all('tr')
            header_done = False
            for row in rows:
                cells = row.find_all(recursive=False)
                cell_texts = [clean_text(c.get_text()) for c in cells]

                if not header_done:
                    if any('변동전' in ct for ct in cell_texts):
                        header_done = True
                    continue

                # Skip empty/dash rows
                if all(ct in ['-', '', '(   )'] for ct in cell_texts):
                    continue
                if len(cell_texts) < 8:
                    continue

                # Structure: [성명, 생년월일, 변동일, 취득/처분방법, 종류, 변동전, 증감, 변동후, 단가, ...]
                name = cell_texts[0].replace(' ', '')
                if not name or name == '-':
                    continue

                change = parse_number(cell_texts[6]) if len(cell_texts) > 6 else None
                price_val = parse_number(cell_texts[8]) if len(cell_texts) > 8 else None
                after = parse_number(cell_texts[7]) if len(cell_texts) > 7 else None
                method = cell_texts[3] if len(cell_texts) > 3 else "-"

                if name not in detail_transactions:
                    detail_transactions[name] = {
                        "changes": [],
                        "last_after": 0,
                        "methods": []
                    }

                if change is not None:
                    detail_transactions[name]["changes"].append((int(change), price_val))
                if after is not None:
                    detail_transactions[name]["last_after"] = int(after)
                if method and method != '-':
                    detail_transactions[name]["methods"].append(method)

        if detail_transactions:
            # Build one row per subject
            for name, txns in detail_transactions.items():
                total_change = sum(c for c, p in txns["changes"])
                # Weighted average price
                w_sum = 0
                w_count = 0
                for c, p in txns["changes"]:
                    if p is not None and abs(c) > 0:
                        w_sum += p * abs(c)
                        w_count += abs(c)
                computed_avg_price = round(w_sum / w_count) if w_count > 0 else None

                # Combine unique methods as change reason
                unique_methods = list(dict.fromkeys(txns["methods"]))
                reason = ', '.join(unique_methods) if unique_methods else change_reason

                results.append({
                    "reporter_name": name,
                    "relationship": "특별관계자",
                    "change_reason": reason,
                    "shares_change": total_change,
                    "avg_price": computed_avg_price,
                    "shares_after": txns["last_after"],
                    "ownership_pct": overall_pct
                })
        else:
            # No detail transactions found; create single summary row
            # Try to get reporter name from summary table
            reporter_name = "-"
            for table in tables:
                all_text = table.get_text()
                if '보고자' in all_text and '이번보고서' in all_text:
                    rows = table.find_all('tr')
                    for row in rows:
                        cells = row.find_all(recursive=False)
                        cell_texts = [clean_text(c.get_text()) for c in cells]
                        if any('이번' in ct for ct in cell_texts):
                            for ct in cell_texts:
                                ct_clean = ct.strip()
                                if ct_clean and ct_clean not in ['-', '이번보고서'] and not ct_clean.replace('.', '').replace(' ', '').isdigit():
                                    pnum = parse_number(ct_clean)
                                    if pnum is None:
                                        reporter_name = ct_clean
                                        break
                            break

            # Get shares from summary (이번보고서 row)
            shares_after = None
            for table in tables:
                all_text = table.get_text()
                if '이번보고서' in all_text and '주식등의' in all_text:
                    rows = table.find_all('tr')
                    for row in rows:
                        cells = row.find_all(recursive=False)
                        cell_texts = [clean_text(c.get_text()) for c in cells]
                        if any('이번' in ct for ct in cell_texts):
                            # Skip first few cells (label, date, reporter name, count)
                            # Structure: [이번보고서, date, name, 특별관계자수, 주식등의수, 비율, 주식수, 비율, 발행주식총수]
                            for ct in cell_texts[4:]:  # Start after metadata cells
                                # Skip date-like strings
                                if '년' in ct or '월' in ct or '일' in ct:
                                    continue
                                n = parse_number(ct)
                                if n is not None and n > 1000:
                                    shares_after = int(n)
                                    break
                            break

            results.append({
                "reporter_name": reporter_name,
                "relationship": "보고자",
                "change_reason": change_reason,
                "shares_change": 0,
                "avg_price": None,
                "shares_after": shares_after or 0,
                "ownership_pct": overall_pct
            })

    return results

def parse_html_options(html_path):
    """Parses local HTML files to extract Call and Put option paragraphs."""
    if not os.path.exists(html_path):
        return "N/A (HTML 파일 없음)", "N/A (HTML 파일 없음)"
        
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
    except Exception as e:
        return f"N/A (파일 에러: {e})", f"N/A (파일 에러: {e})"
        
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n")
    
    # Process text lines
    lines = []
    for line in text.splitlines():
        line = re.sub(r'[ \t]+', ' ', line).strip()
        if line:
            lines.append(line)
    clean_text = "\n".join(lines)
    
    # Heuristics: find all section headers for Put Option
    put_headers = [
        r"(?:조기상환청구권|풋옵션)\s*(?:\([^)]*\))?\s*(?:에\s*관한\s*사항|:|：|\n)"
    ]
    
    put_text = "N/A"
    for pat in put_headers:
        matches = list(re.finditer(pat, clean_text, re.IGNORECASE))
        for m in matches:
            snippet = clean_text[m.start():m.start()+1500].strip()
            intro = clean_text[m.start():m.start()+100]
            if not ('참고' in intro and ('세부' in intro or '공시' in intro or '사항' in intro)):
                if len(snippet) > 1500:
                    snippet = snippet[:1500] + "..."
                put_text = snippet
                break
        if put_text != "N/A":
            break
            
    if put_text == "N/A":
        for kw in ["조기상환청구권", "조기상환청구", "풋옵션"]:
            pos = clean_text.find(kw)
            if pos != -1:
                snippet = clean_text[pos:pos+1500].strip()
                if len(snippet) > 1500:
                    snippet = snippet[:1500] + "..."
                put_text = snippet
                break

    # Heuristics: find all section headers for Call Option
    call_headers = [
        r"(?:매도청구권|콜옵션|사채매수선택권|매수청구권|발행회사의\s*매수청구권)\s*(?:\([^)]*\))?\s*(?:에\s*관한\s*사항|:|：|\n)"
    ]
    
    call_text = "N/A"
    for pat in call_headers:
        matches = list(re.finditer(pat, clean_text, re.IGNORECASE))
        for m in matches:
            snippet = clean_text[m.start():m.start()+1500].strip()
            intro = clean_text[m.start():m.start()+100]
            if not ('참고' in intro and ('세부' in intro or '공시' in intro or '사항' in intro)):
                if len(snippet) > 1500:
                    snippet = snippet[:1500] + "..."
                call_text = snippet
                break
        if call_text != "N/A":
            break
            
    if call_text == "N/A":
        for kw in ["사채매수선택권", "매도청구권", "콜옵션", "매수선택권"]:
            pos = clean_text.find(kw)
            if pos != -1:
                snippet = clean_text[pos:pos+1500].strip()
                if len(snippet) > 1500:
                    snippet = snippet[:1500] + "..."
                call_text = snippet
                break
                
    return call_text, put_text

def parse_mezzanine_html_fallback(html_path, m_type):
    """Parses mezzanine attributes directly from HTML table if API fails or for offline support."""
    res = {
        "total_amount": None,
        "coupon_rate": None,
        "yield_rate": None,
        "maturity_date": "-",
        "conversion_price": None,
        "claim_start": "-",
        "claim_end": "-",
        "share_type": "-"
    }
    if not os.path.exists(html_path):
        return res
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        current_primary_key = ""
        for table in soup.find_all('table'):
            in_claim_period = False
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    first_cell = cols[0].replace(" ", "")
                    if first_cell:
                        is_claim_key = any(k in first_cell for k in ["전환청구기간", "신주인수권행사기간", "교환청구기간", "행사청구기간", "전환청구일", "행사기간", "전환기간", "교환기간"])
                        is_sub_key = first_cell in ["시작일", "시작", "종료일", "종료", "시작일자", "종료일자", "행사시작일", "행사종료일", "전환시작일", "전환종료일"]
                        
                        if is_claim_key:
                            in_claim_period = True
                            current_primary_key = first_cell.replace("ㆍ", "").replace(".", "")
                        elif is_sub_key and in_claim_period:
                            # Keep current claim key as primary key
                            pass
                        else:
                            in_claim_period = False
                            current_primary_key = first_cell.replace("ㆍ", "").replace(".", "")
                    
                    key = current_primary_key
                    
                    if "사채의권면총액" in key or "권면총액" in key:
                        res["total_amount"] = clean_numeric(cols[-1])
                    if "표면이자율" in key:
                        res["coupon_rate"] = clean_numeric(cols[-1])
                    if "만기이자율" in key:
                        res["yield_rate"] = clean_numeric(cols[-1])
                    if "사채만기일" in key:
                        res["maturity_date"] = cols[-1]
                    if any(k in key for k in ["전환가액", "행사가액", "교환가액"]):
                        res["conversion_price"] = clean_numeric(cols[-1])
                        
                    # Parse Conversion Request Start & End dates
                    if any(k in key for k in ["전환청구기간", "신주인수권행사기간", "교환청구기간", "행사청구기간", "전환청구일", "행사기간", "전환기간", "교환기간"]):
                        row_text = "".join(cols).replace(" ", "")
                        if "시작" in row_text:
                            dates = extract_dates(cols[-1])
                            res["claim_start"] = dates[0] if dates else cols[-1]
                        elif "종료" in row_text:
                            dates = extract_dates(cols[-1])
                            res["claim_end"] = dates[0] if dates else cols[-1]
                        elif len(cols) == 2:
                            dates = extract_dates(cols[1])
                            if len(dates) >= 2:
                                res["claim_start"] = dates[0]
                                res["claim_end"] = dates[1]
                            elif len(dates) == 1:
                                res["claim_start"] = dates[0]
                                
                    if any(k in key for k in ["전환에따라발행할주식의종류", "신주인수권행사에따라발행할주식의종류", "교환대상주식의종류", "발행할주식의종류"]):
                        res["share_type"] = cols[-1]
                        
        for k in ["maturity_date", "claim_start", "claim_end"]:
            d_match = re.search(r'(\d{4})[-년\s\.]*(\d{1,2})[-월\s\.]*(\d{1,2})', res[k])
            if d_match:
                y, m, d = d_match.groups()
                res[k] = f"{y}-{int(m):02d}-{int(d):02d}"
    except Exception as e:
        logger.error(f"Error parsing fallback mezzanine HTML: {e}")
    return res

def parse_treasury_html_fallback(html_path, base_type):
    res = {
        "total_amount": None,
        "shares_count": None,
        "start_date": "-",
        "end_date": "-",
        "cancellation_date": "-",
        "method": "-",
        "broker": "-",
        "purpose": "-"
    }
    if not os.path.exists(html_path):
        return res
        
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        current_primary_key = ""
        in_period = False
        
        for table in soup.find_all('table'):
            in_period = False
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    first_cell = cols[0].replace(" ", "")
                    if first_cell:
                        is_period_key = any(k in first_cell for k in ["취득예상기간", "취득예정기간", "취득기간", "계약기간", "소각을위한자기주식취득예정기간"])
                        is_sub_key = first_cell in ["시작일", "시작", "종료일", "종료", "시작일자", "종료일자"]
                        
                        if is_period_key:
                            in_period = True
                            current_primary_key = first_cell.replace("ㆍ", "").replace(".", "")
                        elif is_sub_key and in_period:
                            pass
                        else:
                            in_period = False
                            current_primary_key = first_cell.replace("ㆍ", "").replace(".", "")
                            
                    key = current_primary_key
                    
                    if "취득예정금액" in key or "계약금액" in key or "소각예정금액" in key:
                        val = clean_numeric(cols[-1])
                        if val is not None:
                            res["total_amount"] = val
                    if any(k in key for k in ["취득예정주식", "소각할주식의종류와수", "소각할주식수", "소각예정주식"]):
                        val = clean_numeric(cols[-1])
                        if val is not None:
                            res["shares_count"] = val
                    if "취득방법" in key or "소각할주식의취득방법" in key:
                        res["method"] = cols[-1]
                    if "위탁투자" in key or "계약체결기관" in key or "자기주식취득위탁" in key:
                        res["broker"] = cols[-1]
                    if "취득목적" in key or "계약목적" in key or "소각목적" in key or "목적" in key:
                        res["purpose"] = cols[-1]
                    if "소각예정일" in key or "소각일" in key:
                        res["cancellation_date"] = cols[-1]
                        
                    # Period processing
                    if in_period:
                        row_text = "".join(cols).replace(" ", "")
                        if "시작" in row_text:
                            dates = extract_dates(cols[-1])
                            res["start_date"] = dates[0] if dates else cols[-1]
                        elif "종료" in row_text:
                            dates = extract_dates(cols[-1])
                            res["end_date"] = dates[0] if dates else cols[-1]
                        elif len(cols) == 2:
                            dates = extract_dates(cols[1])
                            if len(dates) >= 2:
                                res["start_date"] = dates[0]
                                res["end_date"] = dates[1]
                            elif len(dates) == 1:
                                res["start_date"] = dates[0]
                                
        # Date normalization
        for k in ["start_date", "end_date", "cancellation_date"]:
            if res[k] and res[k] != "-":
                d_match = re.search(r'(\d{4})[-년\s\.]*(\d{1,2})[-월\s\.]*(\d{1,2})', res[k])
                if d_match:
                    y, m, d = d_match.groups()
                    res[k] = f"{y}-{int(m):02d}-{int(d):02d}"
                    
        # Apply helpful defaults for cancellation purpose
        if base_type == "자기주식소각" and res["purpose"] == "-":
            res["purpose"] = "주주가치 제고 및 주식소각"
    except Exception as e:
        logger.error(f"Error parsing fallback treasury HTML: {e}")
    return res

def fetch_mezzanine_details(api_key, corp_code, date, rcept_no, report_nm):
    """Queries OpenDART to get structured mezzanine details (CB, BW, EB)."""
    nm = report_nm.replace(" ", "")
    if "전환사채" in nm:
        url = "https://opendart.fss.or.kr/api/cvbdIsDecsn.json"
        m_type = "CB"
    elif "신주인수권부사채" in nm:
        url = "https://opendart.fss.or.kr/api/bdwtIsDecsn.json"
        m_type = "BW"
    elif "교환사채" in nm:
        url = "https://opendart.fss.or.kr/api/exbdIsDecsn.json"
        m_type = "EB"
    else:
        return None, None
        
    if not api_key:
        return None, m_type
        
    params = {
        'crtfc_key': api_key,
        'corp_code': corp_code,
        'bgn_de': date,
        'end_de': date
    }
    
    try:
        time.sleep(0.2)
        response = requests.get(url, params=params, timeout=3)
        if response.status_code != 200:
            return None, m_type
            
        data = response.json()
        if data.get("status") != "000":
            prev_date = (datetime.datetime.strptime(date, "%Y%m%d") - datetime.timedelta(days=1)).strftime("%Y%m%d")
            params['bgn_de'] = prev_date
            time.sleep(0.2)
            response = requests.get(url, params=params, timeout=3)
            if response.status_code == 200:
                data = response.json()
                
        if data.get("status") == "000":
            item_list = data.get("list", [])
            for item in item_list:
                if str(item.get("rcept_no")) == str(rcept_no):
                    return item, m_type
            if item_list:
                return item_list[0], m_type
    except Exception as e:
        logger.error(f"Mezzanine API exception: {e}")
        
    return None, m_type

def parse_contract_html(html_path):
    """Parses local supply contract HTML files to extract required fields."""
    res = {
        "content": "-",
        "start_date": "-",
        "end_date": "-",
        "amount": None,
        "backlog": "-",
        "ratio": None,
        "counterparty": "-"
    }
    if not os.path.exists(html_path):
        return res
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        for table in soup.find_all('table'):
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    key = cols[0].replace(" ", "").replace("ㆍ", "").replace(".", "")
                    
                    if any(k in key for k in ["판매공급계약내용", "체결계약명", "계약내용"]):
                        res["content"] = cols[1]
                    elif len(cols) >= 3 and "판매공급계약내용" in cols[1].replace(" ", ""):
                        res["content"] = cols[2]
                        
                    for idx, val in enumerate(cols):
                        val_clean = val.replace(" ", "")
                        if "시작일" in val_clean and idx + 1 < len(cols):
                            res["start_date"] = cols[idx+1]
                        if "종료일" in val_clean and idx + 1 < len(cols):
                            res["end_date"] = cols[idx+1]
                            
                    if any(k in key for k in ["확정계약금액", "계약금액총액", "계약금액(원)", "계약금액총액(원)"]):
                        res["amount"] = clean_numeric(cols[1])
                    elif len(cols) >= 3 and any(k in cols[1].replace(" ", "") for k in ["확정계약금액", "계약금액총액", "계약금액(원)"]):
                        res["amount"] = clean_numeric(cols[2])
                        
                    if any(k in key for k in ["매출액대비(%)", "매출액대비"]):
                        res["ratio"] = clean_numeric(cols[1])
                    elif len(cols) >= 3 and any(k in cols[1].replace(" ", "") for k in ["매출액대비(%)", "매출액대비"]):
                        res["ratio"] = clean_numeric(cols[2])
                        
                    if any(k in key for k in ["계약상대방", "계약상대"]):
                        res["counterparty"] = cols[1]
                    elif len(cols) >= 3 and any(k in cols[1].replace(" ", "") for k in ["계약상대방", "계약상대"]):
                        res["counterparty"] = cols[2]
                        
                    # Backlog check if explicitly in text
                    if "수주잔고" in key:
                        res["backlog"] = cols[1]
                        
        for k in ["start_date", "end_date"]:
            d_match = re.search(r'(\d{4})[-년\s]*(\d{1,2})[-월\s]*(\d{1,2})', res[k])
            if d_match:
                y, m, d = d_match.groups()
                res[k] = f"{y}-{int(m):02d}-{int(d):02d}"
    except Exception as e:
        logger.error(f"Error parsing contract HTML: {e}")
    return res

def parse_capital_increase_html(html_path):
    """Parses local capital increase HTML files to extract required fields."""
    res = {
        "fundraising_amount": 0,
        "issue_price": None,
        "share_type": "보통주식",
        "new_shares_count": 0,
        "payment_date": "-",
        "listing_date": "-",
        "existing_shares": 0,
        "ratio": None,
        "purpose": "-"
    }
    if not os.path.exists(html_path):
        return res
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        purposes = {}
        new_shares = 0
        existing_shares = 0
        
        for table in soup.find_all('table'):
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    key = cols[0].replace(" ", "").replace("ㆍ", "").replace(".", "")
                    
                    if "신주의종류와수" in key:
                        val = clean_numeric(cols[-1])
                        if isinstance(val, (int, float)):
                            new_shares += val
                            if "보통" in cols[1]:
                                res["share_type"] = "보통주식"
                            elif "기타" in cols[1]:
                                res["share_type"] = "기타주식"
                    elif "보통주식(주)" in key and new_shares == 0:
                        val = clean_numeric(cols[-1])
                        if isinstance(val, (int, float)):
                            new_shares = val
                            res["share_type"] = "보통주식"
                            
                    if "증자전발행주식총수" in key or "증자전발행주식총수(주)" in key:
                        val = clean_numeric(cols[-1])
                        if isinstance(val, (int, float)):
                            existing_shares = val
                    elif "보통주식(주)" in key and existing_shares == 0 and "증자전" in key:
                        val = clean_numeric(cols[-1])
                        if isinstance(val, (int, float)):
                            existing_shares = val
                            
                    if "신주발행가액" in key:
                        res["issue_price"] = clean_numeric(cols[-1])
                        
                    if "납입일" in key:
                        res["payment_date"] = cols[-1]
                    if "신주의상장예정일" in key or "상장예정일" in key:
                        res["listing_date"] = cols[-1]
                        
                    for p_key in ["시설자금", "영업양수자금", "운영자금", "채무상환자금", "타법인증권취득자금", "기타자금"]:
                        if p_key in cols[0].replace(" ", "") or (len(cols) >= 2 and p_key in cols[1].replace(" ", "")):
                            val = clean_numeric(cols[-1])
                            if isinstance(val, (int, float)) and val > 0:
                                purposes[p_key] = val
                                
        if purposes:
            res["fundraising_amount"] = sum(purposes.values())
            purpose_strs = []
            for k, v in purposes.items():
                if v >= 100000000:
                    purpose_strs.append(f"{k} ({v/100000000:.1f}억원)")
                else:
                    purpose_strs.append(f"{k} ({v:,.0f}원)")
            res["purpose"] = ", ".join(purpose_strs)
            
        if new_shares > 0:
            res["new_shares_count"] = new_shares
        if existing_shares > 0:
            res["existing_shares"] = existing_shares
            res["ratio"] = (new_shares / existing_shares) * 100
            
        for k in ["payment_date", "listing_date"]:
            d_match = re.search(r'(\d{4})[-년\s]*(\d{1,2})[-월\s]*(\d{1,2})', res[k])
            if d_match:
                y, m, d = d_match.groups()
                res[k] = f"{y}-{int(m):02d}-{int(d):02d}"
    except Exception as e:
        logger.error(f"Error parsing capital increase HTML: {e}")
    return res

def parse_facility_investment_html(html_path):
    """Parses local facility investment HTML files to extract required fields."""
    res = {
        "purpose": "-",
        "start_date": "-",
        "end_date": "-",
        "amount": None,
        "ratio": None
    }
    if not os.path.exists(html_path):
        return res
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        for table in soup.find_all('table'):
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    key = cols[0].replace(" ", "").replace("ㆍ", "").replace(".", "")
                    
                    if "투자목적" in key:
                        res["purpose"] = cols[1]
                    elif len(cols) >= 3 and "투자목적" in cols[1].replace(" ", ""):
                        res["purpose"] = cols[2]
                        
                    for idx, val in enumerate(cols):
                        val_clean = val.replace(" ", "")
                        if "시작일" in val_clean and idx + 1 < len(cols):
                            res["start_date"] = cols[idx+1]
                        if "종료일" in val_clean and idx + 1 < len(cols):
                            res["end_date"] = cols[idx+1]
                            
                    if "투자금액(원)" in key:
                        res["amount"] = clean_numeric(cols[1])
                    elif len(cols) >= 3 and "투자금액(원)" in cols[1].replace(" ", ""):
                        res["amount"] = clean_numeric(cols[2])
                        
                    if "자기자본대비(%)" in key:
                        res["ratio"] = clean_numeric(cols[1])
                    elif len(cols) >= 3 and "자기자본대비(%)" in cols[1].replace(" ", ""):
                        res["ratio"] = clean_numeric(cols[2])
                        
        for k in ["start_date", "end_date"]:
            d_match = re.search(r'(\d{4})[-년\s]*(\d{1,2})[-월\s]*(\d{1,2})', res[k])
            if d_match:
                y, m, d = d_match.groups()
                res[k] = f"{y}-{int(m):02d}-{int(d):02d}"
    except Exception as e:
        logger.error(f"Error parsing facility investment HTML: {e}")
    return res

def parse_guarantee_loan_html(html_path):
    """Parses local debt guarantee / monetary loan HTML files to extract required fields."""
    res = {
        "type": "기타",
        "counterparty": "-",
        "amount": None,
        "ratio": None,
        "start_date": "-",
        "end_date": "-",
        "purpose": "-"
    }
    if not os.path.exists(html_path):
        return res
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            
        for table in soup.find_all('table'):
            for r in table.find_all('tr'):
                children = r.find_all(recursive=False)
                cols = [c.get_text().strip().replace('\n', ' ') for c in children]
                cols = [re.sub(r'\s+', ' ', col) for col in cols]
                if len(cols) >= 2:
                    key = cols[0].replace(" ", "").replace("ㆍ", "").replace(".", "")
                    
                    if "채무보증" in html_path or "보증" in html_path:
                        res["type"] = "채무보증"
                    elif "금전대여" in html_path or "대여" in html_path:
                        res["type"] = "금전대여"
                        
                    if any(k in key for k in ["채무자", "대여상대", "거래상대방", "보증대상"]):
                        res["counterparty"] = cols[1]
                    elif len(cols) >= 3 and any(k in cols[1].replace(" ", "") for k in ["채무자", "대여상대", "거래상대방"]):
                        res["counterparty"] = cols[2]
                        
                    if any(k in key for k in ["채무보증금액(원)", "대여금액(원)", "보증금액", "대여금액"]):
                        res["amount"] = clean_numeric(cols[1])
                    elif len(cols) >= 3 and any(k in cols[1].replace(" ", "") for k in ["채무보증금액(원)", "대여금액(원)"]):
                        res["amount"] = clean_numeric(cols[2])
                        
                    if "자기자본대비(%)" in key:
                        res["ratio"] = clean_numeric(cols[1])
                    elif len(cols) >= 3 and "자기자본대비(%)" in cols[1].replace(" ", ""):
                        res["ratio"] = clean_numeric(cols[2])
                        
                    for idx, val in enumerate(cols):
                        val_clean = val.replace(" ", "")
                        if "시작일" in val_clean and idx + 1 < len(cols):
                            res["start_date"] = cols[idx+1]
                        if "종료일" in val_clean and idx + 1 < len(cols):
                            res["end_date"] = cols[idx+1]
                            
                    if any(k in key for k in ["보증목적", "대여목적", "자금대여목적", "대여용도"]):
                        res["purpose"] = cols[1]
                        
        for k in ["start_date", "end_date"]:
            d_match = re.search(r'(\d{4})[-년\s]*(\d{1,2})[-월\s]*(\d{1,2})', res[k])
            if d_match:
                y, m, d = d_match.groups()
                res[k] = f"{y}-{int(m):02d}-{int(d):02d}"
    except Exception as e:
        logger.error(f"Error parsing guarantee/loan HTML: {e}")
    return res

def send_telegram_document(token, chat_id, file_path, caption=None):
    """Sends a document via Telegram Bot API."""
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    try:
        with open(file_path, "rb") as f:
            files = {"document": f}
            data = {"chat_id": chat_id}
            if caption:
                data["caption"] = caption
            r = requests.post(url, data=data, files=files, timeout=45)
            if r.status_code == 200:
                logger.info(f"Telegram document sent successfully to chat {chat_id}")
                return True
            else:
                logger.error(f"Failed to send telegram document: HTTP {r.status_code}, {r.text}")
                return False
    except Exception as e:
        logger.error(f"Failed to send telegram document: {e}")
        return False

def send_telegram_message(token, chat_id, text, max_retries=3):
    """Sends a plain text message via Telegram Bot API with rate-limit retry."""
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text
    }
    for attempt in range(max_retries + 1):
        try:
            r = requests.post(url, json=payload, timeout=15)
            if r.status_code == 200:
                logger.info(f"Telegram text message sent successfully to chat {chat_id}")
                return True
            elif r.status_code == 429:
                retry_after = r.json().get("parameters", {}).get("retry_after", 5)
                logger.warning(f"Rate limited (429). Retrying after {retry_after}s (attempt {attempt+1}/{max_retries})")
                if attempt < max_retries:
                    time.sleep(retry_after + 1)
                    continue
                else:
                    logger.error(f"Rate limit exceeded after {max_retries} retries.")
            else:
                logger.error(f"Failed to send telegram message: HTTP {r.status_code}, {r.text}")
        except Exception as e:
            logger.error(f"Failed to send telegram message: {e}")
        break
    return False

def build_excel_summary(workspace_dir):
    """Scans disclosures, aggregates them, processes mezzanine/contracts/investments option info, resolves amendments, creates formatted Excel."""
    logger.info("Scanning disclosures in data_dart directory...")
    json_paths = sorted(glob.glob(os.path.join(workspace_dir, "data_dart", "202*", "disclosures.json")))
    
    if not json_paths:
        logger.error("No disclosures.json files found under data_dart/202*/")
        return False
        
    all_disclosures = []
    for p in json_paths:
        date_str = os.path.basename(os.path.dirname(p))
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
                for item in data:
                    item['collected_date'] = date_str
                    all_disclosures.append(item)
        except Exception as e:
            logger.error(f"Error reading {p}: {e}")
            
    logger.info(f"Total raw disclosures loaded: {len(all_disclosures)}")
    if not all_disclosures:
        return False
        
    df_all = pd.DataFrame(all_disclosures)
    df_all = df_all.drop_duplicates(subset=['rcept_no'])
    logger.info(f"Total unique disclosures: {len(df_all)}")
    
    # Classify disclosures
    df_all['category'] = df_all['report_nm'].apply(classify_disclosure)
    df_all['base_type'] = df_all['report_nm'].apply(identify_base_report_type)
    
    # Load cache
    cache = load_cache()
    
    # We will process each disclosure, fetch API and parse HTML, saving all detailed properties.
    parsed_records = []
    
    total_docs = len(df_all)
    logger.info(f"Parsing details for {total_docs} disclosures...")
    
    for idx, row in df_all.iterrows():
        rcept_no = str(row['rcept_no'])
        corp_code = str(row['corp_code'])
        date = str(row['rcept_dt'])
        base_type = row['base_type']
        collected_date = row['collected_date']
        report_nm = row['report_nm']
        
        # Check cache
        if rcept_no in cache:
            record_detail = cache[rcept_no]
            
            # Auto-healing cache: if it is a treasury share disclosure but is classified as "기타" or "기타공시",
            # or has empty data dictionary, re-parse and update it.
            is_treasury = base_type in ["자기주식취득", "자기주식신탁", "자기주식소각"]
            cached_is_other = record_detail.get("base_type") == "기타" or record_detail.get("category") == "기타공시"
            
            if is_treasury and (cached_is_other or not record_detail.get("data")):
                record_detail["category"] = "재무_자기주식"
                record_detail["base_type"] = base_type
                html_path = os.path.join(workspace_dir, "data_dart", collected_date, f"{rcept_no}.html")
                t_data = parse_treasury_html_fallback(html_path, base_type)
                record_detail["data"] = {
                    "total_amount": t_data["total_amount"],
                    "shares_count": t_data["shares_count"],
                    "start_date": t_data["start_date"],
                    "end_date": t_data["end_date"],
                    "cancellation_date": t_data["cancellation_date"],
                    "method": t_data["method"],
                    "broker": t_data["broker"],
                    "purpose": t_data["purpose"]
                }
                cache[rcept_no] = record_detail
                
            # Re-run option date extraction to apply the new request start date logic
            if base_type in ["CB", "BW", "EB"]:
                html_path = os.path.join(workspace_dir, "data_dart", collected_date, f"{rcept_no}.html")
                call_opt, put_opt = parse_html_options(html_path)
                call_start = extract_option_start_date(call_opt, "call")
                put_start = extract_option_start_date(put_opt, "put")
                
                # Parse fallback claim dates in case they are missing from cache
                fallback = parse_mezzanine_html_fallback(html_path, base_type)
                
                if "data" in record_detail:
                    record_detail["data"]["call_start"] = call_start
                    record_detail["data"]["put_start"] = put_start
                    record_detail["data"]["call_option_info"] = call_opt
                    record_detail["data"]["put_option_info"] = put_opt
                    
                    # Fill claim_start and claim_end
                    fb_start = fallback.get("claim_start", "-")
                    fb_end = fallback.get("claim_end", "-")
                    
                    if fb_start and fb_start != "-":
                        record_detail["data"]["claim_start"] = fb_start
                    else:
                        record_detail["data"]["claim_start"] = record_detail["data"].get("claim_start", "-")
                        
                    if fb_end and fb_end != "-":
                        record_detail["data"]["claim_end"] = fb_end
                    else:
                        record_detail["data"]["claim_end"] = record_detail["data"].get("claim_end", "-")
                        
                    record_detail["data"]["listing_date"] = "-"
            parsed_records.append(record_detail)
            continue
            
        html_path = os.path.join(workspace_dir, "data_dart", collected_date, f"{rcept_no}.html")
        
        # Self-healing download if HTML is missing (disabled to run using already-downloaded files)
        if False:
            pass
                
        # Determine original submission date if it's an amendment
        is_amended = any(k in report_nm for k in ["[기재정정]", "[첨부정정]", "정정보고서", "정정공시"])
        original_date = None
        if is_amended:
            original_date = find_original_date_from_html(html_path)
            
        record_detail = {
            "rcept_dt": date,
            "corp_name": row['corp_name'],
            "corp_code": corp_code,
            "corp_cls": row['corp_cls'],
            "stock_code": row['stock_code'],
            "report_nm": report_nm,
            "flr_nm": row['flr_nm'],
            "rcept_no": rcept_no,
            "category": row['category'],
            "base_type": base_type,
            "collected_date": collected_date,
            "is_amended": is_amended,
            "original_date": original_date,
            "merged_into": None, # Will be set during merging
            "data": {} # Specific parsed columns
        }
        
        # 1. Mezzanine detailed parsing (CB, BW, EB)
        if base_type in ["CB", "BW", "EB"]:
            # API
            details, m_type = fetch_mezzanine_details(DART_API_KEY, corp_code, date, rcept_no, report_nm)
            # HTML Falling back + Options Text
            call_opt, put_opt = parse_html_options(html_path)
            fallback = parse_mezzanine_html_fallback(html_path, m_type)
            
            # Extract details
            total_amount = None
            coupon_rate = None
            yield_rate = None
            maturity_date = "-"
            conversion_price = None
            claim_start = "-"
            claim_end = "-"
            share_type = "-"
            
            if details:
                total_amount = clean_numeric(details.get("bd_fta"))
                coupon_rate = clean_numeric(details.get("bd_intr_ex"))
                yield_rate = clean_numeric(details.get("bd_intr_sf"))
                maturity_date = details.get("bd_mtd", "-")
                if m_type == "CB":
                    conversion_price = clean_numeric(details.get("cv_prc"))
                    claim_start = details.get("cvrqpd_bgd", "-")
                    claim_end = details.get("cvrqpd_edd", "-")
                else:
                    conversion_price = clean_numeric(details.get("ex_prc"))
                    claim_start = details.get("exrqpd_bgd", "-")
                    claim_end = details.get("exrqpd_edd", "-")
                share_type = details.get("cvisstk_knd") or details.get("extg") or "-"
            else:
                # Use fallback HTML parsed details
                total_amount = fallback["total_amount"]
                coupon_rate = fallback["coupon_rate"]
                yield_rate = fallback["yield_rate"]
                maturity_date = fallback["maturity_date"]
                conversion_price = fallback["conversion_price"]
                claim_start = fallback["claim_start"]
                claim_end = fallback["claim_end"]
                share_type = fallback["share_type"]
                
            # Extract Option Dates
            call_start = extract_option_start_date(call_opt, "call")
            put_start = extract_option_start_date(put_opt, "put")
            
            record_detail["data"] = {
                "total_amount": total_amount,
                "price": conversion_price,
                "share_type": share_type,
                "new_shares_count": clean_numeric(details.get("cvisstk_cnt")) if details else None,
                "payment_date": details.get("pymd", "-") if details else "-",
                "listing_date": "-",
                "claim_start": claim_start,
                "claim_end": claim_end,
                "ratio": clean_numeric(details.get("cvisstk_tisstk_vs")) if details else None,
                "purpose": "-", # Summarize later
                "call_start": call_start,
                "put_start": put_start,
                "call_option_info": call_opt,
                "put_option_info": put_opt
            }
            
        # 2. Capital Increase Parsing (유상증자)
        elif base_type == "유상증자":
            cap_data = parse_capital_increase_html(html_path)
            record_detail["data"] = {
                "total_amount": cap_data["fundraising_amount"],
                "price": cap_data["issue_price"],
                "share_type": cap_data["share_type"],
                "new_shares_count": cap_data["new_shares_count"],
                "payment_date": cap_data["payment_date"],
                "listing_date": cap_data["listing_date"],
                "ratio": cap_data["ratio"],
                "purpose": cap_data["purpose"],
                "call_start": "-",
                "put_start": "-"
            }
            
        # 3. Supply Contract Parsing (공급계약)
        elif base_type == "공급계약":
            con_data = parse_contract_html(html_path)
            record_detail["data"] = {
                "content": con_data["content"],
                "start_date": con_data["start_date"],
                "end_date": con_data["end_date"],
                "amount": con_data["amount"],
                "backlog": con_data["backlog"],
                "ratio": con_data["ratio"],
                "counterparty": con_data["counterparty"]
            }
            
        # 4. Facility Investment Parsing (신규시설투자)
        elif base_type == "시설투자":
            f_data = parse_facility_investment_html(html_path)
            record_detail["data"] = {
                "purpose": f_data["purpose"],
                "start_date": f_data["start_date"],
                "end_date": f_data["end_date"],
                "amount": f_data["amount"],
                "ratio": f_data["ratio"]
            }
            
        # 5. Guarantee / Loan Parsing (재무_채무보증)
        elif base_type in ["채무보증", "금전대여"]:
            g_data = parse_guarantee_loan_html(html_path)
            record_detail["data"] = {
                "type": g_data["type"],
                "counterparty": g_data["counterparty"],
                "amount": g_data["amount"],
                "ratio": g_data["ratio"],
                "start_date": g_data["start_date"],
                "end_date": g_data["end_date"],
                "purpose": g_data["purpose"]
            }
            
        # 6. Treasury Shares Parsing (자기주식 취득/신탁/소각)
        elif base_type in ["자기주식취득", "자기주식신탁", "자기주식소각"]:
            t_data = parse_treasury_html_fallback(html_path, base_type)
            record_detail["data"] = {
                "total_amount": t_data["total_amount"],
                "shares_count": t_data["shares_count"],
                "start_date": t_data["start_date"],
                "end_date": t_data["end_date"],
                "cancellation_date": t_data["cancellation_date"],
                "method": t_data["method"],
                "broker": t_data["broker"],
                "purpose": t_data["purpose"]
            }
            
        # Cache and save
        cache[rcept_no] = record_detail
        parsed_records.append(record_detail)
        
    save_cache(cache)
    
    # -------------------------------------------------------------
    # Resolving Amendments ('정정' 공시 연동 및 수정 처리)
    # -------------------------------------------------------------
    logger.info("Resolving and merging amendment ('정정') disclosures...")
    
    # Sort records chronologically so that amendments filed later overwrite earlier ones correctly
    parsed_records = sorted(parsed_records, key=lambda x: x['rcept_dt'])
    
    for idx, record in enumerate(parsed_records):
        if record["is_amended"] and record["original_date"]:
            orig_date = record["original_date"]
            corp_code = record["corp_code"]
            base_type = record["base_type"]
            
            # Find the original disclosure: same company, same base type, filed on original_date
            # Or if original_date is slightly off, we check nearby.
            original_match = None
            for prev_record in parsed_records[:idx]:
                if prev_record["corp_code"] == corp_code and prev_record["base_type"] == base_type:
                    # Match date
                    if prev_record["rcept_dt"] == orig_date:
                        original_match = prev_record
                        break
            
            # If still not matched, fallback to matching the most recent report of same type
            if not original_match:
                for prev_record in reversed(parsed_records[:idx]):
                    if prev_record["corp_code"] == corp_code and prev_record["base_type"] == base_type:
                        # Ensure it's not already merged
                        if not prev_record["merged_into"]:
                            original_match = prev_record
                            break
                            
            if original_match:
                logger.info(f"Merging amendment {record['rcept_no']} into original {original_match['rcept_no']}")
                # Overwrite original data with corrected amendment data
                original_match["data"] = record["data"]
                original_match["report_nm"] = f"{original_match['report_nm']} (정정: {record['rcept_dt']})"
                
                # Update date display to show amendment history
                orig_date_fmt = f"{original_match['rcept_dt'][:4]}-{original_match['rcept_dt'][4:6]}-{original_match['rcept_dt'][6:]}"
                amend_date_fmt = f"{record['rcept_dt'][:4]}-{record['rcept_dt'][4:6]}-{record['rcept_dt'][6:]}"
                original_match["rcept_dt_display"] = f"{orig_date_fmt} (정정: {amend_date_fmt})"
                
                # Overwrite receipt number to the latest one so that link points to the corrected version
                original_match["rcept_no"] = record["rcept_no"]
                original_match["collected_date"] = record["collected_date"]
                
                # Mark amendment as merged so we exclude it from the final sheet rows (avoid double counts)
                record["merged_into"] = original_match["rcept_no"]
            else:
                # If original not found in current dataset (e.g. original was 2025), format its date showing it is an amendment
                orig_date_fmt = f"{orig_date[:4]}-{orig_date[4:6]}-{orig_date[6:]}" if orig_date else "과거공시"
                amend_date_fmt = f"{record['rcept_dt'][:4]}-{record['rcept_dt'][4:6]}-{record['rcept_dt'][6:]}"
                record["rcept_dt_display"] = f"{orig_date_fmt} (정정: {amend_date_fmt})"
                
    # Filter out merged amendments
    active_records = [r for r in parsed_records if r["merged_into"] is None]
    logger.info(f"Filtered {len(parsed_records) - len(active_records)} merged amendment disclosures.")
    
    # Map corp_cls
    def map_market(cls):
        return {"Y": "코스피", "K": "코스닥", "N": "코넥스"}.get(cls, "기타")
        
    # Helper to clean date strings for display
    def fmt_date(dt_str):
        if len(dt_str) == 8:
            return f"{dt_str[:4]}-{dt_str[4:6]}-{dt_str[6:]}"
        return dt_str

    # -------------------------------------------------------------
    # Build Excel Workbook
    # -------------------------------------------------------------
    today_str = datetime.datetime.now().strftime('%Y%m%d')
    excel_path = os.path.join(workspace_dir, "data_dart", f"dart_disclosures_summary_{today_str}.xlsx")
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        
        # (Sheets for 정기공시 and 지분공시 have been removed as requested)
        
        # Sheet 3: 자금조달_증자 (Contains detailed columns for Capital Increases & Mezzanine CB/BW/EB)
        fund_rows = [r for r in active_records if r["category"] == "자금조달_증자"]
        fund_data_list = []
        for r in fund_rows:
            d = r["data"]
            u_type = r["base_type"]
            # Map type display name
            if u_type == "CB": type_disp = "전환사채(CB)"
            elif u_type == "BW": type_disp = "신주인수권부사채(BW)"
            elif u_type == "EB": type_disp = "교환사채(EB)"
            elif u_type == "유상증자": type_disp = "유상증자"
            elif u_type == "무상증자": type_disp = "무상증자"
            else: type_disp = u_type
            
            fund_data_list.append({
                "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
                "회사명": r["corp_name"],
                "시장구분": map_market(r["corp_cls"]),
                "종목코드": r["stock_code"],
                "유형": type_disp,
                "공시명": r["report_nm"],
                "조달금액": d.get("total_amount"),
                "발행가/전환가": d.get("price"),
                "신주종류": d.get("share_type", "-"),
                "신주발행수": d.get("new_shares_count"),
                "납입일": fmt_date(str(d.get("payment_date", "-"))),
                "상장예정일": fmt_date(str(d.get("listing_date", "-"))) if r["base_type"] in ["유상증자", "무상증자"] else "-",
                "전환청구시작일": fmt_date(str(d.get("claim_start", "-"))),
                "전환청구종료일": fmt_date(str(d.get("claim_end", "-"))),
                "주식총수대비": d.get("ratio"),
                "조달목적": d.get("purpose", "-"),
                "콜옵션 청구일": d.get("call_start", "-"),
                "풋옵션 청구일": d.get("put_start", "-"),
                "접수번호": r["rcept_no"],
                "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
            })
        df_fund = pd.DataFrame(fund_data_list)
        df_fund.to_excel(writer, sheet_name="자금조달_증자", index=False)
        format_fundraising_sheet(writer.sheets["자금조달_증자"])
        
        # Sheet 4: 영업활동_계약 (Contains supply contract detailed columns)
        contract_rows = [r for r in active_records if r["category"] == "영업활동_계약"]
        contract_data_list = []
        for r in contract_rows:
            d = r["data"]
            contract_data_list.append({
                "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
                "회사명": r["corp_name"],
                "시장구분": map_market(r["corp_cls"]),
                "종목코드": r["stock_code"],
                "유형": "단일판매공급계약체결" if r["base_type"] == "공급계약" else r["base_type"],
                "공시명": r["report_nm"],
                "계약내용": d.get("content", "-") if r["base_type"] == "공급계약" else "-",
                "시작일": fmt_date(str(d.get("start_date", "-"))) if r["base_type"] == "공급계약" else "-",
                "종료일": fmt_date(str(d.get("end_date", "-"))) if r["base_type"] == "공급계약" else "-",
                "수주금액": d.get("amount") if r["base_type"] == "공급계약" else None,
                "수주잔고": d.get("backlog", "-") if r["base_type"] == "공급계약" else "-",
                "최근 매출액 대비": d.get("ratio") if r["base_type"] == "공급계약" else None,
                "수주상대방": d.get("counterparty", "-") if r["base_type"] == "공급계약" else "-",
                "접수번호": r["rcept_no"],
                "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
            })
        df_contract = pd.DataFrame(contract_data_list)
        df_contract.to_excel(writer, sheet_name="영업활동_계약", index=False)
        format_contract_sheet(writer.sheets["영업활동_계약"])
        
        # Sheet 5: 신규시설투자 (Dedicated sheet)
        facility_rows = [r for r in active_records if r["category"] == "신규시설투자"]
        facility_data_list = []
        for r in facility_rows:
            d = r["data"]
            facility_data_list.append({
                "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
                "회사명": r["corp_name"],
                "시장구분": map_market(r["corp_cls"]),
                "종목코드": r["stock_code"],
                "공시명": r["report_nm"],
                "투자목적": d.get("purpose", "-"),
                "시작일": fmt_date(str(d.get("start_date", "-"))),
                "종료일": fmt_date(str(d.get("end_date", "-"))),
                "투자금액": d.get("amount"),
                "자기자본대비": d.get("ratio"),
                "접수번호": r["rcept_no"],
                "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
            })
        df_facility = pd.DataFrame(facility_data_list)
        df_facility.to_excel(writer, sheet_name="신규시설투자", index=False)
        format_facility_sheet(writer.sheets["신규시설투자"])
        
        # Sheet 6: 재무_채무보증 (Guarantees & Loans columns)
        fin_rows = [r for r in active_records if r["category"] == "재무_채무보증"]
        fin_data_list = []
        for r in fin_rows:
            d = r["data"]
            f_type = r["base_type"]
            type_lbl = "채무보증" if f_type == "채무보증" else ("금전대여" if f_type == "금전대여" else "기타")
            
            fin_data_list.append({
                "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
                "회사명": r["corp_name"],
                "시장구분": map_market(r["corp_cls"]),
                "종목코드": r["stock_code"],
                "공시명": r["report_nm"],
                "유형": type_lbl,
                "보증/대여상대방": d.get("counterparty", "-") if f_type in ["채무보증", "금전대여"] else "-",
                "금액": d.get("amount") if f_type in ["채무보증", "금전대여"] else None,
                "자기자본대비": d.get("ratio") if f_type in ["채무보증", "금전대여"] else None,
                "시작일": fmt_date(str(d.get("start_date", "-"))) if f_type in ["채무보증", "금전대여"] else "-",
                "종료일": fmt_date(str(d.get("end_date", "-"))) if f_type in ["채무보증", "금전대여"] else "-",
                "목적/용도": d.get("purpose", "-") if f_type in ["채무보증", "금전대여"] else "-",
                "접수번호": r["rcept_no"],
                "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
            })
        df_fin = pd.DataFrame(fin_data_list)
        df_fin.to_excel(writer, sheet_name="재무_채무보증", index=False)
        format_financial_sheet(writer.sheets["재무_채무보증"])
        
        # Sheet 6-2: 재무_자기주식 (Treasury Shares columns)
        treasury_rows = [r for r in active_records if r["category"] == "재무_자기주식"]
        treasury_data_list = []
        for r in treasury_rows:
            d = r["data"]
            f_type = r["base_type"]
            if f_type == "자기주식취득": type_lbl = "자기주식취득결정"
            elif f_type == "자기주식신탁": type_lbl = "신탁계약체결결정"
            elif f_type == "자기주식소각": type_lbl = "자기주식소각결정"
            else: type_lbl = f_type
            
            treasury_data_list.append({
                "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
                "회사명": r["corp_name"],
                "시장구분": map_market(r["corp_cls"]),
                "종목코드": r["stock_code"],
                "공시명": r["report_nm"],
                "유형": type_lbl,
                "예정금액": d.get("total_amount"),
                "예정주식수": d.get("shares_count"),
                "취득방법": d.get("method", "-"),
                "시작일": fmt_date(str(d.get("start_date", "-"))),
                "종료일": fmt_date(str(d.get("end_date", "-"))),
                "소각예정일": fmt_date(str(d.get("cancellation_date", "-"))),
                "위탁투자업자/수탁기관": d.get("broker", "-"),
                "목적": d.get("purpose", "-"),
                "접수번호": r["rcept_no"],
                "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
            })
        df_treasury = pd.DataFrame(treasury_data_list)
        df_treasury.to_excel(writer, sheet_name="재무_자기주식", index=False)
        format_treasury_sheet(writer.sheets["재무_자기주식"])
        
        # Sheet 7: 경영권_지배구조
        gov_rows = [r for r in active_records if r["category"] == "경영권_지배구조"]
        df_gov = pd.DataFrame([{
            "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
            "회사명": r["corp_name"],
            "시장구분": map_market(r["corp_cls"]),
            "종목코드": r["stock_code"],
            "공시명": r["report_nm"],
            "제출인": r["flr_nm"],
            "접수번호": r["rcept_no"],
            "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
        } for r in gov_rows])
        df_gov.to_excel(writer, sheet_name="경영권_지배구조", index=False)
        format_category_sheet(writer.sheets["경영권_지배구조"])
        
        # Sheet 8: 기타공시
        etc_rows = [r for r in active_records if r["category"] == "기타공시"]
        df_etc = pd.DataFrame([{
            "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
            "회사명": r["corp_name"],
            "시장구분": map_market(r["corp_cls"]),
            "종목코드": r["stock_code"],
            "공시명": r["report_nm"],
            "제출인": r["flr_nm"],
            "접수번호": r["rcept_no"],
            "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
        } for r in etc_rows])
        df_etc.to_excel(writer, sheet_name="기타공시", index=False)
        format_category_sheet(writer.sheets["기타공시"])
        
        # Sheet 9: 5%ㆍ임원보고
        equity_rows = [r for r in active_records if r["category"] == "지분공시"]
        officer_data_list = []
        for r in equity_rows:
            rn = r["report_nm"]
            # Determine report type
            if '대량보유상황보고서' in rn:
                rtype = '대량보유'
            else:
                rtype = '임원보고'
            
            html_path = os.path.join(workspace_dir, "data_dart", r["collected_date"], f"{r['rcept_no']}.html")
            parsed = parse_officer_report_html(html_path, rtype)
            
            for p in parsed:
                officer_data_list.append({
                    "접수일자": r.get("rcept_dt_display") or fmt_date(r["rcept_dt"]),
                    "회사명": r["corp_name"],
                    "종목코드": r["stock_code"],
                    "보고구분": rtype,
                    "보고자(주체)": p["reporter_name"],
                    "관계": p["relationship"],
                    "변동사유": p["change_reason"],
                    "증감(주)": p["shares_change"] if p["shares_change"] != 0 else None,
                    "단가(원)": p["avg_price"],
                    "변동후 보유(주)": p["shares_after"] if p["shares_after"] != 0 else None,
                    "보유비율(%)": p["ownership_pct"],
                    "접수번호": r["rcept_no"],
                    "DART링크": f'=HYPERLINK("https://dart.fss.or.kr/dsaf001/main.do?rcpNo={r["rcept_no"]}", "공시열람")'
                })
        
        df_officer = pd.DataFrame(officer_data_list)
        df_officer.to_excel(writer, sheet_name="5%_임원보고", index=False)
        format_officer_sheet(writer.sheets["5%_임원보고"])
        

    logger.info(f"Excel file built successfully: {excel_path}")
    return True

# -------------------------------------------------------------
# openpyxl Styling Helpers
# -------------------------------------------------------------
def format_officer_sheet(ws):
    """Styles the 5%ㆍ임원보고 sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 25
    
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            if col_idx in [1, 3, 4, 12]:  # 접수일자, 종목코드, 보고구분, 접수번호
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 13:  # DART링크
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            elif col_idx == 8:  # 증감(주)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Color negative red, positive blue
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        cell.font = Font(name="Malgun Gothic", size=9, color="FF0000")
                    elif cell.value > 0:
                        cell.font = Font(name="Malgun Gothic", size=9, color="0000FF")
            elif col_idx == 9:  # 단가(원)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 10:  # 변동후 보유(주)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 11:  # 보유비율(%)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Column widths
    col_widths = {1: 12, 2: 16, 3: 10, 4: 10, 5: 20, 6: 14, 7: 14,
                  8: 14, 9: 12, 10: 16, 11: 10, 12: 18, 13: 10}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
            
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def format_category_sheet(ws):
    """Styles the general category disclosure sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333F48", end_color="333F48", fill_type="solid") # Dark Charcoal
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            if col_idx in [1, 3, 4, 7]: # 접수일자, 시장구분, 종목코드, 접수번호
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [8]: # DART링크 (hyperlink formula)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if val.startswith("="):
                val = "공시열람"
            if len(val) > max_len:
                max_len = len(val)
        width = min(max(max_len * 1.5 + 3, 10), 60)
        ws.column_dimensions[col_letter].width = width
        
    ws.auto_filter.ref = ws.dimensions

def format_fundraising_sheet(ws):
    """Styles the detailed capital increase and mezzanine sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy Blue
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Soft highlights by financing type
    cb_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
    bw_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft orange
    eb_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
    cap_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Soft blue
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 25
    
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36 # Moderately tall row height
        
        type_cell = ws.cell(row=row_idx, column=5)
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            # Alignments
            if col_idx in [1, 3, 4, 5, 9, 11, 12, 13, 14, 17, 18, 19]: # Dates, codes, types, opt dates, claim dates
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 20: # DARTLink
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            elif col_idx == 16: # 조달목적 (Wrap text)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Number formatting
            if col_idx in [7, 8]: # 조달금액, 발행가/전환가
                cell.number_format = '₩#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 10: # 신주발행수
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 15: # 주식총수대비 (%)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
        # Apply highlight to the type cell
        t_val = str(type_cell.value)
        if "CB" in t_val:
            type_cell.fill = cb_fill
        elif "BW" in t_val:
            type_cell.fill = bw_fill
        elif "EB" in t_val:
            type_cell.fill = eb_fill
        elif "유상" in t_val:
            type_cell.fill = cap_fill
            
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 16: # 조달목적 (wider)
            ws.column_dimensions[col_letter].width = 45
        else:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "공시열람"
                if len(val) > max_len:
                    max_len = len(val)
            width = min(max(max_len * 1.4 + 3, 10), 30)
            ws.column_dimensions[col_letter].width = width
            
    ws.auto_filter.ref = ws.dimensions

def format_contract_sheet(ws):
    """Styles the detailed supply contracts sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy Blue
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 25
    
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            if col_idx in [1, 3, 4, 5, 8, 9, 11, 14]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 15:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            elif col_idx in [7, 13]: # 계약내용, 수주상대방
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 10: # 수주금액
                cell.number_format = '₩#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 12: # 최근매출액대비 (%)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
                    
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 7: # 계약내용
            ws.column_dimensions[col_letter].width = 40
        elif col_idx == 13: # 수주상대방
            ws.column_dimensions[col_letter].width = 25
        else:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "공시열람"
                if len(val) > max_len:
                    max_len = len(val)
            width = min(max(max_len * 1.4 + 3, 10), 30)
            ws.column_dimensions[col_letter].width = width
            
    ws.auto_filter.ref = ws.dimensions

def format_facility_sheet(ws):
    """Styles the detailed facility investment sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 25
    
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            if col_idx in [1, 3, 4, 7, 8, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 12:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            elif col_idx == 6: # 투자목적
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 9: # 투자금액
                cell.number_format = '₩#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 10: # 자기자본대비 (%)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
                    
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 6: # 투자목적
            ws.column_dimensions[col_letter].width = 40
        else:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "공시열람"
                if len(val) > max_len:
                    max_len = len(val)
            width = min(max(max_len * 1.4 + 3, 10), 30)
            ws.column_dimensions[col_letter].width = width
            
    ws.auto_filter.ref = ws.dimensions

def format_financial_sheet(ws):
    """Styles the detailed financial debt guarantees and loans sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 25
    
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            if col_idx in [1, 3, 4, 6, 10, 11, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 14:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            elif col_idx in [7, 12]: # 상대방, 목적/용도
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 8: # 보증/대여 금액
                cell.number_format = '₩#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 9: # 자기자본대비 (%)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
                    
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 12: # 목적/용도
            ws.column_dimensions[col_letter].width = 35
        elif col_idx == 7: # 상대방
            ws.column_dimensions[col_letter].width = 25
        else:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "공시열람"
                if len(val) > max_len:
                    max_len = len(val)
            width = min(max(max_len * 1.4 + 3, 10), 30)
            ws.column_dimensions[col_letter].width = width
            
    ws.auto_filter.ref = ws.dimensions

def format_treasury_sheet(ws):
    """Styles the detailed financial treasury shares sheet."""
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy
    data_font = Font(name="Malgun Gothic", size=9)
    link_font = Font(name="Malgun Gothic", size=9, color="0000FF", underline="single")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 25
    
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = data_border
            
            if col_idx in [1, 3, 4, 6, 10, 11, 12, 15]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 16:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = link_font
            elif col_idx in [5, 13, 14]: # 공시명, 위탁/수탁기관, 목적
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 7: # 예정금액
                cell.number_format = '₩#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 8: # 예정주식수
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
    for col_idx, col in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 5: # 공시명
            ws.column_dimensions[col_letter].width = 30
        elif col_idx == 14: # 목적
            ws.column_dimensions[col_letter].width = 35
        elif col_idx == 13: # 위탁투자업자/수탁기관
            ws.column_dimensions[col_letter].width = 25
        else:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    val = "공시열람"
                if len(val) > max_len:
                    max_len = len(val)
            width = min(max(max_len * 1.4 + 3, 10), 30)
            ws.column_dimensions[col_letter].width = width
            
    ws.auto_filter.ref = ws.dimensions

def main():
    parser = argparse.ArgumentParser(description="DART Disclosure Classifier & Mezzanine/Contracts/Investments Parser")
    parser.add_argument("--upload", action="store_true", help="Upload the compiled Excel to Telegram")
    args = parser.parse_args()
    
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. Compile summary and options
    success = build_excel_summary(workspace_dir)
    
    # 2. Upload to Telegram if requested
    if success and args.upload:
        excel_path = os.path.join(workspace_dir, "data_dart", f"dart_disclosures_summary_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx")
        
        if not TELEGRAM_BOT4_TOKEN or not TELEGRAM_TEST_CHAT_ID:
            logger.error("Missing TELEGRAM_BOT4_TOKEN or TELEGRAM_TEST_CHAT_ID in env.")
            return
            
        caption = f"📊 [DART 공시 요약 인덱스 리포트]\n- 공급계약, 유/무상증자, 전환사채(CB)/BW/EB, 시설투자, 채무보증/금전대여, 자기주식취득/신탁/소각 등 세부조항 정밀 파싱 완료\n- 기재정정(정정공시) 발생 시 최초공시 자동 연동 및 데이터 업데이트 반영\n- 일자: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        logger.info("Uploading Excel to Telegram...")
        telegram_sent = send_telegram_document(TELEGRAM_BOT4_TOKEN, TELEGRAM_TEST_CHAT_ID, excel_path, caption=caption)
        if telegram_sent:
            logger.info("Excel successfully uploaded to Telegram.")
        else:
            logger.error("Failed to upload Excel to Telegram.")


if __name__ == "__main__":
    main()

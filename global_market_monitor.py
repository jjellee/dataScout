#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import datetime
import requests
import json
import re
import io
import time
import asyncio
import pandas as pd
import yfinance as yf
import FinanceDataReader as fdr
from pykrx import stock as pykrx_stock
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Setup logging
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Custom env loader
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        logger.info("Loading environment variables from .env file...")
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    val_str = val.strip().strip("'").strip('"')
                    os.environ[key.strip()] = val_str

load_env()
TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_JJANG_GU_CHAT_ID = os.getenv("TELEGRAM_JJANG_GU_CHAT_ID") or "-1003877753638"

# ----------------- Helper Functions ----------------- #

def load_interest_sectors(market):
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "interest_sectors.txt")
    default_us = ["반도체", "소프트웨어", "생명 공학 및 의학 연구", "백화점", "자동차 및 트럭 제조", "온라인 서비스"]
    default_jp = ["Electric Appliances", "Information & Communication", "Transportation Equipment", "Pharmaceutical", "Chemicals"]
    default_kr = ["반도체와반도체장비", "자동차와부품", "제약과생물공학", "소프트웨어와서비스", "미디어와엔터테인먼트", "은행"]
    
    if not os.path.exists(path):
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("# --- US Market Interest Industries (Match with FDR 'Industry' column) ---\n")
                f.write("[US]\n")
                for s in default_us:
                    f.write(f"{s}\n")
                f.write("\n# --- JP Market Interest Sectors (Match with JPX '33 Sector(name)' column) ---\n")
                f.write("[JP]\n")
                for s in default_jp:
                    f.write(f"{s}\n")
                f.write("\n# --- KR Market Interest Sectors (Match with FDR 'Industry' column) ---\n")
                f.write("[KR]\n")
                for s in default_kr:
                    f.write(f"{s}\n")
        except Exception as e:
            logger.error(f"Failed to create default interest_sectors.txt: {e}")
        defaults = {"US": default_us, "JP": default_jp, "KR": default_kr}
        return defaults.get(market, default_us)
        
    sectors = []
    current_section = None
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if line == "[US]":
                    current_section = "US"
                    continue
                elif line == "[JP]":
                    current_section = "JP"
                elif line == "[KR]":
                    current_section = "KR"
                    continue
                if current_section == market:
                    sectors.append(line)
    except Exception as e:
        logger.error(f"Error loading interest_sectors.txt: {e}")
        
    if sectors:
        return sectors
    defaults = {"US": default_us, "JP": default_jp, "KR": default_kr}
    return defaults.get(market, default_us)

def send_telegram_document(token, chat_id, file_path, caption=None):
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    data = {"chat_id": chat_id}
    if caption:
        data["caption"] = caption
        data["parse_mode"] = "Markdown"
        
    try:
        with open(file_path, "rb") as f:
            files = {"document": f}
            resp = requests.post(url, data=data, files=files, timeout=300)
            return resp.json()
    except Exception as e:
        logger.error(f"Failed to send telegram document: {e}")
        return None

def send_telegram_message(token, chat_id, text):
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "Markdown",
        "disable_web_page_preview": True
    }
    try:
        resp = requests.post(url, json=payload, timeout=15)
        return resp.json()
    except Exception as e:
        logger.error(f"Failed to send telegram message: {e}")
        return None

def check_market_holiday(market):
    """
    Checks if the market was closed on the expected trading date.
    Returns (is_holiday, expected_date)
    """
    now = datetime.datetime.now()
    if market == "US":
        # Expected trade date is yesterday (run date - 1 day)
        expected_date = (now - datetime.timedelta(days=1)).date()
        try:
            spy = yf.Ticker("SPY")
            hist = spy.history(period="1d")
            if not hist.empty:
                last_date = hist.index[-1].date()
                if last_date != expected_date:
                    return True, expected_date
                return False, expected_date
        except Exception as e:
            logger.error(f"US holiday check error: {e}")
            return False, expected_date
    elif market == "JP":
        # Expected trade date is today (run date)
        expected_date = now.date()
        try:
            toyota = yf.Ticker("7203.T")
            hist = toyota.history(period="1d")
            if not hist.empty:
                last_date = hist.index[-1].date()
                if last_date != expected_date:
                    return True, expected_date
                return False, expected_date
        except Exception as e:
            logger.error(f"JP holiday check error: {e}")
            return False, expected_date
    else: # KR
        # Expected trade date is today (run after market close)
        expected_date = now.date()
        try:
            # Search last 7 days to find the actual last trading date
            start = (now - datetime.timedelta(days=7)).strftime("%Y%m%d")
            end = expected_date.strftime("%Y%m%d")
            df = pykrx_stock.get_market_ohlcv_by_date(start, end, "005930")
            if df.empty:
                return True, expected_date
            last_trade_date = df.index[-1].date()
            if last_trade_date != expected_date:
                # Return the actual last trading date so test mode can use it
                return True, last_trade_date
            return False, expected_date
        except Exception as e:
            logger.error(f"KR holiday check error: {e}")
            return False, expected_date
            
    return False, now.date()

# ----------------- Core Data Retrieval ----------------- #

def get_us_stock_list():
    logger.info("Fetching S&P 500 / NYSE / NASDAQ / AMEX stock list from FDR...")
    try:
        df_nasdaq = fdr.StockListing('NASDAQ')
        df_nyse = fdr.StockListing('NYSE')
        df_amex = fdr.StockListing('AMEX')
        
        df_all = pd.concat([df_nasdaq, df_nyse, df_amex], ignore_index=True)
        # Clean symbols
        df_all['Symbol'] = df_all['Symbol'].str.strip().str.replace('.', '-', regex=False)
        df_all = df_all.drop_duplicates(subset=['Symbol'])
        return df_all[['Symbol', 'Name', 'Industry']]
    except Exception as e:
        logger.error(f"Failed to fetch US stock list: {e}")
        return pd.DataFrame()

def get_jp_stock_list():
    logger.info("Scraping JPX listed issues Excel URL...")
    try:
        r = requests.get('https://www.jpx.co.jp/english/markets/statistics-equities/misc/01.html', timeout=15)
        xls_links = re.findall(r'href=\"([^\"]+data_e\.xls[x]?)\"', r.text)
        if not xls_links:
            logger.error("No JPX Excel links found.")
            return pd.DataFrame()
            
        url = "https://www.jpx.co.jp" + xls_links[0]
        logger.info(f"Downloading JPX listed issues from: {url}")
        
        r_file = requests.get(url, timeout=45)
        df_jpx = pd.read_excel(io.BytesIO(r_file.content))
        
        # Filter actual stocks
        df_filtered = df_jpx[df_jpx['Section/Products'].str.contains('Market', na=False) & 
                             ~df_jpx['Section/Products'].str.contains('PRO Market', na=False)]
        
        df_filtered = df_filtered.copy()
        df_filtered['Symbol'] = df_filtered['Local Code'].astype(str) + ".T"
        df_filtered['Name'] = df_filtered['Name (English)']
        df_filtered['Industry'] = df_filtered['33 Sector(name)']
        
        return df_filtered[['Symbol', 'Name', 'Industry']]
    except Exception as e:
        logger.error(f"Failed to fetch JP stock list: {e}")
        return pd.DataFrame()

def get_kr_stock_list(trade_date=None):
    """Fetch KOSPI/KOSDAQ stock list with WICS (WiseIndex) industry classification."""
    logger.info("Fetching KR stock list with WICS industry classification from WiseIndex...")
    try:
        date_str = trade_date.strftime("%Y%m%d") if trade_date else datetime.datetime.now().strftime("%Y%m%d")
        
        # WICS 중분류 (28 sectors) - more granular than KRX's ~25 sectors
        wics_mid_sectors = [
            'G1010', 'G1510', 'G2010', 'G2020', 'G2030',
            'G2510', 'G2520', 'G2530', 'G2550', 'G2560',
            'G3010', 'G3020', 'G3030',
            'G3510', 'G3520',
            'G4010', 'G4020', 'G4030', 'G4040', 'G4050',
            'G4510', 'G4520', 'G4530', 'G4535', 'G4540',
            'G5010', 'G5020',
            'G5510',
        ]
        
        url = "http://www.wiseindex.com/Index/GetIndexComponets"
        all_stocks = []
        
        for code in wics_mid_sectors:
            try:
                params = {'ceil_yn': 0, 'dt': date_str, 'sec_cd': code}
                resp = requests.get(url, params=params, timeout=15)
                items = resp.json().get('list', [])
                if not items:
                    continue
                sector_name = items[0].get('IDX_NM_KOR', '').replace('WICS ', '')
                for item in items:
                    all_stocks.append({
                        'Symbol': item['CMP_CD'],
                        'Name': item['CMP_KOR'],
                        'Industry': sector_name,
                    })
            except Exception as e:
                logger.warning(f"Failed to fetch WICS sector {code}: {e}")
                continue
        
        if not all_stocks:
            logger.error("WICS API returned no data. Falling back to pykrx ticker list.")
            # Fallback: get tickers from pykrx without sector mapping
            for mkt in ["KOSPI", "KOSDAQ"]:
                tickers = pykrx_stock.get_market_ticker_list(date_str, market=mkt)
                for t in tickers:
                    name = pykrx_stock.get_market_ticker_name(t)
                    all_stocks.append({'Symbol': t, 'Name': name, 'Industry': '기타'})
        
        df_all = pd.DataFrame(all_stocks).drop_duplicates(subset=['Symbol'])
        logger.info(f"WICS classification loaded: {len(df_all)} stocks, {df_all['Industry'].nunique()} sectors")
        
        # Also add pykrx tickers not covered by WICS (small/micro caps)
        wics_symbols = set(df_all['Symbol'].tolist())
        extra_stocks = []
        for mkt in ["KOSPI", "KOSDAQ"]:
            try:
                tickers = pykrx_stock.get_market_ticker_list(date_str, market=mkt)
                for t in tickers:
                    if t not in wics_symbols:
                        name = pykrx_stock.get_market_ticker_name(t)
                        extra_stocks.append({'Symbol': t, 'Name': name, 'Industry': '기타'})
            except Exception:
                continue
        
        if extra_stocks:
            df_extra = pd.DataFrame(extra_stocks)
            df_all = pd.concat([df_all, df_extra], ignore_index=True)
            logger.info(f"Added {len(extra_stocks)} extra tickers from pykrx (classified as '기타'). Total: {len(df_all)}")
        
        return df_all[['Symbol', 'Name', 'Industry']]
    except Exception as e:
        logger.error(f"Failed to fetch KR stock list: {e}")
        return pd.DataFrame()

def download_kr_prices(expected_date):
    """Download all KR stock prices for a given date using pykrx (single API call per market)."""
    date_str = expected_date.strftime("%Y%m%d")
    logger.info(f"Downloading KR prices for {date_str} using pykrx...")
    
    all_results = []
    for mkt in ["KOSPI", "KOSDAQ"]:
        try:
            df_mkt = pykrx_stock.get_market_ohlcv_by_ticker(date_str, market=mkt)
            if df_mkt.empty:
                logger.warning(f"No pykrx data for {mkt} on {date_str}")
                continue
            for ticker in df_mkt.index:
                close = df_mkt.loc[ticker, '종가']
                change = df_mkt.loc[ticker, '등락률']
                if close > 0:
                    all_results.append({
                        'Symbol': str(ticker),
                        'Price': float(close),
                        'Change': float(change)
                    })
        except Exception as e:
            logger.error(f"Failed to download {mkt} prices: {e}")
    
    logger.info(f"Downloaded price changes for {len(all_results)} KR stocks.")
    return pd.DataFrame(all_results)

def load_kr_investor_trend(expected_date):
    """Load investor trend CSV collected by collector.py for a given date."""
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    date_str = expected_date.strftime("%Y%m%d")
    csv_path = os.path.join(workspace_dir, "data_kr", date_str, "all_stocks_investor_trend.csv")
    
    if not os.path.exists(csv_path):
        logger.warning(f"Investor trend CSV not found: {csv_path}")
        return None
    
    try:
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        # Standardize ticker column
        df = df.rename(columns={'티커': 'Symbol'})
        df['Symbol'] = df['Symbol'].astype(str).str.zfill(6)
        logger.info(f"Loaded investor trend data: {len(df)} stocks from {csv_path}")
        return df
    except Exception as e:
        logger.error(f"Failed to load investor trend CSV: {e}")
        return None

def format_억(val):
    """Format a KRW value (원) to 억원 for readability."""
    억 = val / 1e8
    if abs(억) >= 1:
        return f"{억:+,.0f}억"
    return f"{억:+,.1f}억"

def download_prices_bulk(symbols):
    logger.info(f"Downloading daily prices for {len(symbols)} tickers in chunks of 300...")
    all_results = []
    
    # Split into chunks of 300
    chunks = [symbols[i:i+300] for i in range(0, len(symbols), 300)]
    
    for idx, chunk in enumerate(chunks, start=1):
        chunk_str = " ".join(chunk)
        try:
            # Fetch 5 days to ensure we always have 2+ active trading days
            df_chunk = yf.download(chunk_str, period="5d", group_by="ticker", progress=False)
            
            # Extract close prices
            for ticker in df_chunk.columns.get_level_values(0).unique():
                try:
                    close_series = df_chunk[ticker]['Close'].dropna()
                    if len(close_series) >= 2:
                        prev_close = float(close_series.iloc[-2])
                        close = float(close_series.iloc[-1])
                        if prev_close > 0:
                            change = (close - prev_close) / prev_close * 100
                            all_results.append({
                                'Symbol': ticker,
                                'Price': close,
                                'Change': change
                            })
                except Exception:
                    continue
        except Exception as e:
            logger.error(f"Failed to download chunk {idx}/{len(chunks)}: {e}")
            
        time.sleep(0.3) # Polite delay
        
    return pd.DataFrame(all_results)


def get_news_summary(ticker):
    """Get the most recent news summary for a ticker from yfinance."""
    try:
        news = yf.Ticker(ticker).news
        if not news:
            return None
        for item in news[:3]:
            content = item.get('content', {})
            summary = content.get('summary', '')
            title = content.get('title', '')
            if summary and len(summary) > 10:
                return summary[:120]
            if title and len(title) > 5:
                return title[:120]
    except Exception:
        pass
    return None


def get_news_batch(tickers, max_count=50):
    """Fetch news summaries for a batch of tickers."""
    news_map = {}
    for t in tickers[:max_count]:
        summary = get_news_summary(t)
        if summary:
            news_map[t] = summary
        time.sleep(0.1)
    return news_map

# ----------------- Excel Formatting ----------------- #

def save_to_formatted_excel(df, output_path, market):
    logger.info(f"Saving formatted Excel to {output_path}...")
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Sort by change descending
        df_sorted = df.sort_values(by='Change', ascending=False)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_sorted.to_excel(writer, index=False, sheet_name='Price Changes')
            ws = writer.sheets['Price Changes']
            
            # Styling
            header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="333F48", end_color="333F48", fill_type="solid")
            data_font = Font(name="Malgun Gothic", size=9)
            
            buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
            sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # soft red
            
            border_side = Side(border_style="thin", color="D3D3D3")
            data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            
            currency_map = {"US": '$#,##0.00', "JP": '¥#,##0', "KR": '₩#,##0'}
            currency_format = currency_map.get(market, '#,##0')
            
            # Format Headers
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Format Rows
            for row_idx in range(2, ws.max_row + 1):
                symbol_cell = ws.cell(row=row_idx, column=1)
                price_cell = ws.cell(row=row_idx, column=2)
                change_cell = ws.cell(row=row_idx, column=3)
                name_cell = ws.cell(row=row_idx, column=4)
                industry_cell = ws.cell(row=row_idx, column=5)
                
                # Alignments
                symbol_cell.alignment = Alignment(horizontal="center")
                price_cell.alignment = Alignment(horizontal="right")
                change_cell.alignment = Alignment(horizontal="right")
                name_cell.alignment = Alignment(horizontal="left")
                industry_cell.alignment = Alignment(horizontal="left")
                
                # Fonts
                for cell in [symbol_cell, price_cell, change_cell, name_cell, industry_cell]:
                    cell.font = data_font
                    cell.border = data_border
                    
                # Formats
                price_cell.number_format = currency_format
                change_cell.number_format = '0.00"%"'
                
                # Highlighting
                val = change_cell.value
                if val is not None:
                    try:
                        val_num = float(val)
                        if val_num >= 5.0:
                            change_cell.fill = buy_fill
                        elif val_num <= -5.0:
                            change_cell.fill = sell_fill
                    except (ValueError, TypeError):
                        pass
                        
            # Adjust column widths
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                width = min(max(max_len + 3, 10), 50)
                ws.column_dimensions[col_letter].width = width
                
            ws.auto_filter.ref = ws.dimensions
            
            # Style Comment column if present
            comment_col_idx = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == 'Comment':
                    comment_col_idx = col_idx
                    break
            
            if comment_col_idx:
                comment_font = Font(name="Malgun Gothic", size=8, color="555555")
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=comment_col_idx)
                    cell.font = comment_font
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
                    cell.border = data_border
                ws.column_dimensions[get_column_letter(comment_col_idx)].width = 60
            
        return True
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
        return False

# ----------------- Main Execution ----------------- #

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Global Market Price Monitor")
    parser.add_argument("--market", type=str, required=True, choices=["US", "JP", "KR"], help="Target market: US, JP, or KR")
    parser.add_argument("--test", action="store_true", help="Run in test mode")
    args = parser.parse_args()
    
    market = args.market
    test_mode = args.test
    
    logger.info(f"==================================================")
    logger.info(f" Starting Daily {market} Market Monitor (Test: {test_mode})")
    logger.info(f"==================================================")
    
    # 1. Holiday Check
    is_holiday, expected_date = check_market_holiday(market)
    if is_holiday and not test_mode:
        logger.info(f"{market} market was closed on expected trading date {expected_date}. Skipping report.")
        return
    elif is_holiday and test_mode:
        logger.info(f"[Test Mode] Market was closed on expected trading date {expected_date}, but continuing test run anyway.")
        
    logger.info(f"Processing trading date: {expected_date}")
    
    # 2. Load Stock List Metadata
    if market == "KR":
        df_list = get_kr_stock_list(trade_date=expected_date)
    elif market == "JP":
        df_list = get_jp_stock_list()
    else:
        df_list = get_us_stock_list()
    if df_list.empty:
        logger.error("Failed to load stock list metadata. Exiting.")
        return
        
    logger.info(f"Total metadata stocks loaded: {len(df_list)}")
    
    # 3. Download Prices in Bulk
    if market == "KR":
        # KR uses pykrx (single API call per market, no chunking needed)
        df_prices = download_kr_prices(expected_date)
    else:
        symbols = df_list['Symbol'].tolist()
        if test_mode:
            logger.info("[Test Mode] Limiting symbol list to 30 popular/sample symbols for rapid testing.")
            test_symbols_subset = []
            if market == "US":
                known_tickers = ["AAPL", "MSFT", "NVDA", "TSLA", "AMZN", "GOOGL", "META", "AMD", "INTC", "AVGO"]
                test_symbols_subset = [sym for sym in known_tickers if sym in symbols]
            else: # JP
                known_tickers = ["7203.T", "9984.T", "6758.T", "8035.T", "6857.T", "6501.T", "4502.T", "7974.T"]
                test_symbols_subset = [sym for sym in known_tickers if sym in symbols]
            for sym in symbols:
                if len(test_symbols_subset) >= 30:
                    break
                if sym not in test_symbols_subset:
                    test_symbols_subset.append(sym)
            symbols = test_symbols_subset
        df_prices = download_prices_bulk(symbols)
    
    if df_prices.empty:
        logger.error("No stock prices downloaded. Exiting.")
        return
        
    logger.info(f"Successfully downloaded price changes for {len(df_prices)} stocks.")
    
    # 4. Merge Prices and Metadata
    df_merged = df_prices.merge(df_list, on='Symbol', how='inner')
    # Filter out records without valid industries
    df_merged = df_merged[df_merged['Industry'].notna() & (df_merged['Industry'] != '-')]
    
    if df_merged.empty:
        logger.error("No records matched after merging metadata and prices. Exiting.")
        return
        
    # 5. Add news commentary for sharp movers (±5%)
    logger.info("Fetching news for sharp movers (±5%)...")
    sharp_movers = df_merged[df_merged['Change'].abs() >= 5.0]
    if not sharp_movers.empty:
        mover_symbols = sharp_movers['Symbol'].tolist()
        # For KR market, yfinance needs .KS/.KQ suffix
        if market == "KR":
            yf_symbols = [f"{s}.KS" for s in mover_symbols]
        else:
            yf_symbols = mover_symbols
        
        news_map = get_news_batch(yf_symbols, max_count=50)
        logger.info(f"Got news for {len(news_map)} out of {len(mover_symbols)} sharp movers")
        
        # Map back to original symbols
        comment_map = {}
        if market == "KR":
            for orig, yf_sym in zip(mover_symbols, yf_symbols):
                if yf_sym in news_map:
                    comment_map[orig] = news_map[yf_sym]
        else:
            comment_map = news_map
        
        df_merged['Comment'] = df_merged['Symbol'].map(comment_map).fillna('')
    else:
        df_merged['Comment'] = ''
    
    # 6. Format and Save Excel
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    date_str = expected_date.strftime("%Y%m%d")
    market_lower = market.lower()
    
    excel_filename = f"{market_lower}_prices_{date_str}.xlsx"
    excel_path = os.path.join(workspace_dir, f"data_{market_lower}", "daily_prices", excel_filename)
    
    save_to_formatted_excel(df_merged, excel_path, market)
    
    # 6. Analyze Sectors & Highlight Trades
    # A. Price Movers
    gainers = df_merged.sort_values(by='Change', ascending=False).head(10)
    losers = df_merged.sort_values(by='Change', ascending=True).head(10)
    
    # B. Sector Performance (average price change by Industry)
    sector_perf = df_merged.groupby('Industry')['Change'].mean().reset_index()
    sector_perf_sorted = sector_perf.sort_values(by='Change', ascending=False)
    
    top_sectors = sector_perf_sorted.head(3)
    bottom_sectors = sector_perf_sorted.tail(3)
    
    # C. Load Interest Sectors
    interest_sectors = load_interest_sectors(market)
    interest_perf = sector_perf[sector_perf['Industry'].isin(interest_sectors)].sort_values(by='Change', ascending=False)
    
    # 7. Formulate Telegram Report
    flag_map = {"US": "🇺🇸", "JP": "🇯🇵", "KR": "🇰🇷"}
    name_map = {"US": "미국 증시", "JP": "일본 증시", "KR": "한국 증시"}
    currency_map_sym = {"US": "$", "JP": "¥", "KR": "₩"}
    market_flag = flag_map[market]
    market_name = name_map[market]
    currency_symbol = currency_map_sym[market]
    
    report_lines = []
    report_lines.append(f"{market_flag} *[{market_name} 마감 일일 리포트]* {market_flag}")
    report_lines.append(f"📅 *기준일자: {expected_date.strftime('%Y-%m-%d')}*\n")
    
    # Sector performance
    report_lines.append("🔥 *주요 상승 섹터/산업 TOP 3*")
    for idx, row in enumerate(top_sectors.itertuples(), start=1):
        report_lines.append(f"{idx}. *{row.Industry}* (+{row.Change:.2f}%)")
        
    report_lines.append("\n❄️ *주요 하락 섹터/산업 BOTTOM 3*")
    for idx, row in enumerate(reversed(list(bottom_sectors.itertuples())), start=1):
        report_lines.append(f"{idx}. *{row.Industry}* ({row.Change:.2f}%)")
        
    # Interest Sectors Section
    if not interest_perf.empty:
        report_lines.append("\n📌 *나의 관심 섹터/산업 현황*")
        for row in interest_perf.itertuples():
            # Find top 2 gainers in this interest sector
            sector_stocks = df_merged[df_merged['Industry'] == row.Industry]
            top_stocks = sector_stocks.sort_values(by='Change', ascending=False).head(2)
            stocks_str = ", ".join([f"{r.Name[:8]}({r.Change:+.1f}%)" for r in top_stocks.itertuples()])
            report_lines.append(f"- *{row.Industry}*: {row.Change:+.2f}% (상위: {stocks_str})")
    
    # KR Supply/Demand (수급) Analysis
    if market == "KR":
        df_investor = load_kr_investor_trend(expected_date)
        if df_investor is not None:
            # Merge investor data with WICS industry classification
            inv_cols = ['Symbol', '외국인_순매수대금', '기관합계_순매수대금', '개인_순매수대금', '공매도비중']
            df_inv = df_investor[[c for c in inv_cols if c in df_investor.columns]].copy()
            df_supply = df_merged.merge(df_inv, on='Symbol', how='left')
            
            # Fill NaN with 0
            for col in ['외국인_순매수대금', '기관합계_순매수대금', '개인_순매수대금', '공매도비중']:
                if col in df_supply.columns:
                    df_supply[col] = df_supply[col].fillna(0)
            
            report_lines.append("\n💰 *외국인·기관 수급 분석*")
            
            # Sector-level foreign flow
            if '외국인_순매수대금' in df_supply.columns:
                sector_foreign = df_supply.groupby('Industry')['외국인_순매수대금'].sum().reset_index()
                sector_foreign = sector_foreign.sort_values('외국인_순매수대금', ascending=False)
                
                top_buy_sectors = sector_foreign.head(3)
                top_sell_sectors = sector_foreign.tail(3)
                
                report_lines.append("\n🏦 *섹터별 외국인 순매수 TOP 3*")
                for idx, row in enumerate(top_buy_sectors.itertuples(), start=1):
                    sec_change = sector_perf[sector_perf['Industry'] == row.Industry]['Change'].values
                    chg_str = f" (주가 {sec_change[0]:+.1f}%)" if len(sec_change) > 0 else ""
                    report_lines.append(f"{idx}. *{row.Industry}*: {format_억(row.외국인_순매수대금)}{chg_str}")
                
                report_lines.append("\n📉 *섹터별 외국인 순매도 TOP 3*")
                for idx, row in enumerate(reversed(list(top_sell_sectors.itertuples())), start=1):
                    sec_change = sector_perf[sector_perf['Industry'] == row.Industry]['Change'].values
                    chg_str = f" (주가 {sec_change[0]:+.1f}%)" if len(sec_change) > 0 else ""
                    report_lines.append(f"{idx}. *{row.Industry}*: {format_억(row.외국인_순매수대금)}{chg_str}")
                
                # Contrarian signals: stocks down >2% but foreign/institutional buying
                contrarian = df_supply[
                    (df_supply['Change'] <= -2.0) &
                    (df_supply['외국인_순매수대금'] > 1e8) &  # >1억원 순매수
                    (df_supply['Industry'] != '기타')
                ].sort_values('외국인_순매수대금', ascending=False).head(5)
                
                if not contrarian.empty:
                    report_lines.append("\n🎯 *역발상 시그널 (하락 중 외국인 순매수)*")
                    for row in contrarian.itertuples():
                        기관 = getattr(row, '기관합계_순매수대금', 0)
                        기관_str = f", 기관 {format_억(기관)}" if abs(기관) > 1e8 else ""
                        report_lines.append(f"- *{row.Name[:10]}* {row.Change:+.1f}% | 외국인 {format_억(row.외국인_순매수대금)}{기관_str}")
                
                # Top foreign net buy stocks
                top_foreign_buy = df_supply[df_supply['외국인_순매수대금'] > 0].sort_values('외국인_순매수대금', ascending=False).head(5)
                if not top_foreign_buy.empty:
                    report_lines.append("\n👤 *외국인 순매수 TOP 5*")
                    for idx, row in enumerate(top_foreign_buy.itertuples(), start=1):
                        report_lines.append(f"{idx}. *{row.Name[:10]}* ({row.Change:+.1f}%) | {format_억(row.외국인_순매수대금)}")
            
    # Stock movers
    report_lines.append("\n🚀 *개별 종목 상승 TOP 5*")
    for idx, row in enumerate(gainers.head(5).itertuples(), start=1):
        price_val = f"{row.Price:,.2f}" if market == "US" else f"{int(row.Price):,}"
        report_lines.append(f"{idx}. *{row.Symbol}* ({row.Name[:20]}) | {currency_symbol}{price_val} (*+{row.Change:.2f}%*)")
        
    report_lines.append("\n📉 *개별 종목 하락 TOP 5*")
    for idx, row in enumerate(losers.head(5).itertuples(), start=1):
        price_val = f"{row.Price:,.2f}" if market == "US" else f"{int(row.Price):,}"
        report_lines.append(f"{idx}. *{row.Symbol}* ({row.Name[:20]}) | {currency_symbol}{price_val} (*{row.Change:.2f}%*)")
        
    report_lines.append(f"\n📂 *전체 {len(df_merged):,}개 종목의 마감 등락폭 상세 정보가 담긴 엑셀 파일을 첨부합니다.*")
    
    report_text = "\n".join(report_lines)
    
    # 8. Upload to Telegram
    target_chat = TELEGRAM_JJANG_GU_CHAT_ID
    if test_mode:
        # Use test bot/chat
        target_chat = os.getenv("TELEGRAM_TEST_CHAT_ID") or "-1003843549676"
        
    if TELEGRAM_BOT4_TOKEN and target_chat:
        logger.info(f"Sending daily report text to Telegram chat {target_chat}...")
        res_msg = send_telegram_message(TELEGRAM_BOT4_TOKEN, target_chat, report_text)
        if res_msg and res_msg.get("ok"):
            logger.info("Report text sent successfully to Telegram.")
        else:
            logger.error(f"Failed to send report text to Telegram: {res_msg}")
            
        logger.info(f"Uploading Excel document to Telegram chat {target_chat}...")
        doc_caption = f"📂 전체 {len(df_merged):,}개 종목 마감 등락폭 상세 Excel ({market} - {expected_date.strftime('%Y-%m-%d')})"
        res_doc = send_telegram_document(TELEGRAM_BOT4_TOKEN, target_chat, excel_path, caption=doc_caption)
        if res_doc and res_doc.get("ok"):
            logger.info("Excel document sent successfully to Telegram.")
        else:
            logger.error(f"Failed to send Excel document to Telegram: {res_doc}")
    else:
        logger.error("Telegram credentials or Chat ID is missing in the environment.")
        
    print("\n--- Generated Report Summary ---")
    print(report_text)

if __name__ == "__main__":
    main()

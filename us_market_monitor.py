#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import datetime
import requests
import json
import pandas as pd
import time
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import yfinance as yf
import FinanceDataReader as fdr

# Setup logging
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Custom env loader
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        logger.info("Loading environment variables from .env file...")
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    val_str = val.strip().strip("'").strip('"')
                    os.environ[key.strip()] = val_str

load_env()
TELEGRAM_BOT4_TOKEN = os.getenv("TELEGRAM_BOT4_TOKEN")
TELEGRAM_JJANG_GU_CHAT_ID = os.getenv("TELEGRAM_JJANG_GU_CHAT_ID") or "-1003877753638"
SEC_USER_AGENT = "Norbot norbot.trading@gmail.com"  # Essential for SEC EDGAR API

def get_us_market_movers():
    """
    Downloads S&P 500 stock prices for the last 2 days,
    calculates daily percentage change, and returns top 5 gainers and losers.
    """
    logger.info("Fetching S&P 500 stock list from FinanceDataReader...")
    try:
        sp500 = fdr.StockListing('S&P500')
        symbols = sp500['Symbol'].unique().tolist()
        # Clean symbols (replace . with - for Yahoo Finance compatibility, e.g. BRK.B -> BRK-B)
        symbols = [s.replace('.', '-') for s in symbols if isinstance(s, str)]
    except Exception as e:
        logger.error(f"Failed to fetch S&P 500 symbols: {e}")
        return [], []

    logger.info(f"Downloading daily price data for {len(symbols)} stocks using yfinance...")
    
    # Split into 2 batches to avoid URL length limitations
    batch_size = 250
    all_dfs = []
    
    for i in range(0, len(symbols), batch_size):
        batch = symbols[i:i+batch_size]
        batch_str = " ".join(batch)
        try:
            # Download 2 days of close data
            df = yf.download(batch_str, period="2d", group_by="ticker", progress=False)
            all_dfs.append(df)
            time.sleep(0.5)  # Polite delay
        except Exception as e:
            logger.error(f"Failed to download batch {i//batch_size + 1}: {e}")

    # Process and calculate changes
    results = []
    for df in all_dfs:
        for ticker in df.columns.get_level_values(0).unique():
            try:
                # Check if ticker has Close price
                close_series = df[ticker]['Close'].dropna()
                if len(close_series) >= 2:
                    yesterday_close = float(close_series.iloc[-2])
                    today_close = float(close_series.iloc[-1])
                    if yesterday_close > 0:
                        change_pct = (today_close - yesterday_close) / yesterday_close * 100
                        results.append({
                            'ticker': ticker,
                            'price': today_close,
                            'change': change_pct
                        })
            except Exception:
                continue

    if not results:
        logger.warning("No price change results computed.")
        return [], []

    df_movers = pd.DataFrame(results)
    
    # Sort to get Top Gainers and Top Losers
    gainers = df_movers.sort_values(by='change', ascending=False).head(5).to_dict(orient='records')
    losers = df_movers.sort_values(by='change', ascending=True).head(5).to_dict(orient='records')
    
    return gainers, losers

def get_sec_insider_transactions():
    """
    Queries SEC Atom feed, parses Form 4 XML filings,
    and returns a list of meaningful insider transactions.
    """
    logger.info("Fetching SEC current Form 4 Atom feed...")
    url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcurrent&type=4&count=100&output=atom"
    headers = {"User-Agent": SEC_USER_AGENT}
    
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code != 200:
            logger.error(f"SEC Feed HTTP error: {response.status_code}")
            return []
            
        soup = BeautifulSoup(response.content, "xml")
        entries = soup.find_all("entry")
        logger.info(f"Found {len(entries)} entries in SEC Atom feed.")
    except Exception as e:
        logger.error(f"Failed to fetch or parse SEC Atom feed: {e}")
        return []

    # Filter for unique index links (Issuer and Reporting share same links)
    unique_links = {}
    for entry in entries:
        link_tag = entry.find("link")
        updated_tag = entry.find("updated")
        if link_tag and link_tag.get("href"):
            link_href = link_tag["href"]
            updated_time = updated_tag.get_text().strip() if updated_tag else datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            unique_links[link_href] = updated_time
            
    logger.info(f"Unique Form 4 index URLs to process: {len(unique_links)}")
    
    transactions = []
    
    # Process each index link (limit to first 30 unique links to stay fast and avoid rate limits)
    for index_url, filing_date in list(unique_links.items())[:30]:
        try:
            # 1. Fetch index page
            time.sleep(0.15)  # Obey SEC rate limits (< 10 reqs/sec)
            idx_resp = requests.get(index_url, headers=headers, timeout=10)
            if idx_resp.status_code != 200:
                continue
                
            # 2. Find the raw XML file link
            idx_soup = BeautifulSoup(idx_resp.content, "html.parser")
            xml_href = None
            for link in idx_soup.find_all("a"):
                href = link.get("href", "")
                if href.endswith(".xml") and "xsl" not in href:
                    xml_href = href
                    break
                    
            if not xml_href:
                continue
                
            xml_url = urljoin(index_url, xml_href)
            
            # 3. Fetch raw XML filing
            time.sleep(0.15)
            xml_resp = requests.get(xml_url, headers=headers, timeout=10)
            if xml_resp.status_code != 200:
                continue
                
            # 4. Parse XML filing
            xml_soup = BeautifulSoup(xml_resp.content, "xml")
            
            # Extract basic metadata
            insider = xml_soup.find("rptOwnerName")
            insider_name = insider.get_text().strip() if insider else "Unknown"
            
            ticker = xml_soup.find("issuerTradingSymbol")
            ticker_symbol = ticker.get_text().strip() if ticker else "Unknown"
            
            issuer = xml_soup.find("issuerName")
            issuer_name = issuer.get_text().strip() if issuer else "Unknown"
            
            # Parse relationship
            relationship = "Insider"
            is_director = xml_soup.find("isDirector")
            is_officer = xml_soup.find("isOfficer")
            is_ten_percent = xml_soup.find("isTenPercentOwner")
            officer_title = xml_soup.find("officerTitle")
            
            if is_director and (is_director.get_text() in ["1", "true"]):
                relationship = "Director"
            elif is_officer and (is_officer.get_text() in ["1", "true"]):
                relationship = officer_title.get_text().strip() if officer_title else "Officer"
            elif is_ten_percent and (is_ten_percent.get_text() in ["1", "true"]):
                relationship = "10% Owner"
            
            # Process non-derivative transactions
            nd_trans = xml_soup.find_all("nonDerivativeTransaction")
            for t in nd_trans:
                security_title_tag = t.find("securityTitle")
                security_title = security_title_tag.find('value').get_text().strip() if security_title_tag and security_title_tag.find('value') else "Common Stock"
                
                acq_disp_tag = t.find("transactionAcquiredDisposedCode")
                acq_disp = acq_disp_tag.find('value').get_text().strip() if acq_disp_tag and acq_disp_tag.find('value') else None
                
                shares_tag = t.find("transactionShares")
                shares = float(shares_tag.find('value').get_text().strip()) if shares_tag and shares_tag.find('value') else 0.0
                
                price_tag = t.find("transactionPricePerShare")
                price = float(price_tag.find('value').get_text().strip()) if price_tag and price_tag.find('value') else 0.0
                
                total_value = shares * price
                
                if acq_disp in ["A", "D"] and total_value > 0:
                    is_meaningful = False
                    if acq_disp == "A" and total_value >= 30000:
                        is_meaningful = True
                    elif acq_disp == "D" and total_value >= 200000:
                        is_meaningful = True
                        
                    transactions.append({
                        'filing_date': filing_date,
                        'ticker': ticker_symbol,
                        'company': issuer_name,
                        'insider': insider_name,
                        'role': relationship,
                        'type': 'BUY' if acq_disp == 'A' else 'SELL',
                        'shares': int(shares),
                        'price': price,
                        'value': total_value,
                        'is_meaningful': is_meaningful,
                        'filing_url': xml_url
                    })
        except Exception as e:
            logger.error(f"Failed to process Form 4 index {index_url}: {e}")
            continue
            
    return transactions

def send_telegram_message(token, chat_id, text):
    """Sends a markdown-formatted message to Telegram."""
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "Markdown",
        "disable_web_page_preview": True
    }
    try:
        resp = requests.post(url, json=payload, timeout=10)
        return resp.json()
    except Exception as e:
        logger.error(f"Failed to send telegram request: {e}")
        return None

def save_insider_transactions_to_excel(transactions):
    """
    Saves the list of insider transactions cumulatively into an Excel file.
    Deduplicates based on filing_url, ticker, insider, shares, price, and type.
    Returns True if new transactions were added.
    """
    if not transactions:
        return False
        
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(workspace_dir, "data_us")
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, "us_insider_transactions.xlsx")
    
    df_new = pd.DataFrame(transactions)
    if 'is_meaningful' in df_new.columns:
        df_new = df_new.drop(columns=['is_meaningful'])
    
    old_row_count = 0
    if os.path.exists(excel_path):
        try:
            df_old = pd.read_excel(excel_path)
            if 'is_meaningful' in df_old.columns:
                df_old = df_old.drop(columns=['is_meaningful'])
            old_row_count = len(df_old)
            # Combine old and new
            df_combined = pd.concat([df_old, df_new], ignore_index=True)
        except Exception as e:
            logger.error(f"Failed to read existing Excel file: {e}")
            df_combined = df_new
    else:
        df_combined = df_new
        
    # Deduplicate
    dup_cols = ['ticker', 'insider', 'shares', 'price', 'type']
    if 'filing_url' in df_combined.columns:
        dup_cols.append('filing_url')
        
    df_combined = df_combined.drop_duplicates(subset=dup_cols, keep='first')
    
    # Sort by filing_date descending (if available)
    if 'filing_date' in df_combined.columns:
        df_combined = df_combined.sort_values(by='filing_date', ascending=False)
        
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_combined.to_excel(writer, index=False, sheet_name='Transactions')
            
            worksheet = writer.sheets['Transactions']
            
            # Define fills
            buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
            sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # soft red
            meaningful_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # soft yellow/gold
            
            # Color Column 6 (type) based on BUY/SELL and highlight meaningful values in Column 9 (value)
            for row_idx in range(2, worksheet.max_row + 1):
                type_cell = worksheet.cell(row=row_idx, column=6)
                shares_cell = worksheet.cell(row=row_idx, column=7)
                price_cell = worksheet.cell(row=row_idx, column=8)
                value_cell = worksheet.cell(row=row_idx, column=9)
                
                # Apply 3-digit comma formatting
                shares_cell.number_format = '#,##0'
                price_cell.number_format = '#,##0.00'
                value_cell.number_format = '#,##0'
                
                if type_cell.value == "BUY":
                    type_cell.fill = buy_fill
                    try:
                        val_num = float(value_cell.value)
                        if val_num >= 30000:
                            value_cell.fill = meaningful_fill
                    except (ValueError, TypeError):
                        pass
                elif type_cell.value == "SELL":
                    type_cell.fill = sell_fill
                    try:
                        val_num = float(value_cell.value)
                        if val_num >= 200000:
                            value_cell.fill = meaningful_fill
                    except (ValueError, TypeError):
                        pass
            
            # Adjust column widths
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                width = min(max(max_len + 3, 10), 50)
                if col_idx in [3, 4, 5]: # company, insider, role
                    width = width / 2
                worksheet.column_dimensions[col_letter].width = width
                
        logger.info(f"Successfully saved {len(df_combined)} cumulative insider transactions to {excel_path}")
        if len(df_combined) > old_row_count:
            return True
    except Exception as e:
        logger.error(f"Failed to save cumulative Excel file: {e}")
        
    return False

def send_telegram_document(token, chat_id, file_path, caption=None):
    """Sends a document/file to Telegram."""
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    data = {"chat_id": chat_id}
    if caption:
        data["caption"] = caption
        data["parse_mode"] = "Markdown"
        
    try:
        with open(file_path, "rb") as f:
            files = {"document": f}
            resp = requests.post(url, data=data, files=files, timeout=300)
            return resp.json()
    except Exception as e:
        logger.error(f"Failed to send telegram document: {e}")
        return None

def main():
    logger.info("Starting Daily US Market Monitor...")
    
    # 1. Get market movers
    gainers, losers = get_us_market_movers()
    
    # Fetch short interest for movers
    logger.info("Fetching short interest data for movers...")
    for item in gainers + losers:
        ticker_symbol = item['ticker']
        try:
            ticker = yf.Ticker(ticker_symbol)
            info = ticker.info
            short_percent = info.get("shortPercentOfFloat")
            short_ratio = info.get("shortRatio")
            
            if short_percent is not None:
                item['short_percent'] = float(short_percent) * 100
            else:
                item['short_percent'] = None
                
            if short_ratio is not None:
                item['short_ratio'] = float(short_ratio)
            else:
                item['short_ratio'] = None
        except Exception as e:
            logger.error(f"Failed to fetch short interest for {ticker_symbol}: {e}")
            item['short_percent'] = None
            item['short_ratio'] = None
            
    # 2. Get insider transactions
    insiders = get_sec_insider_transactions()
    
    # Save all parsed transactions to Excel and check if updated
    is_updated = save_insider_transactions_to_excel(insiders)
    
    if is_updated:
        workspace_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(workspace_dir, "data_us", "us_insider_transactions.xlsx")
        if TELEGRAM_BOT4_TOKEN and TELEGRAM_JJANG_GU_CHAT_ID:
            logger.info(f"Uploading updated Excel sheet to Telegram chat: {TELEGRAM_JJANG_GU_CHAT_ID}...")
            caption_text = "📁 *[미국 내부자 지분 변동 Excel 업데이트]*\n새로운 내부자 거래가 추가되어 누적 엑셀 파일을 업로드합니다."
            res_file = send_telegram_document(TELEGRAM_BOT4_TOKEN, TELEGRAM_JJANG_GU_CHAT_ID, excel_path, caption=caption_text)
            if res_file and res_file.get("ok"):
                logger.info("Successfully uploaded Excel sheet to Telegram.")
            else:
                logger.error(f"Failed to upload Excel sheet to Telegram: {res_file}")
                
    # Filter for meaningful transactions to send to Telegram
    meaningful_insiders = [item for item in insiders if item.get('is_meaningful')]
    
    # 3. Format Telegram Report
    report_lines = []
    report_lines.append("🇺🇸 *[미국 증시 일일 요약 리포트]* 🇺🇸\n")
    
    # A. Price Movers Section
    report_lines.append("📈 *S&P 500 상승 TOP 5*")
    if gainers:
        for idx, item in enumerate(gainers):
            short_str = ""
            if item.get('short_percent') is not None:
                short_str = f" (공매도 비율: {item['short_percent']:.2f}%"
                if item.get('short_ratio') is not None:
                    short_str += f", Ratio: {item['short_ratio']:.1f}"
                short_str += ")"
            report_lines.append(f"{idx+1}. *{item['ticker']}* | ${item['price']:.2f} (+{item['change']:.2f}%){short_str}")
    else:
        report_lines.append("- 데이터 없음")
        
    report_lines.append("\n📉 *S&P 500 하락 TOP 5*")
    if losers:
        for idx, item in enumerate(losers):
            short_str = ""
            if item.get('short_percent') is not None:
                short_str = f" (공매도 비율: {item['short_percent']:.2f}%"
                if item.get('short_ratio') is not None:
                    short_str += f", Ratio: {item['short_ratio']:.1f}"
                short_str += ")"
            report_lines.append(f"{idx+1}. *{item['ticker']}* | ${item['price']:.2f} ({item['change']:.2f}%){short_str}")
    else:
        report_lines.append("- 데이터 없음")
        
    # B. Insider Trading Section
    report_lines.append("\n🕵️ *주요 기업 내부자 지분 변동 (SEC Form 4)*")
    if meaningful_insiders:
        # Sort by value descending
        insiders_sorted = sorted(meaningful_insiders, key=lambda x: x['value'], reverse=True)
        # Limit to top 10 insider trades
        for item in insiders_sorted[:10]:
            action_emoji = "🟢 매수" if item['type'] == 'BUY' else "🔴 매도"
            value_str = f"${item['value']:,.0f}"
            report_lines.append(
                f"- *{item['ticker']}* ({item['company']}) \n"
                f"  └ {action_emoji} | {item['insider']} ({item['role']}) \n"
                f"  └ 수량: {item['shares']:,}주 | 평단가: ${item['price']:.2f} | 총액: *{value_str}*"
            )
    else:
        report_lines.append("- 당일 $30,000 이상의 내부자 매매 내역 없음")
        
    report_text = "\n".join(report_lines)
    
    # Send to Telegram
    if TELEGRAM_BOT4_TOKEN and TELEGRAM_JJANG_GU_CHAT_ID:
        logger.info(f"Uploading US Market Report to Telegram chat: {TELEGRAM_JJANG_GU_CHAT_ID}...")
        res = send_telegram_message(TELEGRAM_BOT4_TOKEN, TELEGRAM_JJANG_GU_CHAT_ID, report_text)
        if res and res.get("ok"):
            logger.info("Successfully uploaded report to Telegram.")
        else:
            logger.error(f"Failed to upload report to Telegram: {res}")
    else:
        logger.error("Telegram credentials or Chat ID is missing in environment.")
        
    # Also output to stdout
    print("\n--- Generated Report ---")
    print(report_text)

if __name__ == "__main__":
    main()
